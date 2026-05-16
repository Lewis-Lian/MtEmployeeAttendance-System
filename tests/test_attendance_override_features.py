import io
import os
from pathlib import Path
import tempfile
import unittest
import urllib.parse
from types import SimpleNamespace

import openpyxl
from flask import Flask, g, render_template

from models import db
from models.account_set import AccountSet
from models.department import Department
from models.employee import Employee
from models.manager_month_stat import ManagerMonthStat
from models.user import EMPLOYEE_PAGE_PERMISSION_KEYS, HOME_PAGE_PERMISSION_KEYS, MANAGER_PAGE_PERMISSION_KEYS, User
from routes import register_routes
from utils.app_navigation import module_by_slug, nav_context, visible_modules


class AttendanceOverrideFeatureTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmpdir.name, "test.db")
        self.upload_dir = os.path.join(self.tmpdir.name, "uploads")

        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            __name__,
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        app.config.update(
            TESTING=True,
            SECRET_KEY="test-secret",
            SQLALCHEMY_DATABASE_URI=f"sqlite:///{self.db_path}",
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
            JWT_EXPIRES_HOURS=12,
            JWT_EXPIRES_DELTA=__import__("datetime").timedelta(hours=12),
            UPLOAD_FOLDER=self.upload_dir,
        )
        os.makedirs(self.upload_dir, exist_ok=True)

        db.init_app(app)
        register_routes(app)
        self.app = app

        with self.app.app_context():
            db.create_all()
            admin = User(username="admin", role="admin")
            admin.set_password("admin123")
            dept = Department(dept_no="D001", dept_name="行政部")
            employee = Employee(emp_no="E001", name="员工甲", dept_id=None, is_manager=False)
            manager = Employee(emp_no="M001", name="经理甲", dept_id=None, is_manager=True)
            db.session.add_all([admin, dept])
            db.session.flush()
            employee.dept_id = dept.id
            manager.dept_id = dept.id
            db.session.add_all([employee, manager])
            employee_b = Employee(emp_no="E002", name="员工乙", dept_id=dept.id, is_manager=False)
            db.session.add(employee_b)
            db.session.add(AccountSet(month="2026-05", name="2026-05", is_active=True, is_locked=False))
            db.session.commit()
            self.dept_id = dept.id
            self.employee_id = employee.id
            self.employee_b_id = employee_b.id
            self.manager_id = manager.id

        self.client = self.app.test_client()
        self.client.post("/login", data={"username": "admin", "password": "admin123"})

    def tearDown(self) -> None:
        with self.app.app_context():
            db.session.remove()
            db.drop_all()
        self.tmpdir.cleanup()

    def test_employee_manual_save_creates_history_once(self) -> None:
        payload = {
            "month": "2026-05",
            "emp_id": self.employee_id,
            "attendance_days": "3",
            "work_hours": "21.5",
            "half_days": "1",
            "late_early_minutes": "10",
            "remark": "手工修正",
        }

        first = self.client.put("/admin/employee-attendance-overrides/record", json=payload)
        self.assertEqual(first.status_code, 200)

        history = self.client.get(
            f"/admin/employee-attendance-overrides/history?emp_id={self.employee_id}&month=2026-05"
        )
        self.assertEqual(history.status_code, 200)
        first_rows = history.get_json()["rows"]
        self.assertEqual(len(first_rows), 1)
        self.assertEqual(first_rows[0]["action_type"], "manual_save")

        second = self.client.put("/admin/employee-attendance-overrides/record", json=payload)
        self.assertEqual(second.status_code, 200)
        history_again = self.client.get(
            f"/admin/employee-attendance-overrides/history?emp_id={self.employee_id}&month=2026-05"
        )
        self.assertEqual(len(history_again.get_json()["rows"]), 1)

    def test_employee_clear_creates_clear_history(self) -> None:
        self.client.put(
            "/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_id,
                "attendance_days": "2",
                "remark": "先保存",
            },
        )

        cleared = self.client.delete(
            f"/admin/employee-attendance-overrides/record?emp_id={self.employee_id}&month=2026-05"
        )
        self.assertEqual(cleared.status_code, 200)

        history = self.client.get(
            f"/admin/employee-attendance-overrides/history?emp_id={self.employee_id}&month=2026-05"
        )
        rows = history.get_json()["rows"]
        self.assertEqual(rows[0]["action_type"], "clear")
        self.assertEqual(rows[1]["action_type"], "manual_save")

    def test_employee_import_only_overrides_provided_values(self) -> None:
        self.client.put(
            "/admin/employee-attendance-overrides/record",
            json={
                "month": "2026-05",
                "emp_id": self.employee_id,
                "attendance_days": "2",
                "work_hours": "8",
                "remark": "原值",
            },
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工考勤修正"
        ws.append(
            [
                "月份",
                "工号",
                "姓名",
                "系统考勤天数",
                "系统工时",
                "系统半勤天数",
                "系统迟到早退",
                "考勤天数",
                "工时",
                "半勤天数",
                "迟到早退",
                "备注",
            ]
        )
        ws.append(["2026-05", "E001", "员工甲", "", "", "", "", "5", "", "", "", "导入修正"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        res = self.client.post(
            "/admin/employee-attendance-overrides/import",
            data={"month": "2026-05", "file": (buf, "employee-import.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(res.status_code, 200)
        data = res.get_json()
        self.assertEqual(data["changed_count"], 1)

        record = self.client.get(
            f"/admin/employee-attendance-overrides/record?emp_id={self.employee_id}&month=2026-05"
        ).get_json()
        self.assertEqual(record["override"]["attendance_days"], 5.0)
        self.assertEqual(record["override"]["work_hours"], 8.0)
        self.assertEqual(record["override"]["remark"], "导入修正")

        history = self.client.get(
            f"/admin/employee-attendance-overrides/history?emp_id={self.employee_id}&month=2026-05"
        ).get_json()["rows"]
        self.assertEqual(history[0]["action_type"], "import")

    def test_employee_history_lists_all_overrides_in_month(self) -> None:
        self.client.put(
            "/admin/employee-attendance-overrides/record",
            json={"month": "2026-05", "emp_id": self.employee_id, "attendance_days": "2", "remark": "员工甲修正"},
        )
        self.client.put(
            "/admin/employee-attendance-overrides/record",
            json={"month": "2026-05", "emp_id": self.employee_b_id, "attendance_days": "4", "remark": "员工乙修正"},
        )

        history = self.client.get("/admin/employee-attendance-overrides/history?month=2026-05")
        self.assertEqual(history.status_code, 200)
        rows = history.get_json()["rows"]
        names = {row["employee_name"] for row in rows}
        self.assertEqual(names, {"员工甲", "员工乙"})

    def test_attendance_override_pages_default_to_active_account_set_month(self) -> None:
        employee_res = self.client.get("/admin/employee-attendance-overrides")
        self.assertEqual(employee_res.status_code, 200)
        employee_html = employee_res.get_data(as_text=True)
        self.assertIn('id="employeeAttendanceOverrideMonth" value="2026-05"', employee_html)

        manager_res = self.client.get("/admin/manager-attendance-overrides")
        self.assertEqual(manager_res.status_code, 200)
        manager_html = manager_res.get_data(as_text=True)
        self.assertIn('id="managerAttendanceOverrideMonth" value="2026-05"', manager_html)

    def test_admin_route_extraction_modules_preserve_admin_contract(self) -> None:
        from routes.admin_accounts import register_admin_account_routes
        from routes.admin_attendance_overrides import register_admin_attendance_override_routes
        from routes.admin_imports import register_admin_import_routes

        self.assertTrue(callable(register_admin_account_routes))
        self.assertTrue(callable(register_admin_attendance_override_routes))
        self.assertTrue(callable(register_admin_import_routes))

        rules: dict[str, set[str]] = {}
        for rule in self.app.url_map.iter_rules():
            if not rule.rule.startswith("/admin/"):
                continue
            rules.setdefault(rule.rule, set()).update(rule.methods - {"HEAD", "OPTIONS"})

        self.assertEqual(rules["/admin/users"], {"GET", "POST"})
        self.assertEqual(rules["/admin/users/manager-batch"], {"POST"})
        self.assertEqual(rules["/admin/users/batch"], {"POST"})
        self.assertEqual(rules["/admin/users/<int:user_id>"], {"DELETE", "PUT"})
        self.assertEqual(rules["/admin/employee-attendance-overrides/record"], {"DELETE", "GET", "PUT"})
        self.assertEqual(rules["/admin/manager-attendance-overrides/import"], {"POST"})
        self.assertEqual(rules["/admin/departments/import"], {"POST"})
        self.assertEqual(rules["/admin/departments/template"], {"GET"})
        self.assertEqual(rules["/admin/employees/import"], {"POST"})

    def test_user_list_includes_employee_department_info(self) -> None:
        create_res = self.client.post(
            "/admin/users",
            json={
                "username": "M001",
                "password": "mt@123",
                "role": "readonly",
                "emp_ids": [self.manager_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(create_res.status_code, 200)

        list_res = self.client.get("/admin/users")
        self.assertEqual(list_res.status_code, 200)
        users = list_res.get_json()
        target = next((row for row in users if row["username"] == "M001"), None)

        self.assertIsNotNone(target)
        self.assertEqual(len(target["employees"]), 1)
        self.assertEqual(target["profile_emp_no"], "M001")
        self.assertEqual(target["profile_name"], "经理甲")
        self.assertEqual(target["profile_department"]["dept_name"], "行政部")
        self.assertEqual(target["employees"][0]["emp_no"], "M001")
        self.assertEqual(target["employees"][0]["name"], "经理甲")
        self.assertEqual(target["employees"][0]["dept_name"], "行政部")

    def test_manager_batch_create_uses_emp_no_and_self_scope(self) -> None:
        res = self.client.post("/admin/users/manager-batch")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["created_count"], 1)
        self.assertEqual(payload["skipped_count"], 0)
        self.assertEqual(payload["created_users"][0]["username"], "M001")

        with self.app.app_context():
            user = User.query.filter_by(username="M001").first()
            self.assertIsNotNone(user)
            self.assertEqual(user.role, "readonly")
            self.assertTrue(user.check_password("mt@123"))
            self.assertEqual(user.profile_emp_no, "M001")
            self.assertEqual(user.profile_name, "经理甲")
            self.assertEqual(user.profile_dept_id, self.dept_id)
            self.assertEqual([row.emp_id for row in user.employee_assignments], [self.manager_id])
            self.assertEqual([row.dept_id for row in user.department_assignments], [])

            permissions = user.effective_page_permissions()
            for key in MANAGER_PAGE_PERMISSION_KEYS:
                self.assertTrue(permissions[key])
            for key in EMPLOYEE_PAGE_PERMISSION_KEYS:
                self.assertFalse(permissions[key])

    def test_manager_batch_create_returns_skipped_employee_reasons(self) -> None:
        first_res = self.client.post("/admin/users/manager-batch")
        self.assertEqual(first_res.status_code, 200)

        second_res = self.client.post("/admin/users/manager-batch")
        self.assertEqual(second_res.status_code, 200)
        payload = second_res.get_json()

        self.assertEqual(payload["created_count"], 0)
        self.assertEqual(payload["skipped_count"], 1)
        self.assertEqual(payload["skipped_users"][0]["emp_no"], "M001")
        self.assertEqual(payload["skipped_users"][0]["name"], "经理甲")
        self.assertEqual(payload["skipped_users"][0]["reason"], "账号已存在")

    def test_user_batch_operations_support_reset_role_permissions_assignments_and_delete(self) -> None:
        create_a = self.client.post(
            "/admin/users",
            json={
                "username": "reader-a",
                "password": "old-a",
                "role": "readonly",
                "emp_ids": [self.employee_id],
                "dept_ids": [],
                "page_permissions": {key: key in MANAGER_PAGE_PERMISSION_KEYS for key in (*HOME_PAGE_PERMISSION_KEYS, *MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)},
            },
        )
        create_b = self.client.post(
            "/admin/users",
            json={
                "username": "reader-b",
                "password": "old-b",
                "role": "readonly",
                "emp_ids": [self.employee_b_id],
                "dept_ids": [],
                "page_permissions": {key: key in EMPLOYEE_PAGE_PERMISSION_KEYS for key in (*HOME_PAGE_PERMISSION_KEYS, *MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)},
            },
        )
        self.assertEqual(create_a.status_code, 200)
        self.assertEqual(create_b.status_code, 200)
        user_a_id = create_a.get_json()["user"]["id"]
        user_b_id = create_b.get_json()["user"]["id"]

        reset_res = self.client.post(
            "/admin/users/batch",
            json={"action": "reset_password", "user_ids": [user_a_id, user_b_id]},
        )
        self.assertEqual(reset_res.status_code, 200)

        with self.app.app_context():
            user_a = db.session.get(User, user_a_id)
            user_b = db.session.get(User, user_b_id)
            self.assertTrue(user_a.check_password("mt@123"))
            self.assertTrue(user_b.check_password("mt@123"))

        role_res = self.client.post(
            "/admin/users/batch",
            json={"action": "update_role", "user_ids": [user_a_id, user_b_id], "role": "admin"},
        )
        self.assertEqual(role_res.status_code, 200)

        with self.app.app_context():
            user_a = db.session.get(User, user_a_id)
            user_b = db.session.get(User, user_b_id)
            self.assertEqual(user_a.role, "admin")
            self.assertEqual(user_b.role, "admin")

        next_permissions = {key: key in MANAGER_PAGE_PERMISSION_KEYS for key in (*HOME_PAGE_PERMISSION_KEYS, *MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)}
        permissions_res = self.client.post(
            "/admin/users/batch",
            json={"action": "update_permissions", "user_ids": [user_a_id, user_b_id], "page_permissions": next_permissions},
        )
        self.assertEqual(permissions_res.status_code, 200)

        with self.app.app_context():
            user_a = db.session.get(User, user_a_id)
            user_b = db.session.get(User, user_b_id)
            self.assertEqual(user_a.effective_page_permissions(), {key: True for key in (*HOME_PAGE_PERMISSION_KEYS, *MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)})
            self.assertEqual(user_b.effective_page_permissions(), {key: True for key in (*HOME_PAGE_PERMISSION_KEYS, *MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)})

        downgrade_res = self.client.post(
            "/admin/users/batch",
            json={"action": "update_role", "user_ids": [user_a_id, user_b_id], "role": "readonly"},
        )
        self.assertEqual(downgrade_res.status_code, 200)

        permissions_res = self.client.post(
            "/admin/users/batch",
            json={"action": "update_permissions", "user_ids": [user_a_id, user_b_id], "page_permissions": next_permissions},
        )
        self.assertEqual(permissions_res.status_code, 200)

        with self.app.app_context():
            user_a = db.session.get(User, user_a_id)
            user_b = db.session.get(User, user_b_id)
            self.assertEqual(user_a.effective_page_permissions(), next_permissions)
            self.assertEqual(user_b.effective_page_permissions(), next_permissions)

        employee_batch_res = self.client.post(
            "/admin/users/batch",
            json={"action": "update_employees", "user_ids": [user_a_id, user_b_id], "emp_ids": [self.employee_id]},
        )
        self.assertEqual(employee_batch_res.status_code, 200)

        department_batch_res = self.client.post(
            "/admin/users/batch",
            json={"action": "update_departments", "user_ids": [user_a_id, user_b_id], "dept_ids": [self.dept_id]},
        )
        self.assertEqual(department_batch_res.status_code, 200)

        with self.app.app_context():
            user_a = db.session.get(User, user_a_id)
            user_b = db.session.get(User, user_b_id)
            self.assertEqual([row.emp_id for row in user_a.employee_assignments], [self.employee_id])
            self.assertEqual([row.emp_id for row in user_b.employee_assignments], [self.employee_id])
            self.assertEqual([row.dept_id for row in user_a.department_assignments], [self.dept_id])
            self.assertEqual([row.dept_id for row in user_b.department_assignments], [self.dept_id])

        delete_res = self.client.post(
            "/admin/users/batch",
            json={"action": "delete", "user_ids": [user_a_id, user_b_id]},
        )
        self.assertEqual(delete_res.status_code, 200)

        with self.app.app_context():
            self.assertIsNone(db.session.get(User, user_a_id))
            self.assertIsNone(db.session.get(User, user_b_id))

    def test_manager_batch_account_login_redirects_to_accessible_page(self) -> None:
        create_res = self.client.post("/admin/users/manager-batch")
        self.assertEqual(create_res.status_code, 200)

        manager_client = self.app.test_client()
        login_res = manager_client.post(
            "/login",
            data={"username": "M001", "password": "mt@123"},
            follow_redirects=False,
        )
        self.assertEqual(login_res.status_code, 302)
        self.assertTrue(login_res.headers["Location"].endswith("/employee/home"))

        root_res = manager_client.get("/", follow_redirects=False)
        self.assertEqual(root_res.status_code, 302)
        self.assertTrue(root_res.headers["Location"].endswith("/employee/home"))

    def test_account_create_and_update_allow_empty_employee_assignment(self) -> None:
        create_res = self.client.post(
            "/admin/users",
            json={
                "username": "no-emp",
                "password": "123456",
                "role": "readonly",
                "emp_ids": [],
                "dept_ids": [],
            },
        )
        self.assertEqual(create_res.status_code, 200)
        user_id = create_res.get_json()["user"]["id"]

        with self.app.app_context():
            user = db.session.get(User, user_id)
            self.assertEqual(user.employee_assignments, [])
            self.assertEqual(user.profile_emp_no or "", "")
            self.assertEqual(user.profile_name or "", "")

        create_ok = self.client.post(
            "/admin/users",
            json={
                "username": "with-emp",
                "password": "123456",
                "role": "readonly",
                "emp_ids": [self.employee_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(create_ok.status_code, 200)
        user_id = create_ok.get_json()["user"]["id"]

        update_res = self.client.put(
            f"/admin/users/{user_id}",
            json={
                "emp_ids": [],
                "dept_ids": [],
            },
        )
        self.assertEqual(update_res.status_code, 200)

        with self.app.app_context():
            user = db.session.get(User, user_id)
            self.assertEqual(user.employee_assignments, [])

    def test_top_nav_renders_username_and_employee_name(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "top_nav_user_display_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        nav_user = SimpleNamespace(
            username="reader",
            role="readonly",
            profile_emp_no="E001",
            profile_name="员工甲",
            employee_assignments=[SimpleNamespace(employee=SimpleNamespace(emp_no="E999", name="临时关联员工"))],
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )
        with app.test_request_context("/employee/dashboard"):
            g.current_user = nav_user
            html = render_template("dashboard.html", employees=[])

        self.assertIn("top-nav-user-code", html)
        self.assertIn("top-nav-user-person", html)
        self.assertIn(">E001<", html)
        self.assertIn(">员工甲<", html)

    def test_account_profile_identity_is_bound_on_create_only(self) -> None:
        create_res = self.client.post(
            "/admin/users",
            json={
                "username": "fixed-reader",
                "password": "123456",
                "role": "readonly",
                "emp_ids": [self.employee_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(create_res.status_code, 200)
        user_id = create_res.get_json()["user"]["id"]

        update_res = self.client.put(
            f"/admin/users/{user_id}",
            json={
                "emp_ids": [self.employee_b_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(update_res.status_code, 200)
        payload = update_res.get_json()["user"]
        self.assertEqual(payload["profile_emp_no"], "E001")
        self.assertEqual(payload["profile_name"], "员工甲")
        self.assertEqual(payload["profile_department"]["dept_name"], "行政部")
        self.assertEqual(payload["employees"][0]["emp_no"], "E002")
        self.assertEqual(payload["employees"][0]["name"], "员工乙")

        with self.app.app_context():
            user = db.session.get(User, user_id)
            self.assertEqual(user.profile_emp_no, "E001")
            self.assertEqual(user.profile_name, "员工甲")
            self.assertEqual(user.profile_dept_id, self.dept_id)

    def test_account_profile_identity_can_be_edited_manually(self) -> None:
        create_res = self.client.post(
            "/admin/users",
            json={
                "username": "editable-reader",
                "password": "123456",
                "role": "readonly",
                "emp_ids": [self.employee_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(create_res.status_code, 200)
        user_id = create_res.get_json()["user"]["id"]

        update_res = self.client.put(
            f"/admin/users/{user_id}",
            json={
                "profile_emp_no": "A1001",
                "profile_name": "测试账号",
                "profile_dept_id": self.dept_id,
                "emp_ids": [self.employee_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(update_res.status_code, 200)
        payload = update_res.get_json()["user"]
        self.assertEqual(payload["profile_emp_no"], "A1001")
        self.assertEqual(payload["profile_name"], "测试账号")
        self.assertEqual(payload["profile_department"]["dept_name"], "行政部")

        with self.app.app_context():
            user = db.session.get(User, user_id)
            self.assertEqual(user.profile_emp_no, "A1001")
            self.assertEqual(user.profile_name, "测试账号")
            self.assertEqual(user.profile_dept_id, self.dept_id)

    def test_account_profile_department_can_be_edited_by_single_selector_value(self) -> None:
        with self.app.app_context():
            dept_b = Department(dept_no="D002", dept_name="生产部")
            db.session.add(dept_b)
            db.session.commit()
            dept_b_id = dept_b.id

        create_res = self.client.post(
            "/admin/users",
            json={
                "username": "dept-reader",
                "password": "123456",
                "role": "readonly",
                "emp_ids": [self.employee_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(create_res.status_code, 200)
        user_id = create_res.get_json()["user"]["id"]

        update_res = self.client.put(
            f"/admin/users/{user_id}",
            json={
                "profile_emp_no": "E001",
                "profile_name": "员工甲",
                "profile_dept_id": dept_b_id,
                "emp_ids": [self.employee_id],
                "dept_ids": [],
            },
        )
        self.assertEqual(update_res.status_code, 200)
        payload = update_res.get_json()["user"]
        self.assertEqual(payload["profile_department"]["id"], dept_b_id)
        self.assertEqual(payload["profile_department"]["dept_name"], "生产部")

    def test_admin_module_exports_symbols_needed_by_extracted_routes(self) -> None:
        from routes import admin as admin_module

        self.assertIsInstance(admin_module.MANAGER_PAGE_PERMISSION_KEYS, (list, tuple))
        self.assertIsInstance(admin_module.EMPLOYEE_PAGE_PERMISSION_KEYS, (list, tuple))
        self.assertIsInstance(admin_module.PAGE_PERMISSION_LABELS, dict)
        self.assertTrue(callable(admin_module.parse_bool_zh))
        self.assertTrue(admin_module.parse_bool_zh("是"))

    def test_runtime_code_no_longer_uses_legacy_query_lookup_helpers(self) -> None:
        project_root = Path(os.path.dirname(os.path.dirname(__file__)))
        runtime_files = [
            project_root / "routes" / "auth.py",
            project_root / "routes" / "admin.py",
            project_root / "routes" / "admin_accounts.py",
            project_root / "routes" / "admin_attendance_overrides.py",
            project_root / "routes" / "admin_imports.py",
            project_root / "routes" / "employee.py",
            project_root / "services" / "attendance_service.py",
            project_root / "services" / "manager_attendance_service.py",
        ]

        offenders = []
        for file_path in runtime_files:
            content = file_path.read_text(encoding="utf-8")
            if ".query.get(" in content or ".get_or_404(" in content:
                offenders.append(str(file_path.relative_to(project_root)))

        self.assertEqual(offenders, [])

    def test_manager_example_download_returns_expected_headers(self) -> None:
        res = self.client.get("/admin/manager-attendance-overrides/template?month=2026-05")
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("月份", headers)
        self.assertIn("工号", headers)
        self.assertIn("出勤天数", headers)
        self.assertIn("备注", headers)

    def test_manager_overtime_export_download_returns_workbook(self) -> None:
        res = self.client.get("/admin/manager-overtime/export?year=2026")
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("前年累积天数", headers)
        self.assertIn("剩余调休天数", headers)

    def test_manager_annual_leave_export_download_returns_workbook(self) -> None:
        res = self.client.get("/admin/manager-annual-leave/export?year=2026")
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("1月", headers)
        self.assertIn("剩余年休天数", headers)

    def test_manager_overtime_import_defaults_year_when_omitted(self) -> None:
        res = self.client.post("/admin/manager-overtime/import")
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.get_json()["error"], "请选择要导入的Excel文件")

    def test_manager_overtime_import_checks_locked_prev_dec_when_year_omitted(self) -> None:
        current_year = __import__("datetime").datetime.now().year
        with self.app.app_context():
            db.session.add(
                AccountSet(
                    month=f"{current_year - 1}-12",
                    name=f"{current_year - 1}-12",
                    is_active=False,
                    is_locked=True,
                )
            )
            db.session.commit()

        res = self.client.post("/admin/manager-overtime/import")
        self.assertEqual(res.status_code, 400)
        self.assertIn(f"{current_year - 1}-12 账套已锁定，不能导入管理人员加班统计", res.get_json()["error"])

    def test_employee_import_preserves_boolean_parsing_through_admin_import_routes(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工导入模板"
        ws.append(
            [
                "人员编号",
                "人员姓名",
                "部门名称",
                "班次编号",
                "是否管理人员",
                "是否哺乳假",
                "员工考勤统计来源",
                "管理人员考勤统计来源",
            ]
        )
        ws.append(["E003", "员工丙", "行政部", "", "是", "是", "员工考勤源文件取值", "管理人员考勤源文件取值"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        res = self.client.post(
            "/admin/employees/import",
            data={"file": (buf, "employees-import.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(res.status_code, 200)

        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E003").first()
            self.assertIsNotNone(employee)
            self.assertTrue(employee.is_manager)
            self.assertTrue(employee.is_nursing)

    def test_employee_batch_update_keeps_employee_list_unique(self) -> None:
        before_rows = self.client.get("/admin/employees").get_json()
        before_ids = [row["id"] for row in before_rows]

        target_ids = [self.employee_id, self.employee_b_id]
        update_res = self.client.post(
            "/admin/employees/batch",
            json={
                "action": "set_employee_stats_attendance_source",
                "ids": target_ids,
                "employee_stats_attendance_source": "manager",
            },
        )
        self.assertEqual(update_res.status_code, 200)

        after_rows = self.client.get("/admin/employees").get_json()
        after_ids = [row["id"] for row in after_rows]
        self.assertEqual(len(after_ids), len(set(after_ids)))
        self.assertEqual(set(after_ids), set(before_ids))

        updated_rows = {row["id"]: row for row in after_rows if row["id"] in target_ids}
        self.assertEqual(updated_rows[self.employee_id]["employee_stats_attendance_source"], "manager")
        self.assertEqual(updated_rows[self.employee_b_id]["employee_stats_attendance_source"], "manager")

    def test_unique_employees_helper_drops_duplicate_employee_objects(self) -> None:
        from routes import admin as admin_module

        with self.app.app_context():
            employee = db.session.get(Employee, self.employee_id)
            manager = db.session.get(Employee, self.manager_id)

        unique_rows = admin_module._unique_employees([employee, employee, manager, manager])
        self.assertEqual([row.id for row in unique_rows], [self.employee_id, self.manager_id])

    def test_manager_annual_leave_import_defaults_year_when_omitted(self) -> None:
        res = self.client.post("/admin/manager-annual-leave/import")
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.get_json()["error"], "请选择要导入的Excel文件")

    def test_manager_annual_leave_import_checks_locked_months_when_year_omitted(self) -> None:
        current_year = __import__("datetime").datetime.now().year
        with self.app.app_context():
            db.session.add(
                AccountSet(
                    month=f"{current_year}-01",
                    name=f"{current_year}-01",
                    is_active=False,
                    is_locked=True,
                )
            )
            db.session.commit()

        res = self.client.post("/admin/manager-annual-leave/import")
        self.assertEqual(res.status_code, 400)
        self.assertIn(f"{current_year}-01 账套已锁定，不能导入管理人员年休统计", res.get_json()["error"])

    def test_departments_template_download_uses_importable_headers(self) -> None:
        res = self.client.get("/admin/departments/template")
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[:3], ["部门编号", "部门名称", "上级部门编号"])
        self.assertEqual(headers[3], "原始部门ID")
        self.assertEqual(ws.max_column, 4)
        self.assertTrue(ws.column_dimensions["D"].hidden)
        self.assertNotIn("部门导入元数据", wb.sheetnames)

    def test_departments_page_renders_with_export_link(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "departments_template_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        admin_user = SimpleNamespace(
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )
        with app.test_request_context("/admin/departments/manage"):
            g.current_user = admin_user
            html = render_template("admin/departments.html")

        self.assertIn("/admin/departments/export", html)

    def test_accounts_page_renders_filter_controls(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "accounts_filter_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        admin_user = SimpleNamespace(
            id=1,
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )
        with app.test_request_context("/admin/accounts"):
            g.current_user = admin_user
            html = render_template(
                "admin/accounts.html",
                current_user_id=admin_user.id,
                manager_page_permissions=[],
                employee_page_permissions=[],
            )

        self.assertIn('class="query-page-shell"', html)
        self.assertIn("创建账号", html)
        self.assertIn("账号列表", html)
        self.assertIn('id="createPermissionInput"', html)
        self.assertIn('id="editPermissionInput"', html)
        self.assertIn('id="accountPermissionModal"', html)
        self.assertIn('id="filterEmpLookup"', html)
        self.assertIn("账号筛选", html)
        self.assertIn("姓名", html)
        self.assertIn('name="profile_emp_no"', html)
        self.assertIn('name="profile_name"', html)
        self.assertIn('id="editProfileDeptLookup"', html)
        self.assertIn('name="profile_dept_id"', html)
        self.assertIn('id="accountSingleDeptPickerModal"', html)
        self.assertIn('id="toggleSelectAllUsers"', html)
        self.assertIn('id="batchResetPasswordBtn"', html)
        self.assertIn('id="batchDeleteUsersBtn"', html)
        self.assertIn('id="applyBatchRoleBtn"', html)
        self.assertIn('id="batchRoleModal"', html)
        self.assertIn('id="batchRoleSelect"', html)
        self.assertIn('id="confirmBatchRoleBtn"', html)
        self.assertIn('id="openBatchEmployeeBtn"', html)
        self.assertIn('id="batchEmployeeModal"', html)
        self.assertIn('id="confirmBatchEmployeeBtn"', html)
        self.assertIn('id="openBatchDepartmentBtn"', html)
        self.assertIn('id="batchDepartmentModal"', html)
        self.assertIn('id="confirmBatchDepartmentBtn"', html)
        self.assertIn('id="openBatchPermissionBtn"', html)
        self.assertIn('id="filterAdminRole"', html)
        self.assertIn('id="applyUserFiltersBtn"', html)
        self.assertIn('id="resetUserFiltersBtn"', html)

    def test_authenticated_shell_renders_enterprise_navigation(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "enterprise_shell_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        admin_user = SimpleNamespace(
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )
        with app.test_request_context("/admin/dashboard"):
            g.current_user = admin_user
            html = render_template("admin/dashboard.html")

        self.assertIn("企业考勤处理中心", html)
        self.assertIn("app-top-modules", html)
        self.assertIn("app-top-module", html)
        self.assertIn("app-module-sidebar", html)
        self.assertIn("module-bottom-nav", html)
        self.assertIn("module-bottom-link", html)
        self.assertIn("/module/query", html)
        self.assertIn("/module/account", html)
        self.assertIn("查询中心", html)
        self.assertIn("账套中心", html)
        self.assertIn("主数据", html)
        self.assertIn("修正中心", html)
        self.assertIn("系统设置", html)

    def test_authenticated_shell_hides_restricted_modules_for_readonly_user(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "enterprise_shell_readonly_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        readonly_user = SimpleNamespace(
            username="reader",
            role="readonly",
            has_any_page_access=lambda keys: "employee_dashboard" in keys,
            can_access_page=lambda key: key in ("employee_dashboard", "query_home"),
        )
        with app.test_request_context("/employee/dashboard"):
            g.current_user = readonly_user
            html = render_template("dashboard.html", employees=[])

        self.assertIn("/module/query", html)
        self.assertIn("/employee/dashboard", html)
        self.assertNotIn("/module/account", html)
        self.assertNotIn("/admin/dashboard", html)
        self.assertNotIn("/module/master-data", html)
        self.assertNotIn("/module/corrections", html)
        self.assertNotIn("/module/settings", html)

    def test_authenticated_shell_renders_icon_class_for_every_sidebar_entry(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "enterprise_shell_icon_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        admin_user = SimpleNamespace(
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )

        expected_entry_keys = {
            "query_home",
            "employee_dashboard",
            "abnormal_query",
            "punch_records",
            "department_hours_query",
            "summary_download",
            "manager_query",
            "manager_overtime_query",
            "manager_annual_leave_query",
            "manager_department_hours_query",
            "account_dashboard",
            "employees",
            "departments",
            "shifts",
            "employee_attendance_overrides",
            "manager_attendance_overrides",
            "manager_overtime",
            "manager_annual_leave",
            "accounts",
        }
        seen_keys = set()
        modules = visible_modules(admin_user)

        for module in modules:
            request_path = module["entries"][0]["href"]
            with app.test_request_context(request_path):
                g.current_user = admin_user
                html = render_template("partials/app_nav.html", app_nav=nav_context(admin_user, request_path))

            expected_module_icon_class = f"icon-{module.get('icon_key', '').replace('_', '-')}"
            if module.get("icon_key"):
                with self.subTest(module=module["slug"], module_icon=module["icon_key"]):
                    self.assertIn(expected_module_icon_class, html)

            for entry in module["entries"]:
                seen_keys.add(entry["key"])
                expected_icon_class = f"icon-{entry['key'].replace('_', '-')}"
                with self.subTest(module=module["slug"], entry_key=entry["key"]):
                    self.assertIn(expected_icon_class, html)

        self.assertEqual(seen_keys, expected_entry_keys)

    def test_login_page_renders_enterprise_entry_copy(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "enterprise_login_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        with app.test_request_context("/login"):
            html = render_template("login.html", error=None)

        self.assertIn("企业考勤处理中心", html)
        self.assertIn("登录考勤系统", html)
        self.assertIn("login-surface", html)
        self.assertIn("login-brand-panel", html)
        self.assertIn("login-brand-mark", html)
        self.assertIn("login-brand-slider", html)
        self.assertIn("login-backdrop-curve", html)
        self.assertIn("记住我", html)
        self.assertIn("修改密码", html)
        self.assertIn("auth-slider-track", html)

    def test_change_password_page_renders_login_style_layout(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "enterprise_change_password_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        with app.test_request_context("/change-password"):
            html = render_template("change_password.html", error=None, success=None)
            success_html = render_template("change_password.html", error=None, success="密码修改成功")

        self.assertIn("修改密码", html)
        self.assertIn("auth-slider-track", html)
        self.assertIn("返回登录", html)
        self.assertIn("data-success-redirect", success_html)
        self.assertIn("2</span> 秒后自动返回登录页", success_html)

    def test_login_with_remember_me_sets_persistent_cookie(self) -> None:
        response = self.client.post(
            "/login",
            data={"username": "admin", "password": "admin123", "remember_me": "1"},
        )

        self.assertEqual(response.status_code, 302)
        cookie_header = response.headers.get("Set-Cookie", "")
        self.assertIn("Max-Age=2592000", cookie_header)

    def test_change_password_rejects_wrong_current_password(self) -> None:
        response = self.client.post(
            "/change-password",
            data={
                "username": "admin",
                "current_password": "wrong-password",
                "new_password": "new-admin-123",
                "confirm_password": "new-admin-123",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("用户不存在或原密码错误", response.get_data(as_text=True))

    def test_change_password_updates_password_and_supports_new_login(self) -> None:
        response = self.client.post(
            "/change-password",
            data={
                "username": "admin",
                "current_password": "admin123",
                "new_password": "new-admin-123",
                "confirm_password": "new-admin-123",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("密码修改成功", response.get_data(as_text=True))

        failed_login = self.app.test_client().post(
            "/login",
            data={"username": "admin", "password": "admin123"},
        )
        self.assertEqual(failed_login.status_code, 401)

        success_login = self.app.test_client().post(
            "/login",
            data={"username": "admin", "password": "new-admin-123"},
        )
        self.assertEqual(success_login.status_code, 302)

    def test_representative_pages_render_workflow_classes(self) -> None:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            "enterprise_workflow_render_test",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        admin_user = SimpleNamespace(
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )
        with app.test_request_context("/admin/dashboard"):
            g.current_user = admin_user
            admin_html = render_template("admin/dashboard.html")

        with app.test_request_context("/employee/dashboard"):
            g.current_user = admin_user
            query_html = render_template("dashboard.html", employees=[])

        self.assertIn("account-workflow", admin_html)
        self.assertIn("account-status-card", admin_html)
        self.assertIn("account-audit-card", admin_html)
        self.assertIn("query-page-shell", query_html)
        self.assertIn("query-filter-rail", query_html)
        self.assertIn("query-workspace", query_html)
        self.assertIn("query-metric-grid", query_html)
        self.assertIn("query-result-panel", query_html)
        self.assertIn("employeePickerModal", query_html)
        self.assertIn("static/js/dashboard.js", query_html)

    def render_query_template(self, path: str, template_name: str, **context: object) -> str:
        project_root = os.path.dirname(os.path.dirname(__file__))
        app = Flask(
            f"query_center_render_{template_name.replace('/', '_')}",
            template_folder=os.path.join(project_root, "templates"),
            static_folder=os.path.join(project_root, "static"),
        )
        register_routes(app)

        admin_user = SimpleNamespace(
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )
        with app.test_request_context(path):
            g.current_user = admin_user
            return render_template(template_name, **context)

    def test_query_center_table_pages_render_query_workspace(self) -> None:
        pages = [
            (
                "/employee/home",
                "employee_home.html",
                {},
                ["managerHomeAccountSetSelect", "managerHomeSummary", "managerHomeEmptyState", "static/js/employee_home.js"],
            ),
            (
                "/employee/dashboard",
                "dashboard.html",
                {"employees": []},
                ["selectedEmpIds", "accountSetSelect", "refreshBtn", "downloadBtn", "finalDataTable", "static/js/dashboard.js"],
            ),
            (
                "/employee/abnormal-query",
                "abnormal_query.html",
                {"employees": []},
                ["selectedEmpIds", "accountSetSelect", "queryBtn", "downloadBtn", "abnormalTableBody", "static/js/abnormal_query.js"],
            ),
            (
                "/employee/punch-records",
                "punch_records.html",
                {"employees": []},
                ["selectedEmpIds", "accountSetSelect", "queryBtn", "downloadBtn", "punchTableBody", "static/js/punch_records.js"],
            ),
            (
                "/employee/department-hours-query",
                "department_hours_query.html",
                {},
                ["accountSetSelect", "queryBtn", "downloadBtn", "departmentHoursBody", "static/js/department_hours_query.js"],
            ),
            (
                "/employee/manager-query",
                "manager_query.html",
                {"employees": []},
                ["selectedEmpIds", "managerAccountSetSelect", "managerQueryBtn", "managerDownloadBtn", "managerQueryBody", "static/js/manager_query.js"],
            ),
            (
                "/employee/manager-overtime-query",
                "manager_overtime_query.html",
                {"employees": []},
                ["selectedEmpIds", "managerOvertimeQueryYear", "managerOvertimeQueryBtn", "managerOvertimeQueryBody", "static/js/manager_overtime_query.js"],
            ),
            (
                "/employee/manager-annual-leave-query",
                "manager_annual_leave_query.html",
                {"employees": []},
                ["selectedEmpIds", "managerAnnualLeaveQueryYear", "managerAnnualLeaveQueryBtn", "managerAnnualLeaveQueryBody", "static/js/manager_annual_leave_query.js"],
            ),
            (
                "/employee/manager-department-hours-query",
                "manager_department_hours_query.html",
                {},
                ["accountSetSelect", "queryBtn", "downloadBtn", "managerDepartmentHoursBody", "static/js/manager_department_hours_query.js"],
            ),
        ]

        for path, template_name, context, expected_fragments in pages:
            with self.subTest(template=template_name):
                html = self.render_query_template(path, template_name, **context)
                self.assertIn("query-page-shell", html)
                self.assertIn("query-filter-rail", html)
                self.assertIn("query-workspace", html)
                self.assertIn("query-result-panel", html)
                self.assertNotIn("query-page-heading", html)
                self.assertNotIn("Query Center", html)
                for fragment in expected_fragments:
                    self.assertIn(fragment, html)

    def test_summary_download_renders_download_task_layout(self) -> None:
        html = self.render_query_template("/employee/summary-download", "summary_download.html", employees=[])

        self.assertIn("download-page-shell", html)
        self.assertIn("download-task-panel", html)
        self.assertIn("download-report-grid", html)
        self.assertIn("download-header-panel", html)
        self.assertIn("download-help-panel", html)
        self.assertIn("employeePickerModal", html)
        self.assertIn("accountSetSelect", html)
        self.assertIn("selectedEmpIds", html)
        self.assertIn("includeFinalData", html)
        self.assertIn("includePunchRecords", html)
        self.assertIn("downloadBtn", html)
        self.assertIn("finalHeaderCheckboxes", html)
        self.assertIn("punchHeaderCheckboxes", html)
        self.assertIn("static/js/summary_download.js", html)

    def test_employee_query_pages_default_to_all_accessible_employees_when_selector_empty(self) -> None:
        project_root = Path(__file__).resolve().parent.parent
        js_files = [
            project_root / "static/js/dashboard.js",
            project_root / "static/js/abnormal_query.js",
            project_root / "static/js/punch_records.js",
            project_root / "static/js/summary_download.js",
        ]

        for path in js_files:
            with self.subTest(script=path.name):
                source = path.read_text(encoding="utf-8")
                self.assertIn("未手动选择时，默认查询当前账号下全部可见员工", source)
                self.assertNotIn("请先选择员工", source)

    def test_accounts_filter_uses_account_profile_name_and_department_fields(self) -> None:
        project_root = Path(__file__).resolve().parent.parent
        source = (project_root / "static/js/accounts.js").read_text(encoding="utf-8")
        self.assertIn("user.profile_name", source)
        self.assertIn("user.profile_dept_id", source)
        self.assertNotIn("const userEmpIds = Array.isArray(user.emp_ids) ? user.emp_ids : [];", source)

    def test_global_metric_sections_are_toggle_ready(self) -> None:
        project_root = Path(__file__).resolve().parent.parent
        base_html = (project_root / "templates/base.html").read_text(encoding="utf-8")
        toggle_js = (project_root / "static/js/metric_section_toggle.js").read_text(encoding="utf-8")
        style_css = (project_root / "static/css/style.css").read_text(encoding="utf-8")

        self.assertIn("js/metric_section_toggle.js", base_html)
        self.assertIn('collect(".query-metric-grid")', toggle_js)
        self.assertNotIn('collect(".manager-home-metric-grid")', toggle_js)
        self.assertNotIn('collect(".module-summary-grid")', toggle_js)
        self.assertIn("summary-card.dashboard-metric-card", toggle_js)
        self.assertIn('.top-nav-actions', toggle_js)
        self.assertIn(".top-nav-user", toggle_js)
        self.assertIn("展开卡片", toggle_js)
        self.assertIn("收起卡片", toggle_js)
        self.assertIn(".top-nav-metric-toggle", style_css)
        self.assertIn(".metric-toggle-target.is-collapsed", style_css)

    def test_global_tables_support_sticky_header_and_sorting(self) -> None:
        project_root = Path(__file__).resolve().parent.parent
        pager_js = (project_root / "static/js/table_pager.js").read_text(encoding="utf-8")
        style_css = (project_root / "static/css/style.css").read_text(encoding="utf-8")

        self.assertIn("bindSort(table, tbody, state)", pager_js)
        self.assertIn("table-sortable", pager_js)
        self.assertIn("data-sort-direction", pager_js)
        self.assertIn("position: sticky;", style_css)
        self.assertIn(".table-sortable", style_css)
        self.assertIn('.table thead th[data-sort-direction="asc"]::after', style_css)
        self.assertIn('.table thead th[data-sort-direction="desc"]::after', style_css)

    def test_product_navigation_groups_pages_into_modules(self) -> None:
        admin_user = SimpleNamespace(
            username="admin",
            role="admin",
            has_any_page_access=lambda _keys: True,
            can_access_page=lambda _key: True,
        )

        modules = visible_modules(admin_user)
        slugs = [module["slug"] for module in modules]
        self.assertEqual(slugs, ["home", "query", "account", "master-data", "corrections", "settings"])

        home = module_by_slug("home")
        self.assertIsNotNone(home)
        self.assertEqual(home["label"], "首页")
        self.assertEqual([entry["label"] for entry in home["entries"]], ["首页"])

        query = module_by_slug("query")
        self.assertIsNotNone(query)
        self.assertEqual(query["label"], "查询中心")
        self.assertIn("/employee/dashboard", [entry["href"] for entry in query["entries"]])
        self.assertEqual(
            [entry["label"] for entry in query["entries"]],
            [
                "员工考勤数据查询",
                "员工异常查询",
                "员工打卡数据查询",
                "员工部门工时",
                "管理人员考勤数据查询",
                "管理人员加班查询",
                "管理人员年休查询",
                "管理人员部门工时",
                "汇总下载",
            ],
        )

        context = nav_context(admin_user, "/admin/departments/manage")
        self.assertEqual(context["current_module"]["slug"], "master-data")
        self.assertIn("/admin/departments/manage", [entry["href"] for entry in context["current_entries"]])

    def test_product_navigation_filters_readonly_permissions(self) -> None:
        readonly_user = SimpleNamespace(
            username="reader",
            role="readonly",
            has_any_page_access=lambda keys: "employee_dashboard" in keys,
            can_access_page=lambda key: key in ("employee_dashboard", "query_home"),
        )

        modules = visible_modules(readonly_user)
        self.assertEqual([module["slug"] for module in modules], ["home", "query"])

        context = nav_context(readonly_user, "/employee/home")
        self.assertEqual(context["current_module"]["slug"], "home")
        self.assertEqual([entry["href"] for entry in context["current_entries"]], ["/employee/dashboard"])

    def test_module_home_routes_render_accessible_entries(self) -> None:
        home_module_res = self.client.get("/module/home", follow_redirects=False)
        self.assertEqual(home_module_res.status_code, 302)
        self.assertTrue(home_module_res.headers["Location"].endswith("/employee/home"))

        res = self.client.get("/module/query")
        self.assertEqual(res.status_code, 200)
        query_html = res.get_data(as_text=True)
        self.assertIn("查询中心", query_html)
        self.assertIn("/employee/dashboard", query_html)
        self.assertIn("/employee/manager-query", query_html)

        home_res = self.client.get("/employee/home")
        self.assertEqual(home_res.status_code, 200)
        home_html = home_res.get_data(as_text=True)
        self.assertIn("manager-home-shell", home_html)
        self.assertIn("managerHomeAccountSetSelect", home_html)

        account_res = self.client.get("/module/account")
        self.assertEqual(account_res.status_code, 200)
        self.assertIn("/admin/dashboard", account_res.get_data(as_text=True))

    def test_module_home_rejects_inaccessible_module(self) -> None:
        with self.app.app_context():
            reader = User(username="reader", role="readonly", page_permissions={"employee_dashboard": True})
            reader.set_password("reader123")
            db.session.add(reader)
            db.session.commit()

        reader_client = self.app.test_client()
        reader_client.post("/login", data={"username": "reader", "password": "reader123"})

        allowed = reader_client.get("/module/query")
        self.assertEqual(allowed.status_code, 200)
        self.assertIn("/employee/dashboard", allowed.get_data(as_text=True))

        home = reader_client.get("/employee/home")
        self.assertEqual(home.status_code, 200)
        home_html = home.get_data(as_text=True)
        self.assertIn("manager-home-shell", home_html)
        self.assertNotIn("/employee/abnormal-query", home_html)

        denied = reader_client.get("/module/account")
        self.assertEqual(denied.status_code, 403)

    def test_employee_home_api_returns_empty_when_no_account_set(self) -> None:
        with self.app.app_context():
            AccountSet.query.delete()
            db.session.commit()

        res = self.client.get("/employee/api/home-manager-summary")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertFalse(payload["has_data"])
        self.assertEqual(payload["empty_state"], "暂无账套，暂无数据")

    def test_employee_home_api_returns_empty_when_profile_emp_no_missing(self) -> None:
        with self.app.app_context():
            viewer = User(username="viewer", role="readonly", page_permissions={"employee_dashboard": True})
            viewer.profile_name = "查看者"
            viewer.set_password("viewer123")
            db.session.add(viewer)
            db.session.commit()

        viewer_client = self.app.test_client()
        viewer_client.post("/login", data={"username": "viewer", "password": "viewer123"})
        res = viewer_client.get("/employee/api/home-manager-summary")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertFalse(payload["has_data"])
        self.assertEqual(payload["empty_state"], "账号未绑定管理人员工号，暂无数据")

    def test_employee_home_api_returns_manager_summary_by_profile_emp_no(self) -> None:
        with self.app.app_context():
            viewer = User(username="viewer", role="readonly", page_permissions={"manager_query": True})
            viewer.profile_emp_no = "M001"
            viewer.profile_name = "经理甲账号"
            viewer.set_password("viewer123")
            db.session.add(viewer)
            db.session.add(AccountSet(month="2026-04", name="2026-04", is_active=False, is_locked=False))
            db.session.add(AccountSet(month="2026-06", name="2026-06", is_active=False, is_locked=False))
            db.session.add(ManagerMonthStat(emp_id=self.manager_id, year=2026, stat_type="annual_leave", m4=1, m5=2, remaining=9))
            db.session.add(ManagerMonthStat(emp_id=self.manager_id, year=2026, stat_type="overtime", prev_dec=2, m4=1, m5=-0.5, m6=3, remaining=5.5))
            db.session.commit()

        viewer_client = self.app.test_client()
        viewer_client.post("/login", data={"username": "viewer", "password": "viewer123"})
        april_res = viewer_client.get("/employee/api/home-manager-summary?month=2026-04")
        self.assertEqual(april_res.status_code, 200)
        april_payload = april_res.get_json()
        self.assertTrue(april_payload["has_data"])
        self.assertEqual(april_payload["month"], "2026-04")
        self.assertEqual(april_payload["manager"]["emp_no"], "M001")
        self.assertEqual(april_payload["manager"]["name"], "经理甲")
        self.assertEqual(april_payload["manager"]["dept_name"], "行政部")
        self.assertIn("attendance_days", april_payload["summary"])
        self.assertIn("late_early_minutes", april_payload["summary"])
        self.assertIn("benefit_days", april_payload["summary"])
        self.assertEqual(april_payload["summary"]["benefit_days"], 11)
        self.assertEqual(april_payload["summary"]["overtime_remaining_days"], 3)

        payload = viewer_client.get("/employee/api/home-manager-summary?month=2026-05").get_json()
        self.assertEqual(payload["summary"]["benefit_days"], 9)
        self.assertEqual(payload["summary"]["overtime_remaining_days"], 2.5)
        self.assertIn("overtime_remaining_days", payload["summary"])

    def test_departments_export_downloads_importable_rows(self) -> None:
        with self.app.app_context():
            parent = Department.query.filter_by(dept_no="D001").first()
            child = Department(dept_no="D010", dept_name="行政一部", parent_id=parent.id)
            db.session.add(child)
            db.session.commit()

        res = self.client.get("/admin/departments/export")
        self.assertEqual(res.status_code, 200)
        content_disposition = res.headers.get("Content-Disposition", "")
        self.assertIn("attachment", content_disposition)
        filename = ""
        for part in [segment.strip() for segment in content_disposition.split(";")]:
            if part.startswith("filename*="):
                value = part.split("=", 1)[1].strip().strip('"')
                if value.lower().startswith("utf-8''"):
                    value = value[7:]
                filename = urllib.parse.unquote(value)
                break
            if part.startswith("filename="):
                filename = part.split("=", 1)[1].strip().strip('"')
        self.assertEqual(filename, "部门导出.xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        normalized_rows = [
            (row[0], row[1], row[2] or "")
            for row in rows[1:]
            if row[0] and row[1]
        ]
        self.assertEqual(rows[0][:3], ("部门编号", "部门名称", "上级部门编号"))
        self.assertEqual(rows[0][3], "原始部门ID")
        self.assertEqual(ws.max_column, 4)
        self.assertTrue(ws.column_dimensions["D"].hidden)
        self.assertIn("部门导入元数据", wb.sheetnames)
        self.assertIn(("D001", "行政部", ""), normalized_rows)
        self.assertIn(("D010", "行政一部", "D001"), normalized_rows)
        exported_ids = {
            str(row[3])
            for row in rows[1:]
            if row[0] and row[1] and row[3] is not None
        }
        self.assertEqual(len(exported_ids), 2)

    def test_departments_import_updates_existing_department_when_exported_dept_no_changes(self) -> None:
        with self.app.app_context():
            parent = Department.query.filter_by(dept_no="D001").first()
            child = Department(dept_no="D010", dept_name="行政一部", parent_id=parent.id)
            db.session.add(child)
            db.session.commit()

        export_res = self.client.get("/admin/departments/export")
        self.assertEqual(export_res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(export_res.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        dept_no_idx = headers.index("部门编号") + 1
        dept_name_idx = headers.index("部门名称") + 1
        parent_no_idx = headers.index("上级部门编号") + 1

        for row_idx in range(2, ws.max_row + 1):
            dept_no = ws.cell(row=row_idx, column=dept_no_idx).value
            dept_name = ws.cell(row=row_idx, column=dept_name_idx).value
            if dept_no == "D001" and dept_name == "行政部":
                ws.cell(row=row_idx, column=dept_no_idx, value="D099")
                continue
            if dept_no == "D010" and dept_name == "行政一部":
                ws.cell(row=row_idx, column=parent_no_idx, value="D099")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import_res = self.client.post(
            "/admin/departments/import",
            data={"file": (buf, "departments-renumbered.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_res.status_code, 200)

        with self.app.app_context():
            departments = Department.query.order_by(Department.dept_no.asc()).all()
            self.assertEqual(len(departments), 2)
            self.assertIsNone(Department.query.filter_by(dept_no="D001").first())

            renamed_parent = Department.query.filter_by(dept_no="D099").first()
            self.assertIsNotNone(renamed_parent)
            self.assertEqual(renamed_parent.dept_name, "行政部")

            child = Department.query.filter_by(dept_no="D010").first()
            self.assertIsNotNone(child)
            self.assertEqual(child.parent_id, renamed_parent.id)

    def test_departments_import_keeps_identity_when_rows_are_reordered(self) -> None:
        with self.app.app_context():
            parent = Department.query.filter_by(dept_no="D001").first()
            child = Department(dept_no="D010", dept_name="行政一部", parent_id=parent.id)
            db.session.add(child)
            db.session.commit()

        export_res = self.client.get("/admin/departments/export")
        self.assertEqual(export_res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(export_res.data))
        ws = wb.active
        row_two = [ws.cell(row=2, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        row_three = [ws.cell(row=3, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        for col_idx, value in enumerate(row_three, start=1):
            ws.cell(row=2, column=col_idx, value=value)
        for col_idx, value in enumerate(row_two, start=1):
            ws.cell(row=3, column=col_idx, value=value)

        ws.cell(row=2, column=2, value="行政一部-已排序")
        ws.cell(row=3, column=2, value="行政部-已排序")
        ws.cell(row=2, column=3, value="D001")
        ws.cell(row=3, column=3, value="")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import_res = self.client.post(
            "/admin/departments/import",
            data={"file": (buf, "departments-reordered.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_res.status_code, 200)

        with self.app.app_context():
            departments = Department.query.order_by(Department.dept_no.asc()).all()
            self.assertEqual(len(departments), 2)

            parent = Department.query.filter_by(dept_no="D001").first()
            self.assertIsNotNone(parent)
            self.assertEqual(parent.dept_name, "行政部-已排序")

            child = Department.query.filter_by(dept_no="D010").first()
            self.assertIsNotNone(child)
            self.assertEqual(child.dept_name, "行政一部-已排序")
            self.assertEqual(child.parent_id, parent.id)

    def test_departments_import_ignores_malformed_hidden_metadata(self) -> None:
        for hidden_id in ("not-a-number", "Infinity", "1e309"):
            with self.subTest(hidden_id=hidden_id):
                export_res = self.client.get("/admin/departments/export")
                self.assertEqual(export_res.status_code, 200)

                wb = openpyxl.load_workbook(io.BytesIO(export_res.data))
                ws = wb.active
                ws.cell(row=2, column=4, value=hidden_id)
                ws.cell(row=2, column=2, value=f"行政部-安全导入-{hidden_id}")

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                import_res = self.client.post(
                    "/admin/departments/import",
                    data={"file": (buf, f"departments-malformed-metadata-{hidden_id}.xlsx")},
                    content_type="multipart/form-data",
                )
                self.assertEqual(import_res.status_code, 200)

                with self.app.app_context():
                    departments = Department.query.order_by(Department.id.asc()).all()
                    self.assertEqual(len(departments), 1)
                    self.assertEqual(departments[0].dept_no, "D001")
                    self.assertEqual(departments[0].dept_name, f"行政部-安全导入-{hidden_id}")

    def test_departments_import_rejects_rows_when_hidden_ids_mismatch_current_identity(self) -> None:
        export_res = self.client.get("/admin/departments/export")
        self.assertEqual(export_res.status_code, 200)

        other_db_path = os.path.join(self.tmpdir.name, "other.db")
        other_upload_dir = os.path.join(self.tmpdir.name, "other_uploads")
        other_app = Flask("departments_cross_db_test")
        other_app.config.update(
            TESTING=True,
            SECRET_KEY="test-secret",
            SQLALCHEMY_DATABASE_URI=f"sqlite:///{other_db_path}",
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
            JWT_EXPIRES_HOURS=12,
            JWT_EXPIRES_DELTA=__import__("datetime").timedelta(hours=12),
            UPLOAD_FOLDER=other_upload_dir,
        )
        os.makedirs(other_upload_dir, exist_ok=True)
        db.init_app(other_app)
        register_routes(other_app)

        with other_app.app_context():
            db.create_all()
            admin = User(username="other-admin", role="admin")
            admin.set_password("admin123")
            unrelated = Department(dept_no="X001", dept_name="外部部门")
            db.session.add_all([admin, unrelated])
            db.session.commit()
            self.assertEqual(unrelated.id, 1)

        other_client = other_app.test_client()
        other_client.post("/login", data={"username": "other-admin", "password": "admin123"})
        import_res = other_client.post(
            "/admin/departments/import",
            data={"file": (io.BytesIO(export_res.data), "departments-cross-db.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_res.status_code, 400)
        self.assertIn("原始部门", import_res.get_json()["error"])

        with other_app.app_context():
            unrelated = Department.query.filter_by(dept_no="X001").first()
            imported = Department.query.filter_by(dept_no="D001").first()

            self.assertIsNotNone(unrelated)
            self.assertEqual(unrelated.dept_name, "外部部门")
            self.assertIsNone(imported)
            self.assertEqual(Department.query.count(), 1)

    def test_departments_import_handles_duplicate_new_rows_for_same_dept_no(self) -> None:
        export_res = self.client.get("/admin/departments/export")
        self.assertEqual(export_res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(export_res.data))
        ws = wb.active
        ws.append(["D020", "新部门一版", "", ""])
        ws.append(["D020", "新部门二版", "", ""])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import_res = self.client.post(
            "/admin/departments/import",
            data={"file": (buf, "departments-duplicate-new-rows.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_res.status_code, 200)

        with self.app.app_context():
            matches = Department.query.filter_by(dept_no="D020").all()
            self.assertEqual(len(matches), 1)
            self.assertEqual(matches[0].dept_name, "新部门二版")

    def test_departments_import_allows_swapping_existing_department_numbers(self) -> None:
        with self.app.app_context():
            db.session.add(Department(dept_no="D002", dept_name="生产部"))
            db.session.commit()

        export_res = self.client.get("/admin/departments/export")
        self.assertEqual(export_res.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(export_res.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        dept_no_idx = headers.index("部门编号") + 1
        dept_name_idx = headers.index("部门名称") + 1

        row_by_name = {}
        for row_idx in range(2, ws.max_row + 1):
            dept_name = ws.cell(row=row_idx, column=dept_name_idx).value
            if dept_name:
                row_by_name[dept_name] = row_idx

        admin_row = row_by_name["行政部"]
        prod_row = row_by_name["生产部"]
        admin_dept_no = ws.cell(row=admin_row, column=dept_no_idx).value
        prod_dept_no = ws.cell(row=prod_row, column=dept_no_idx).value
        ws.cell(row=admin_row, column=dept_no_idx, value=prod_dept_no)
        ws.cell(row=prod_row, column=dept_no_idx, value=admin_dept_no)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import_res = self.client.post(
            "/admin/departments/import",
            data={"file": (buf, "departments-swapped.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(import_res.status_code, 200)

        with self.app.app_context():
            admin = Department.query.filter_by(dept_name="行政部").first()
            production = Department.query.filter_by(dept_name="生产部").first()

            self.assertIsNotNone(admin)
            self.assertIsNotNone(production)
            self.assertEqual(admin.dept_no, "D002")
            self.assertEqual(production.dept_no, "D001")


if __name__ == "__main__":
    unittest.main()
