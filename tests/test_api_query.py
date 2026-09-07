import io
import tempfile
import unittest
from datetime import datetime
from datetime import timedelta
from datetime import date
from types import SimpleNamespace

import openpyxl
from flask import Flask

from models import db
from models.account_set import AccountSet
from models.department import Department
from models.employee import Employee
from models.daily_record import DailyRecord
from models.leave import LeaveRecord
from models.manager_month_stat import ManagerMonthStat
from models.user import User, UserEmployeeAssignment
from routes import register_routes
from routes.auth_helpers import issue_slider_verified_token
from routes.query_core import _effective_attendance_day_value
from tests.csrf_helper import attach_origin


class ApiQueryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)
        self.app = Flask(__name__)
        self.app.config.update(
            TESTING=True,
            SECRET_KEY="test-secret",
            SQLALCHEMY_DATABASE_URI=f"sqlite:///{self.tmpdir.name}/query.db",
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
            JWT_EXPIRES_DELTA=timedelta(hours=12),
            FRONTEND_ORIGIN="http://localhost:5173",
            SESSION_COOKIE_NAME="api_access_token",
            SESSION_COOKIE_SAMESITE="None",
            SESSION_COOKIE_SECURE=False,
        )
        db.init_app(self.app)
        register_routes(self.app)

        with self.app.app_context():
            db.create_all()

            dept_a = Department(dept_no="D001", dept_name="制造一部")
            dept_b = Department(dept_no="D002", dept_name="制造二部")
            db.session.add_all([dept_a, dept_b])
            db.session.flush()

            employee_a = Employee(emp_no="E001", name="员工甲", dept_id=dept_a.id, is_manager=False)
            employee_b = Employee(emp_no="E002", name="员工乙", dept_id=dept_b.id, is_manager=False)
            manager = Employee(emp_no="M001", name="经理甲", dept_id=dept_a.id, is_manager=True)
            db.session.add_all([employee_a, employee_b, manager])
            db.session.flush()

            db.session.add(AccountSet(month="2026-05", name="2026年5月", is_active=True, is_locked=False))

            viewer = User(
                username="viewer",
                role="readonly",
                page_permissions={
                    "employee_dashboard": True,
                    "summary_download": True,
                },
            )
            viewer.set_password("viewer123")
            blocked = User(username="blocked", role="readonly", page_permissions={"employee_dashboard": True})
            blocked.set_password("blocked123")
            home_only = User(username="home-only", role="readonly", page_permissions={"query_home": True})
            home_only.set_password("home123")
            manager_viewer = User(
                username="manager-viewer",
                role="readonly",
                page_permissions={
                    "manager_overtime_query": True,
                    "manager_annual_leave_query": True,
                },
                profile_emp_no="M001",
                profile_name="经理甲账号",
            )
            manager_viewer.set_password("manager123")
            admin = User(username="admin", role="admin", profile_emp_no="A001", profile_name="系统管理员")
            admin.set_password("admin123")
            db.session.add_all([viewer, blocked, home_only, manager_viewer, admin])
            db.session.flush()

            db.session.add(UserEmployeeAssignment(user_id=viewer.id, emp_id=employee_a.id))
            db.session.add(UserEmployeeAssignment(user_id=blocked.id, emp_id=employee_a.id))
            db.session.add(ManagerMonthStat(emp_id=manager.id, year=2026, stat_type="overtime", prev_dec=2, m5=-0.5))
            db.session.add(ManagerMonthStat(emp_id=manager.id, year=2026, stat_type="annual_leave", m5=1))
            db.session.commit()

        self.client = attach_origin(self.app.test_client())

    def tearDown(self) -> None:
        with self.app.app_context():
            db.session.remove()
            db.drop_all()

    def _login(self, username: str, password: str):
        with self.app.app_context():
            captcha_token = issue_slider_verified_token()
        return self.client.post(
            "/api/auth/login",
            json={"username": username, "password": password, "captcha_token": captcha_token},
        )

    def test_query_routes_are_registered_under_api_prefix(self) -> None:
        rules = {rule.rule for rule in self.app.url_map.iter_rules()}

        self.assertTrue(
            {
                "/api/query/bootstrap",
                "/api/query/navigation",
                "/api/query/employee-dashboard",
                "/api/query/abnormal",
                "/api/query/punch-records",
                "/api/query/punch-records/modal-export",
                "/api/query/leave-records",
                "/api/query/leave-records/export",
                "/api/query/department-hours",
                "/api/query/manager-attendance",
                "/api/query/manager-punch-records",
                "/api/query/manager-leave-records",
                "/api/query/manager-overtime",
                "/api/query/manager-annual-leave",
                "/api/query/manager-department-hours",
                "/api/query/summary-download/export",
            }.issubset(rules)
        )

    def test_query_bootstrap_returns_account_sets_and_scoped_people(self) -> None:
        self._login("viewer", "viewer123")

        response = self.client.get("/api/query/bootstrap")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["account_sets"][0]["month"], "2026-05")
        # is_locked 供修正页/弹窗禁用编辑，缺失会导致锁定账套仍可提交
        self.assertIs(payload["account_sets"][0]["is_locked"], False)
        self.assertEqual([row["emp_no"] for row in payload["employees"]], ["E001"])
        self.assertEqual([row["dept_no"] for row in payload["departments"]], ["D001"])

    def test_query_bootstrap_succeeds_for_home_only_user(self) -> None:
        self._login("home-only", "home123")

        response = self.client.get("/api/query/bootstrap")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        # 纯首页权限用户也应拿到账套列表（首页月份选择器与摘要依赖它定位数据）
        self.assertEqual([row["month"] for row in payload["account_sets"]], ["2026-05"])
        # 但 departments 仍受查询中心权限约束，无查询权限时为空
        self.assertEqual(payload["departments"], [])
        self.assertEqual(payload["employees"], [])

    def test_query_navigation_api_returns_visible_query_entries(self) -> None:
        self._login("viewer", "viewer123")

        response = self.client.get("/api/query/navigation")

        self.assertEqual(response.status_code, 200)
        modules = response.get_json()["modules"]
        query_module = next(module for module in modules if module["slug"] == "query")
        entry_keys = {entry["key"] for entry in query_module["entries"]}
        self.assertIn("employee_dashboard", entry_keys)
        self.assertIn("summary_download", entry_keys)
        self.assertTrue(all("href" in entry for entry in query_module["entries"]))

    def test_query_navigation_api_exposes_disabled_users_in_settings(self) -> None:
        self._login("admin", "admin123")

        response = self.client.get("/api/query/navigation")

        self.assertEqual(response.status_code, 200)
        modules = response.get_json()["modules"]
        settings_module = next(module for module in modules if module["slug"] == "settings")
        entry_hrefs = {entry["href"] for entry in settings_module["entries"]}
        self.assertIn("/admin/accounts", entry_hrefs)
        self.assertIn("/admin/disabled-users", entry_hrefs)
        # 数据库设置是部署/初始化阶段页面：能登录说明数据库已配好，
        # 该入口不应出现在登录后的导航菜单中
        self.assertNotIn("/admin/database-settings", entry_hrefs)

    def test_legacy_employee_dashboard_route_is_not_available(self) -> None:
        self._login("viewer", "viewer123")

        response = self.client.get("/employee/dashboard")

        self.assertEqual(response.status_code, 404)

    def test_summary_download_wrapper_requires_matching_permission(self) -> None:
        self._login("blocked", "blocked123")

        response = self.client.get("/api/query/summary-download/export?month=2026-05")

        self.assertEqual(response.status_code, 403)

    def test_summary_download_export_includes_manager_hours_and_count(self) -> None:
        self._login("admin", "admin123")

        response = self.client.get(
            "/api/query/summary-download/export?month=2026-05&sheets=emp_dept_hours,mgr_dept_hours"
        )
        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        self.assertIn("员工部门工时查询", wb.sheetnames)
        self.assertIn("管理人员部门工时查询", wb.sheetnames)

        ws_emp = wb["员工部门工时查询"]
        self.assertEqual(ws_emp.cell(row=1, column=3).value, "部门人数")

        ws_mgr = wb["管理人员部门工时查询"]
        self.assertEqual(ws_mgr.cell(row=1, column=3).value, "部门人数")

        rows_mgr = list(ws_mgr.values)
        dept_names = [r[0] for r in rows_mgr]
        self.assertIn("制造一部", dept_names)
        
        # 检查是否成功统计到人数。经理甲属于制造一部，所以在经理部门工时里制造一部人数应当是1
        d001_row = next(r for r in rows_mgr if r[0] == "制造一部")
        self.assertEqual(d001_row[2], 1)

    def test_summary_download_export_manager_overtime_and_annual_leave(self) -> None:
        # 复现：勾选"管理人员加班查询"和"管理人员年假查询"后下载会报错。
        # 经理甲已存在 overtime/annual_leave 两份 ManagerMonthStat 数据。
        self._login("admin", "admin123")

        response = self.client.get(
            "/api/query/summary-download/export?month=2026-05&sheets=mgr_overtime,mgr_annual_leave"
        )

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        self.assertIn("管理人员加班查询", wb.sheetnames)
        self.assertIn("管理人员年假查询", wb.sheetnames)

        # 加班表表头应按预期顺序落地，且应包含经理甲一行（部门=制造一部）
        ws_overtime = wb["管理人员加班查询"]
        overtime_rows = list(ws_overtime.values)
        self.assertEqual(
            overtime_rows[0],
            ("部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"),
        )
        overtime_names = [row[1] for row in overtime_rows[1:]]
        self.assertIn("经理甲", overtime_names)

        # 年假表表头应按预期顺序落地（无"前年累积天数"列）
        ws_leave = wb["管理人员年假查询"]
        leave_rows = list(ws_leave.values)
        self.assertEqual(
            leave_rows[0],
            ("部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"),
        )
        leave_names = [row[1] for row in leave_rows[1:]]
        self.assertIn("经理甲", leave_names)

    def test_manager_overtime_export_returns_xlsx_with_expected_headers(self) -> None:
        self._login("admin", "admin123")

        response = self.client.get("/api/query/manager-overtime/export?year=2026")

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        rows = list(ws.values)
        self.assertEqual(
            rows[0],
            ("部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"),
        )
        self.assertIn("经理甲", [row[1] for row in rows[1:]])

    def test_manager_annual_leave_export_returns_xlsx_with_expected_headers(self) -> None:
        self._login("admin", "admin123")

        response = self.client.get("/api/query/manager-annual-leave/export?year=2026")

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        rows = list(ws.values)
        self.assertEqual(
            rows[0],
            ("部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"),
        )
        self.assertIn("经理甲", [row[1] for row in rows[1:]])

    def test_manager_profile_binding_can_query_overtime_and_annual_leave(self) -> None:
        self._login("manager-viewer", "manager123")

        overtime_response = self.client.get("/api/query/manager-overtime?year=2026")
        annual_leave_response = self.client.get("/api/query/manager-annual-leave?year=2026")

        self.assertEqual(overtime_response.status_code, 200)
        self.assertEqual(annual_leave_response.status_code, 200)

        overtime_payload = overtime_response.get_json()
        annual_leave_payload = annual_leave_response.get_json()

        self.assertEqual([row["name"] for row in overtime_payload["rows"]], ["经理甲"])
        self.assertEqual([row["name"] for row in annual_leave_payload["rows"]], ["经理甲"])
        self.assertEqual(overtime_payload["rows"][0]["remaining"], 0)
        self.assertEqual(annual_leave_payload["rows"][0]["remaining"], 0)

    def test_manager_attendance_api_toggles_punch_days_column(self) -> None:
        with self.app.app_context():
            manager = Employee.query.filter_by(emp_no="M001").first()
            db.session.add(
                DailyRecord(
                    emp_id=manager.id,
                    record_date=date(2026, 5, 1),
                    manager_payload={
                        "raw_data": {"上班1打卡时间": "08:00", "下班1打卡时间": "17:30"}
                    },
                )
            )
            db.session.commit()

        self._login("admin", "admin123")
        hidden_response = self.client.get(
            "/api/query/manager-attendance?month=2026-05&show_actual_attendance_days=1"
        )
        shown_response = self.client.get(
            "/api/query/manager-attendance?month=2026-05&show_actual_attendance_days=1&show_punch_days=1"
        )

        self.assertEqual(hidden_response.status_code, 200)
        self.assertEqual(shown_response.status_code, 200)
        hidden = hidden_response.get_json()
        shown = shown_response.get_json()
        self.assertNotIn("打卡天数", hidden["headers"])
        self.assertIn("打卡天数", shown["headers"])
        self.assertEqual(shown["headers"].index("打卡天数"), shown["headers"].index("实际出勤天数") + 1)
        self.assertEqual(len(shown["rows"][0]), len(hidden["rows"][0]) + 1)
        punch_index = shown["headers"].index("打卡天数")
        self.assertEqual(shown["rows"][0][punch_index], 1.0)

    def test_manager_attendance_export_toggles_punch_days_column(self) -> None:
        self._login("admin", "admin123")

        response = self.client.get("/api/query/manager-attendance/export?month=2026-05&show_punch_days=1")

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        header_row = [cell.value for cell in wb["管理人员查询"][1]]
        self.assertIn("打卡天数", header_row)

    def test_manager_punch_records_returns_manager_attendance_rows(self) -> None:
        with self.app.app_context():
            manager = Employee.query.filter_by(emp_no="M001").first()
            db.session.add(
                DailyRecord(
                    emp_id=manager.id,
                    record_date=date(2026, 5, 1),
                    actual_hours=240,
                    raw_data={
                        "姓名": "经理甲",
                        "上班1打卡时间": "08:00",
                        "下班1打卡时间": "12:00",
                        "上班2打卡时间": "13:00",
                    },
                    manager_payload={
                        "actual_hours": 240,
                        "late_minutes": 5,
                        "early_leave_minutes": 0,
                        "raw_data": {
                            "姓名": "经理甲",
                            "上班1打卡时间": "08:00",
                            "下班1打卡时间": "12:00",
                            "上班2打卡时间": "13:00",
                        },
                    },
                )
            )
            db.session.commit()

        self._login("admin", "admin123")
        response = self.client.get("/api/query/manager-punch-records?month=2026-05&emp_ids=3")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload[0]["name"], "经理甲")
        self.assertEqual(payload[0]["raw_punch_data"], "08:00,12:00,13:00")
        self.assertEqual(payload[0]["late_minutes"], 5)

    def test_manager_leave_records_filters_manager_leave_bucket(self) -> None:
        with self.app.app_context():
            manager = Employee.query.filter_by(emp_no="M001").first()
            db.session.add(
                LeaveRecord(
                    emp_id=manager.id,
                    leave_no="ML001",
                    leave_type="婚假",
                    start_time=datetime(2026, 5, 2, 9, 0),
                    end_time=datetime(2026, 5, 3, 9, 0),
                    duration=1,
                    reason="结婚",
                )
            )
            db.session.commit()

        self._login("admin", "admin123")
        response = self.client.get("/api/query/manager-leave-records?month=2026-05&emp_ids=3&leave_bucket=marriage")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload[0]["name"], "经理甲")
        self.assertEqual(payload[0]["leave_type"], "婚假")
        self.assertEqual(payload[0]["reason"], "结婚")

    def test_punch_records_modal_export_returns_requested_xlsx_columns(self) -> None:
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            db.session.add(
                DailyRecord(
                    emp_id=employee.id,
                    record_date=date(2026, 5, 1),
                    actual_hours=8,
                    check_in_times=["08:00"],
                    check_out_times=["17:30"],
                    raw_data={"打卡记录": ["08:00", "17:30"]},
                    employee_payload={
                        "actual_hours": 8,
                        "check_in_times": ["08:00"],
                        "check_out_times": ["17:30"],
                        "raw_data": {"打卡记录": ["08:00", "17:30"]},
                    },
                )
            )
            db.session.commit()

        self._login("admin", "admin123")
        response = self.client.get("/api/query/punch-records/modal-export?month=2026-05&emp_ids=1")

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        rows = list(ws.values)

        self.assertEqual(rows[0], ("部门", "姓名", "日期", "原始打卡数据"))
        self.assertEqual(rows[1], ("制造一部", "员工甲", "2026-05-01", "08:00,17:30"))

    def test_employee_dashboard_only_user_can_open_dashboard_punch_record_modal(self) -> None:
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            db.session.add(
                DailyRecord(
                    emp_id=employee.id,
                    record_date=date(2026, 5, 1),
                    actual_hours=8,
                    raw_data={"刷卡时间数据": "08:00,17:30"},
                    employee_payload={
                        "actual_hours": 8,
                        "raw_data": {"刷卡时间数据": "08:00,17:30"},
                    },
                )
            )
            db.session.commit()

        self._login("blocked", "blocked123")

        records_response = self.client.get("/api/query/punch-records?month=2026-05&emp_ids=1")
        export_response = self.client.get("/api/query/punch-records/modal-export?month=2026-05&emp_ids=1")
        standalone_export_response = self.client.get("/api/query/punch-records/export?month=2026-05&emp_ids=1")

        self.assertEqual(records_response.status_code, 200)
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(standalone_export_response.status_code, 403)
        self.assertEqual(records_response.get_json()[0]["raw_punch_data"], "08:00,17:30")

    def test_punch_records_api_falls_back_to_manager_slot_times_for_raw_punch_data(self) -> None:
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            db.session.add(
                DailyRecord(
                    emp_id=employee.id,
                    record_date=date(2026, 5, 2),
                    actual_hours=4,
                    check_in_times=[],
                    check_out_times=[],
                    raw_data={
                        "姓名": "员工甲",
                        "上班1打卡时间": "06:33",
                        "下班1打卡时间": "11:30",
                        "上班2打卡时间": None,
                        "下班2打卡时间": None,
                    },
                    employee_payload={
                        "actual_hours": 0,
                        "check_in_times": [],
                        "check_out_times": [],
                        "raw_data": {"刷卡时间数据": " "},
                    },
                    manager_payload={
                        "actual_hours": 240,
                        "late_minutes": 0,
                        "early_leave_minutes": 0,
                        "raw_data": {
                            "姓名": "员工甲",
                            "上班1打卡时间": "06:33",
                            "下班1打卡时间": "11:30",
                            "上班2打卡时间": None,
                            "下班2打卡时间": None,
                        },
                    },
                )
            )
            db.session.commit()

        self._login("admin", "admin123")
        response = self.client.get("/api/query/punch-records?month=2026-05&emp_ids=1")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        target = next(row for row in payload if row["date"] == "2026-05-02")
        self.assertEqual(target["raw_punch_data"], "06:33,11:30")

    def test_leave_records_api_filters_details_by_leave_bucket(self) -> None:
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            db.session.add_all(
                [
                    LeaveRecord(
                        emp_id=employee.id,
                        leave_no="L001",
                        leave_type="病假",
                        start_time=datetime(2026, 5, 2, 9, 0),
                        end_time=datetime(2026, 5, 3, 9, 0),
                        duration=1,
                        reason="发烧",
                    ),
                    LeaveRecord(
                        emp_id=employee.id,
                        leave_no="L002",
                        leave_type="事假",
                        start_time=datetime(2026, 5, 4, 9, 0),
                        end_time=datetime(2026, 5, 5, 9, 0),
                        duration=1,
                        reason="家中有事",
                    ),
                ]
            )
            db.session.commit()

        self._login("viewer", "viewer123")
        response = self.client.get("/api/query/leave-records?month=2026-05&emp_ids=1&leave_type=病假")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(
            payload,
            [
                {
                    "dept_name": "制造一部",
                    "name": "员工甲",
                    "leave_type": "病假",
                    "start_time": "2026-05-02 09:00",
                    "end_time": "2026-05-03 09:00",
                    "duration": 1.0,
                    "reason": "发烧",
                }
            ],
        )

    def test_leave_records_export_returns_requested_xlsx_columns(self) -> None:
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            db.session.add(
                LeaveRecord(
                    emp_id=employee.id,
                    leave_no="L003",
                    leave_type="补休(调休)",
                    start_time=datetime(2026, 5, 6, 9, 0),
                    end_time=datetime(2026, 5, 7, 9, 0),
                    duration=1,
                    reason="调休",
                )
            )
            db.session.commit()

        self._login("viewer", "viewer123")
        response = self.client.get("/api/query/leave-records/export?month=2026-05&emp_ids=1&leave_type=补休（调休）")

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        rows = list(ws.values)

        self.assertEqual(rows[0], ("部门", "姓名", "请假类型", "开始时间", "结束时间", "时长", "事由"))
        self.assertEqual(rows[1], ("制造一部", "员工甲", "补休（调休）", "2026-05-06 09:00", "2026-05-07 09:00", 1, "调休"))

    def test_half_day_detected_when_check_arrays_contain_shift_slot_pollution(self) -> None:
        # 复刻真实 bug：导入时「段2实际上班时间」(班次整点 12:00) 被塞进 check_in_times，
        # 导致 _punch_count 把真实 2 次刷卡 (13:06/17:27, 工时 5h) 错算成 3 次，半天漏判。
        # raw_data["刷卡时间数据"] 保留干净的真实刷卡源 "13:06,17:27"。
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            db.session.add(
                DailyRecord(
                    emp_id=employee.id,
                    record_date=date(2026, 5, 15),
                    actual_hours=5,
                    check_in_times=["13:06", "2026-05-15 12:00"],
                    check_out_times=["17:27", "2026-05-15 17:27"],
                    raw_data={
                        "刷卡时间数据": "13:06,17:27",
                        "段2实际上班时间": "2026-05-15 12:00",
                        "段2实际下班时间": "2026-05-15 17:27",
                    },
                    employee_payload={
                        "actual_hours": 5,
                        "check_in_times": ["13:06", "2026-05-15 12:00"],
                        "check_out_times": ["17:27", "2026-05-15 17:27"],
                        "raw_data": {"刷卡时间数据": "13:06,17:27"},
                    },
                )
            )
            db.session.commit()

        self._login("viewer", "viewer123")
        response = self.client.get("/api/query/employee-dashboard?month=2026-05")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        half_days_idx = payload["headers"].index("半勤天数")
        attendance_days_idx = payload["headers"].index("考勤天数")
        emp_row = next(row for row in payload["rows"] if row[1] == "E001")
        # 半勤天数应识别出这 1 天
        self.assertEqual(emp_row[half_days_idx], 1)
        # 半天考勤应计入考勤天数 0.5 天（而非整天 1.0）
        self.assertEqual(emp_row[attendance_days_idx], 0.5)

    def test_employee_dashboard_actual_attendance_days_uses_punch_count(self) -> None:
        # 打卡天数口径：当日有真实刷卡（≥1 次）记 1 天，无刷卡记 0 天。
        # setUp 未给 E001 建 5 月打卡记录，故本测试两天即全部聚合值。
        with self.app.app_context():
            employee = Employee.query.filter_by(emp_no="E001").first()
            # 全勤日：2 次真实刷卡 -> 1 天
            db.session.add(DailyRecord(
                emp_id=employee.id, record_date=date(2026, 5, 20),
                check_in_times=["08:00"], check_out_times=["17:00"],
                employee_payload={
                    "check_in_times": ["08:00"],
                    "check_out_times": ["17:00"],
                    "raw_data": {"刷卡时间数据": "08:00,17:00"},
                },
            ))
            # 仅 1 次真实刷卡 -> 1 天
            db.session.add(DailyRecord(
                emp_id=employee.id, record_date=date(2026, 5, 21),
                check_in_times=["08:00"], check_out_times=[],
                employee_payload={
                    "check_in_times": ["08:00"],
                    "check_out_times": [],
                    "raw_data": {"刷卡时间数据": "08:00"},
                },
            ))
            db.session.commit()

        self._login("viewer", "viewer123")
        # 带开关：打卡天数列存在，聚合 = 1（全勤）+ 1（单刷）
        res = self.client.get("/api/query/employee-dashboard?month=2026-05&show_actual_attendance_days=1")
        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        actual_idx = payload["headers"].index("打卡天数")
        emp_row = next(row for row in payload["rows"] if row[1] == "E001")
        self.assertEqual(emp_row[actual_idx], 2.0)

        # 不带开关：该列不存在
        res2 = self.client.get("/api/query/employee-dashboard?month=2026-05")
        self.assertNotIn("打卡天数", res2.get_json()["headers"])


class EffectiveAttendanceDayValueTests(unittest.TestCase):
    """晚加班条日的单日考勤值：白班出勤 + 晚加班顶班记 1.5 天；纯晚顶班记 0.5 天。"""

    def test_day_shift_with_evening_overtime_counts_one_and_half(self) -> None:
        record = SimpleNamespace(
            check_in_times=["07:54", "12:58"],
            check_out_times=["12:19", "20:08"],
            raw_data={},
            record_date=date(2026, 8, 31),
            employee=None,
            shift=None,
        )
        self.assertEqual(_effective_attendance_day_value(record, None, True), 1.5)

    def test_pure_evening_topup_counts_half_day(self) -> None:
        record = SimpleNamespace(
            check_in_times=["17:55"],
            check_out_times=["21:46"],
            raw_data={},
            record_date=date(2026, 7, 31),
            employee=None,
            shift=None,
        )
        self.assertEqual(_effective_attendance_day_value(record, None, True), 0.5)

    def test_manager_raw_punch_day_shift_with_evening_overtime(self) -> None:
        record = SimpleNamespace(
            check_in_times=[],
            check_out_times=[],
            raw_data={
                "上班1打卡时间": "07:54",
                "下班2打卡时间": "17:02",
                "上班3打卡时间": "17:02",
                "下班3打卡时间": "20:08",
            },
            record_date=date(2026, 8, 31),
            employee=None,
            shift=None,
        )
        self.assertEqual(_effective_attendance_day_value(record, None, True), 1.5)


if __name__ == "__main__":
    unittest.main()
