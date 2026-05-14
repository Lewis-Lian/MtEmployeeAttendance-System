from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO

import openpyxl
from flask import current_app, jsonify, request, send_file

from routes.auth import admin_required


def register_admin_import_routes(admin_bp) -> None:
    from . import admin as admin_module

    @admin_bp.route("/account-sets/<int:account_set_id>/imports", methods=["GET"])
    @admin_required
    def list_account_set_imports(account_set_id: int):
        row = admin_module.AccountSet.query.get_or_404(account_set_id)
        records = (
            admin_module.AccountSetImport.query.filter_by(account_set_id=row.id)
            .order_by(admin_module.AccountSetImport.id.desc())
            .all()
        )
        return jsonify(
            [
                {
                    "id": record.id,
                    "source_filename": record.source_filename,
                    "stored_path": record.stored_path,
                    "file_type": record.file_type,
                    "status": record.status,
                    "imported_count": record.imported_count,
                    "error_message": record.error_message,
                    "created_at": record.created_at.isoformat() if record.created_at else None,
                }
                for record in records
            ]
        )

    @admin_bp.route("/upload", methods=["POST"])
    @admin_required
    def upload_excel():
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]
        if not file.filename:
            return jsonify({"error": "Invalid filename"}), 400

        save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], file.filename)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        file.save(save_path)

        result = admin_module.ImportService.import_file(save_path)
        return jsonify(result)

    @admin_bp.route("/import/raw-files", methods=["POST"])
    @admin_required
    def import_raw_files():
        account_set_id = request.form.get("account_set_id", type=int)
        account_set = (
            admin_module.AccountSet.query.get(account_set_id)
            if account_set_id
            else admin_module.AccountSet.query.filter_by(is_active=True).first()
        )
        if not account_set:
            return jsonify({"status": "error", "message": "请先创建并选择账套"}), 400
        locked_error = admin_module._ensure_account_set_unlocked(account_set, "上传原始文件")
        if locked_error:
            return locked_error

        uploaded_files = [file for file in request.files.getlist("files") if (file.filename or "").strip()]
        if not uploaded_files:
            return jsonify({"status": "error", "message": "请至少选择一个要上传的源文件"}), 400

        results = []
        success = 0
        failed = 0
        for file in uploaded_files:
            filename = file.filename.strip()
            file_type = admin_module._account_set_file_type(filename)
            previous_record = admin_module.AccountSetImport.query.filter_by(
                account_set_id=account_set.id, file_type=file_type
            ).first()
            replaced = previous_record is not None

            if previous_record:
                old_path = (previous_record.stored_path or "").strip()
                if old_path and os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception:
                        pass
                admin_module.db.session.delete(previous_record)
                admin_module.db.session.flush()

            account_set_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "account_sets", account_set.month)
            os.makedirs(account_set_dir, exist_ok=True)
            save_name = f"{int(datetime.now().timestamp())}_{filename}"
            save_path = os.path.join(account_set_dir, save_name)
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            file.save(save_path)

            import_record = admin_module.AccountSetImport(
                account_set_id=account_set.id,
                source_filename=filename,
                stored_path=save_path,
                file_type=file_type,
                status="uploaded",
                imported_count=0,
            )
            admin_module.db.session.add(import_record)

            try:
                success += 1
                import_record.error_message = None
                results.append({"file": filename, "status": "ok", "message": "replaced" if replaced else "uploaded"})
            except Exception as exc:
                failed += 1
                import_record.status = "error"
                import_record.error_message = str(exc)
                results.append({"file": filename, "status": "error", "error": str(exc)})
            admin_module.db.session.commit()

        return jsonify(
            {
                "status": "ok" if failed == 0 else "partial",
                "account_set": admin_module._serialize_account_set(account_set),
                "total": len(uploaded_files),
                "success": success,
                "failed": failed,
                "results": results,
            }
        )

    @admin_bp.route("/manager-overtime/template", methods=["GET"])
    @admin_required
    def download_manager_overtime_template():
        return admin_module._download_manager_stat_template("overtime")

    @admin_bp.route("/manager-overtime/import", methods=["POST"])
    @admin_required
    def import_manager_overtime():
        year = request.form.get("year", type=int) or datetime.now().year
        locked_error = admin_module._ensure_year_months_unlocked(
            year, "导入管理人员加班统计", include_prev_dec=True
        )
        if locked_error:
            return locked_error
        return admin_module._import_manager_stat_file("overtime", year)

    @admin_bp.route("/manager-overtime/export", methods=["GET"])
    @admin_required
    def export_manager_overtime():
        year = request.args.get("year", type=int) or datetime.now().year
        return admin_module._export_manager_overtime_workbook(year)

    @admin_bp.route("/manager-annual-leave/template", methods=["GET"])
    @admin_required
    def download_manager_annual_leave_template():
        return admin_module._download_manager_stat_template("annual_leave")

    @admin_bp.route("/manager-annual-leave/import", methods=["POST"])
    @admin_required
    def import_manager_annual_leave():
        year = request.form.get("year", type=int) or datetime.now().year
        locked_error = admin_module._ensure_year_months_unlocked(year, "导入管理人员年休统计")
        if locked_error:
            return locked_error
        return admin_module._import_manager_stat_file("annual_leave", year)

    @admin_bp.route("/manager-annual-leave/export", methods=["GET"])
    @admin_required
    def export_manager_annual_leave():
        year = request.args.get("year", type=int) or datetime.now().year
        return admin_module._export_manager_annual_leave_workbook(year)

    @admin_bp.route("/departments/import", methods=["POST"])
    @admin_required
    def import_departments_xlsx():
        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "file is required"}), 400
        if not file.filename.lower().endswith(".xlsx"):
            return jsonify({"error": "only .xlsx is supported"}), 400

        save_path = os.path.join(
            current_app.config["UPLOAD_FOLDER"], f"departments_{int(datetime.now().timestamp())}.xlsx"
        )
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        file.save(save_path)

        wb = openpyxl.load_workbook(save_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not raw_rows:
            return jsonify({"error": "empty file"}), 400

        header_idx, header_map = admin_module._parse_header_row(raw_rows, ["部门编号", "部门名称", "上级部门编号"])
        dept_no_idx = header_map.get("部门编号", -1)
        dept_name_idx = header_map.get("部门名称", -1)
        parent_no_idx = header_map.get("上级部门编号", -1)
        original_id_idx = header_map.get(admin_module._DEPARTMENT_ORIGINAL_ID_HEADER, -1)
        if dept_no_idx < 0 or dept_name_idx < 0:
            return jsonify({"error": "missing required headers: 部门编号, 部门名称"}), 400

        imported = 0
        pending_parent_links: list[tuple[admin_module.Department, str]] = []
        existing_departments = admin_module.Department.query.order_by(admin_module.Department.id.asc()).all()
        departments_by_id = {department.id: department for department in existing_departments}
        departments_by_dept_no = {
            department.dept_no: department for department in existing_departments if department.dept_no
        }
        original_identities_by_id = admin_module._load_department_identity_metadata(wb)
        staged_rows: list[tuple[admin_module.Department, str, str, str]] = []
        staged_departments_by_dept_no = dict(departments_by_dept_no)
        used_temp_dept_nos = set(departments_by_dept_no)
        temp_index = 1

        for row_idx, row in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
            dept_no = (
                str(row[dept_no_idx]).strip()
                if dept_no_idx < len(row) and row[dept_no_idx] is not None
                else ""
            )
            dept_name = (
                str(row[dept_name_idx]).strip()
                if dept_name_idx < len(row) and row[dept_name_idx] is not None
                else ""
            )
            parent_no = (
                str(row[parent_no_idx]).strip()
                if parent_no_idx >= 0 and parent_no_idx < len(row) and row[parent_no_idx] is not None
                else ""
            )
            original_id = (
                admin_module._parse_department_original_id(row[original_id_idx])
                if original_id_idx >= 0 and original_id_idx < len(row)
                else None
            )
            if not dept_no or not dept_name:
                continue

            department = None
            if original_id is not None:
                original_identity = original_identities_by_id.get(original_id)
                if original_identity:
                    candidate = departments_by_id.get(original_id)
                    if not admin_module._department_matches_original_identity(candidate, original_identity):
                        admin_module.db.session.rollback()
                        return (
                            jsonify({"error": f"第 {row_idx} 行的原始部门元数据与当前数据不匹配，请重新导出后再导入"}),
                            400,
                        )
                    department = candidate
            if department is None:
                department = staged_departments_by_dept_no.get(dept_no)
            if department is None:
                while True:
                    temp_dept_no = f"{admin_module._DEPARTMENT_IMPORT_TEMP_PREFIX}{temp_index:04d}"
                    temp_index += 1
                    if temp_dept_no not in used_temp_dept_nos:
                        used_temp_dept_nos.add(temp_dept_no)
                        break
                department = admin_module.Department(dept_no=temp_dept_no, dept_name=dept_name)
                admin_module.db.session.add(department)
            if department.id is not None:
                departments_by_id[department.id] = department
            staged_departments_by_dept_no[dept_no] = department
            staged_rows.append((department, dept_no, dept_name, parent_no))
            imported += 1

        for department, dept_no, _dept_name, _parent_no in staged_rows:
            current_dept_no = (department.dept_no or "").strip()
            if current_dept_no == dept_no or current_dept_no.startswith(admin_module._DEPARTMENT_IMPORT_TEMP_PREFIX):
                continue
            while True:
                temp_dept_no = f"{admin_module._DEPARTMENT_IMPORT_TEMP_PREFIX}{temp_index:04d}"
                temp_index += 1
                if temp_dept_no not in used_temp_dept_nos:
                    used_temp_dept_nos.add(temp_dept_no)
                    break
            department.dept_no = temp_dept_no

        admin_module.db.session.flush()
        for department, dept_no, dept_name, parent_no in staged_rows:
            department.dept_no = dept_no
            department.dept_name = dept_name
            pending_parent_links.append((department, parent_no))

        admin_module.db.session.flush()
        for department, parent_no in pending_parent_links:
            if not parent_no:
                department.parent_id = None
                continue
            parent = admin_module.Department.query.filter_by(dept_no=parent_no).first()
            if not parent or parent.id == department.id:
                department.parent_id = None
                continue
            department.parent_id = parent.id

        admin_module.db.session.commit()
        return jsonify({"status": "ok", "imported": imported})

    @admin_bp.route("/departments/template", methods=["GET"])
    @admin_required
    def download_departments_template():
        wb = admin_module._build_departments_workbook(
            [
                ("D001", "行政部", "", ""),
                ("D002", "生产中心", "", ""),
                ("D003", "生产一部", "D002", ""),
            ]
        )
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="部门导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @admin_bp.route("/departments/export", methods=["GET"])
    @admin_required
    def export_departments_xlsx():
        departments = admin_module.Department.query.order_by(
            admin_module.Department.dept_no.asc(), admin_module.Department.dept_name.asc()
        ).all()
        rows = [
            (
                department.dept_no or "",
                department.dept_name or "",
                department.parent.dept_no if department.parent else "",
                str(department.id),
            )
            for department in departments
        ]
        wb = admin_module._build_departments_workbook(rows, include_identity_metadata=True)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="部门导出.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @admin_bp.route("/employees/import", methods=["POST"])
    @admin_required
    def import_employees_xlsx():
        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "file is required"}), 400
        if not file.filename.lower().endswith(".xlsx"):
            return jsonify({"error": "only .xlsx is supported"}), 400

        save_path = os.path.join(
            current_app.config["UPLOAD_FOLDER"], f"employees_{int(datetime.now().timestamp())}.xlsx"
        )
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        file.save(save_path)

        wb = openpyxl.load_workbook(save_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not raw_rows:
            return jsonify({"error": "empty file"}), 400

        header_idx, header_map = admin_module._parse_header_row(raw_rows, ["人员编号", "人员姓名", "部门名称", "班次编号"])
        emp_no_idx = header_map.get("人员编号", -1)
        name_idx = header_map.get("人员姓名", -1)
        dept_idx = header_map.get("部门名称", -1)
        shift_idx = header_map.get("班次编号", -1)
        manager_idx = header_map.get("是否管理人员", -1)
        nursing_idx = header_map.get("是否哺乳假", -1)
        employee_source_idx = header_map.get("员工考勤统计来源", -1)
        manager_source_idx = header_map.get("管理人员考勤统计来源", -1)
        if emp_no_idx < 0 or name_idx < 0:
            return jsonify({"error": "missing required headers: 人员编号, 人员姓名"}), 400

        imported = 0
        for row in raw_rows[header_idx + 1 :]:
            emp_no = (
                str(row[emp_no_idx]).strip()
                if emp_no_idx < len(row) and row[emp_no_idx] is not None
                else ""
            )
            name = (
                str(row[name_idx]).strip()
                if name_idx < len(row) and row[name_idx] is not None
                else ""
            )
            dept_name = (
                str(row[dept_idx]).strip()
                if dept_idx >= 0 and dept_idx < len(row) and row[dept_idx] is not None
                else ""
            )
            shift_no = (
                str(row[shift_idx]).strip()
                if shift_idx >= 0 and shift_idx < len(row) and row[shift_idx] is not None
                else ""
            )
            is_manager = (
                admin_module.parse_bool_zh(row[manager_idx])
                if manager_idx >= 0 and manager_idx < len(row)
                else False
            )
            is_nursing = (
                admin_module.parse_bool_zh(row[nursing_idx])
                if nursing_idx >= 0 and nursing_idx < len(row)
                else False
            )
            employee_stats_attendance_source = admin_module._parse_attendance_source(
                row[employee_source_idx] if employee_source_idx >= 0 and employee_source_idx < len(row) else None,
                admin_module.ATTENDANCE_SOURCE_EMPLOYEE,
            )
            manager_stats_attendance_source = admin_module._parse_attendance_source(
                row[manager_source_idx] if manager_source_idx >= 0 and manager_source_idx < len(row) else None,
                admin_module.ATTENDANCE_SOURCE_MANAGER,
            )
            if not emp_no or not name:
                continue
            department = admin_module._resolve_department(dept_name) if dept_name else None
            shift = admin_module._resolve_shift(shift_no) if shift_no else None
            employee = admin_module.Employee.query.filter_by(emp_no=emp_no).first()
            if not employee:
                employee = admin_module.Employee(
                    emp_no=emp_no,
                    name=name,
                    dept_id=department.id if department else None,
                    is_manager=is_manager,
                    is_nursing=is_nursing,
                    employee_stats_attendance_source=employee_stats_attendance_source,
                    manager_stats_attendance_source=manager_stats_attendance_source,
                )
                admin_module.db.session.add(employee)
                admin_module.db.session.flush()
            else:
                employee.name = name
                employee.dept_id = department.id if department else None
                employee.is_manager = is_manager
                employee.is_nursing = is_nursing
                employee.employee_stats_attendance_source = employee_stats_attendance_source
                employee.manager_stats_attendance_source = manager_stats_attendance_source
            admin_module._assign_employee_shift(employee, shift)
            imported += 1

        admin_module.db.session.commit()
        return jsonify({"status": "ok", "imported": imported})

    @admin_bp.route("/employees/template", methods=["GET"])
    @admin_required
    def download_employees_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工导入模板"
        ws.append(
            [
                "人员编号",
                "人员姓名",
                "部门名称",
                "班次编号",
                "是否管理人员",
                "是否哺乳假",
                "员工考勤统计来源",
                "管理人员考勤统计来源",
            ]
        )
        ws.append(
            [
                "1001001",
                "张三",
                "生产中心",
                "A00001",
                "否",
                "否",
                "员工考勤源文件取值",
                "管理人员考勤源文件取值",
            ]
        )
        ws.append(
            [
                "1001002",
                "李四",
                "行政部",
                "A00002",
                "是",
                "是",
                "员工考勤源文件取值",
                "管理人员考勤源文件取值",
            ]
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="员工导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @admin_bp.route("/employees/export", methods=["GET"])
    @admin_required
    def export_employees_xlsx():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工主数据导出"
        ws.append(
            [
                "人员编号",
                "人员姓名",
                "部门名称",
                "班次编号",
                "是否管理人员",
                "是否哺乳假",
                "员工考勤统计来源",
                "管理人员考勤统计来源",
            ]
        )

        requested_ids: list[int] = []
        for raw in request.args.getlist("ids"):
            for part in str(raw).split(","):
                text = part.strip()
                if text.isdigit():
                    requested_ids.append(int(text))

        query = admin_module.Employee.query
        if requested_ids:
            query = query.filter(admin_module.Employee.id.in_(requested_ids))

        employees = query.order_by(admin_module.Employee.emp_no.asc(), admin_module.Employee.name.asc()).all()
        for employee in employees:
            shift = employee.shift_assignment.shift if employee.shift_assignment else None
            ws.append(
                [
                    employee.emp_no or "",
                    employee.name or "",
                    employee.department.dept_name if employee.department else "",
                    shift.shift_no if shift else "",
                    "是" if employee.is_manager else "否",
                    "是" if employee.is_nursing else "否",
                    (
                        "管理人员考勤源文件取值"
                        if employee.employee_stats_attendance_source == admin_module.ATTENDANCE_SOURCE_MANAGER
                        else (
                            "自动回退"
                            if employee.employee_stats_attendance_source
                            == admin_module.ATTENDANCE_SOURCE_AUTO_FALLBACK
                            else "员工考勤源文件取值"
                        )
                    ),
                    (
                        "员工考勤源文件取值"
                        if employee.manager_stats_attendance_source == admin_module.ATTENDANCE_SOURCE_EMPLOYEE
                        else (
                            "自动回退"
                            if employee.manager_stats_attendance_source
                            == admin_module.ATTENDANCE_SOURCE_AUTO_FALLBACK
                            else "管理人员考勤源文件取值"
                        )
                    ),
                ]
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="员工主数据导出.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
