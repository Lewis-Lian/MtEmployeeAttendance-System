from __future__ import annotations

import os
from collections import defaultdict
from copy import copy
from datetime import date, datetime, time, timedelta
import math
import re
from io import BytesIO

from flask import Blueprint, jsonify, render_template, request, g, send_file
from sqlalchemy.orm import joinedload
import openpyxl

from models.employee import Employee
from models.department import Department
from models.daily_record import DailyRecord
from models.overtime import OvertimeRecord
from models.leave import LeaveRecord
from models.annual_leave import AnnualLeave
from models.monthly_report import MonthlyReport
from models.account_set import AccountSet
from models.user import EMPLOYEE_PAGE_PERMISSION_KEYS, MANAGER_PAGE_PERMISSION_KEYS, UserEmployeeAssignment, UserDepartmentAssignment
from services.attendance_service import AttendanceService
from services.attendance_source_service import (
    EMPLOYEE_STATS_CONTEXT,
    MANAGER_STATS_CONTEXT,
    attendance_views_by_employee,
)
from services.manager_attendance_service import (
    MANAGER_HEADERS,
    ManagerAttendanceOptions,
    build_manager_rows,
    manager_headers,
    rows_as_table,
)
from routes.auth import login_required, page_permission_required
from utils.helpers import overlap_duration_days


employee_bp = Blueprint("employee", __name__, url_prefix="/employee")


FINAL_HEADERS = [
    "部门名称",
    "人员编号",
    "人员名称",
    "考勤天数",
    "病假（次数）",
    "工伤（次数）",
    "丧假（次数）",
    "事假（次数）",
    "补休（调休）(次)",
    "婚假（次）",
    "病假时长（天）",
    "工伤时长（天）",
    "丧假时长（天）",
    "事假时长（天）",
    "补休（调休）(天)",
    "婚假（天）",
    "工时",
    "半勤天数",
    "备注",
]

LEAVE_COUNT_HEADERS = {
    "病假（次数）",
    "工伤（次数）",
    "丧假（次数）",
    "事假（次数）",
    "补休（调休）(次)",
    "婚假（次）",
}

LEAVE_DURATION_HEADERS = {
    "病假时长（天）",
    "工伤时长（天）",
    "丧假时长（天）",
    "事假时长（天）",
    "补休（调休）(天)",
    "婚假（天）",
}


def _filter_final_columns(headers: list[str], rows: list[list[object]]) -> tuple[list[str], list[list[object]]]:
    requested = (request.args.get("final_headers") or "").strip()
    if requested:
        wanted = [h.strip() for h in requested.split(",") if h.strip()]
        wanted_set = set(wanted)
        keep_indexes: list[int] = []
        filtered_headers: list[str] = []
        for idx, header in enumerate(headers):
            if header in wanted_set:
                keep_indexes.append(idx)
                filtered_headers.append(header)
        filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
        return filtered_headers, filtered_rows

    show_leave_counts = request.args.get("show_leave_counts", "").strip() in {"1", "true", "True"}
    show_leave_durations = request.args.get("show_leave_durations", "").strip() in {"1", "true", "True"}

    keep_indexes: list[int] = []
    filtered_headers: list[str] = []
    for idx, header in enumerate(headers):
        if header in LEAVE_COUNT_HEADERS and not show_leave_counts:
            continue
        if header in LEAVE_DURATION_HEADERS and not show_leave_durations:
            continue
        keep_indexes.append(idx)
        filtered_headers.append(header)

    filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
    return filtered_headers, filtered_rows


def _filter_punch_columns(headers: list[str], rows: list[list[object]]) -> tuple[list[str], list[list[object]]]:
    requested = (request.args.get("punch_headers") or "").strip()
    if not requested:
        return headers, rows
    wanted = [h.strip() for h in requested.split(",") if h.strip()]
    wanted_set = set(wanted)
    keep_indexes: list[int] = []
    filtered_headers: list[str] = []
    for idx, header in enumerate(headers):
        if header in wanted_set:
            keep_indexes.append(idx)
            filtered_headers.append(header)
    filtered_rows = [[row[idx] if idx < len(row) else "" for idx in keep_indexes] for row in rows]
    return filtered_headers, filtered_rows


def _normalize_leave_type(value: str | None) -> str:
    text = (value or "").strip()
    if text in {"补休(调休)", "补休（调休）"}:
        return "补休（调休）"
    return text


def _leave_bucket(value: str | None) -> str | None:
    text = _normalize_leave_type(value)
    if not text:
        return None
    if "病假" in text:
        return "病假"
    if "工伤" in text:
        return "工伤"
    if "丧假" in text:
        return "丧假"
    if "事假" in text:
        return "事假"
    if "补休" in text or "调休" in text:
        return "补休（调休）"
    if "婚假" in text:
        return "婚假"
    # Source file may use generic "请假"; align it to personal leave bucket.
    if "请假" in text:
        return "事假"
    return None


def _normalized_leave_days(duration: float | int | None) -> float:
    value = float(duration or 0)
    if value <= 0:
        return 0.0
    int_days = math.floor(value)
    frac = value - int_days
    if frac > 0.17:
        return float(int_days + 1)
    if frac > 0.08:
        return float(int_days) + 0.5
    return float(int_days)


def _leave_days_in_month(record: LeaveRecord, month: str) -> float:
    datetime_range = _month_datetime_range(month)
    if not datetime_range:
        return 0.0
    start_dt, end_dt = datetime_range
    return overlap_duration_days(record.start_time, record.end_time, start_dt, end_dt)


def _month_date_range(month: str) -> tuple[date, date] | None:
    try:
        start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError:
        return None
    if start.month == 12:
        end = date(start.year + 1, 1, 1)
    else:
        end = date(start.year, start.month + 1, 1)
    return start, end


def _month_datetime_range(month: str) -> tuple[datetime, datetime] | None:
    bounds = _month_date_range(month)
    if not bounds:
        return None
    start, end = bounds
    return datetime.combine(start, time.min), datetime.combine(end, time.min)


def _has_punch_record(record) -> bool:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    in_times = [str(x).strip() for x in (record.check_in_times or []) if x is not None and str(x).strip()]
    out_times = [str(x).strip() for x in (record.check_out_times or []) if x is not None and str(x).strip()]
    if in_times or out_times:
        return True

    manager_time_keys = (
        "上班1打卡时间",
        "下班1打卡时间",
        "上班2打卡时间",
        "下班2打卡时间",
        "上班3打卡时间",
        "下班3打卡时间",
        "上班4打卡时间",
        "下班4打卡时间",
    )
    punch_data = str(raw.get("刷卡时间数据") or "").strip()
    if punch_data:
        return True
    return any(str(raw.get(key) or "").strip() for key in manager_time_keys)


def _attendance_day_value(record) -> float:
    if not _has_punch_record(record):
        return 0.0
    if (record.actual_hours or 0) < 2:
        return 0.5
    return 1.0


def _normalize_punch_token(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    m = re.search(r"(\d{1,2}:\d{2})", text)
    if not m:
        return text
    hh, mm = m.group(1).split(":")
    return f"{int(hh):02d}:{mm}"


def _punch_events(record) -> set[str]:
    events: set[str] = set()
    for raw in (record.check_in_times or []):
        token = _normalize_punch_token(raw)
        if token:
            events.add(token)
    for raw in (record.check_out_times or []):
        token = _normalize_punch_token(raw)
        if token:
            events.add(token)
    return events


def _punch_count(record) -> int:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    in_events = {_normalize_punch_token(x) for x in (record.check_in_times or [])}
    in_events = {x for x in in_events if x}
    out_events = {_normalize_punch_token(x) for x in (record.check_out_times or [])}
    out_events = {x for x in out_events if x}

    # Some source rows duplicate the same swipe into both in/out arrays.
    overlap = in_events & out_events
    if overlap:
        in_events -= overlap
        out_events -= overlap

    count = len(in_events) + len(out_events)
    if count:
        return count

    manager_time_keys = (
        "上班1打卡时间",
        "下班1打卡时间",
        "上班2打卡时间",
        "下班2打卡时间",
        "上班3打卡时间",
        "下班3打卡时间",
        "上班4打卡时间",
        "下班4打卡时间",
    )
    manager_events = {
        _normalize_punch_token(raw.get(key))
        for key in manager_time_keys
        if _normalize_punch_token(raw.get(key))
    }
    if manager_events:
        return len(manager_events)

    punch_data = str(raw.get("刷卡时间数据") or "").strip()
    if punch_data:
        return len(re.findall(r"(\d{1,2}:\d{2})", punch_data))
    return 0


def _punch_round_count(record) -> int:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    in_events = {_normalize_punch_token(x) for x in (record.check_in_times or [])}
    in_events = {x for x in in_events if x}
    out_events = {_normalize_punch_token(x) for x in (record.check_out_times or [])}
    out_events = {x for x in out_events if x}

    overlap = in_events & out_events
    if overlap:
        in_events -= overlap
        out_events -= overlap

    # Query page shows "打卡轮次" (e.g. 上午+下午 = 2), not raw swipe points.
    rounds = max(len(in_events), len(out_events))
    if rounds:
        return rounds

    manager_pairs = (
        ("上班1打卡时间", "下班1打卡时间"),
        ("上班2打卡时间", "下班2打卡时间"),
        ("上班3打卡时间", "下班3打卡时间"),
        ("上班4打卡时间", "下班4打卡时间"),
    )
    manager_rounds = 0
    for in_key, out_key in manager_pairs:
        if str(raw.get(in_key) or "").strip() or str(raw.get(out_key) or "").strip():
            manager_rounds += 1
    return manager_rounds


def _format_punch_times(values: list[object] | None) -> str:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in (values or []):
        token = _normalize_punch_token(raw)
        if not token or token in seen:
            continue
        seen.add(token)
        normalized.append(token)
    return " / ".join(normalized)


def _repair_mojibake(text: str) -> str:
    try:
        return text.encode("latin1").decode("gbk")
    except Exception:
        return text


def _extract_raw_punch_data(record) -> str:
    raw = record.raw_data or {}
    if not isinstance(raw, dict):
        raw = {}
    if isinstance(raw.get("raw_data"), dict):
        raw = raw.get("raw_data") or {}

    direct_keys = {"刷卡时间数据", "原始刷卡数据", "刷卡时间", "打卡记录"}
    for key in direct_keys:
        value = raw.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()

    for key, value in raw.items():
        if value is None or not str(value).strip():
            continue
        repaired = _repair_mojibake(str(key))
        if ("刷卡" in repaired and "时间" in repaired) or ("打卡" in repaired and "记录" in repaired):
            return str(value).strip()

    return ""


def _raw_punch_count(record) -> int:
    raw = _extract_raw_punch_data(record)
    if raw:
        tokens = re.findall(r"(\d{1,2}:\d{2})", raw)
        if tokens:
            return len(tokens)
    return _punch_round_count(record)


def _parse_punch_dt(value: object, record_date) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    # Full datetime like 2026-03-11 18:00
    m_dt = re.search(r"(\d{4}-\d{2}-\d{2})\s+(\d{1,2}:\d{2})", text)
    if m_dt:
        try:
            dt_text = f"{m_dt.group(1)} {m_dt.group(2)}"
            return datetime.strptime(dt_text, "%Y-%m-%d %H:%M")
        except Exception:
            pass
    # Time only like 07:31, attach record_date
    m_tm = re.search(r"(\d{1,2}):(\d{2})", text)
    if m_tm and record_date:
        try:
            hh = int(m_tm.group(1))
            mm = int(m_tm.group(2))
            return datetime.combine(record_date, datetime.min.time()).replace(hour=hh, minute=mm)
        except Exception:
            return None
    return None


def _resolve_shift_for_record(record):
    if record.employee and record.employee.shift_assignment and record.employee.shift_assignment.shift:
        return record.employee.shift_assignment.shift
    if record.shift:
        return record.shift
    return None


def _parse_slot_dt(record_date, hhmm: str, anchor: datetime | None = None) -> datetime | None:
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", str(hhmm or ""))
    if not m or not record_date:
        return None
    base = datetime.combine(record_date, datetime.min.time()).replace(hour=int(m.group(1)), minute=int(m.group(2)))
    if anchor and base < anchor:
        base += timedelta(days=1)
    return base


def _build_shift_break_windows(record) -> list[tuple[datetime, datetime]]:
    shift = _resolve_shift_for_record(record)
    slots = (shift.time_slots if shift else None) or []
    parsed_slots: list[tuple[datetime, datetime]] = []
    prev_end: datetime | None = None

    for slot in slots:
        if not isinstance(slot, (list, tuple)) or len(slot) != 2:
            continue
        s0 = str(slot[0] or "").strip()
        s1 = str(slot[1] or "").strip()

        # normal pair: ["08:00","11:20"]
        if re.match(r"^\d{1,2}:\d{2}$", s0) and re.match(r"^\d{1,2}:\d{2}$", s1):
            start = _parse_slot_dt(record.record_date, s0, prev_end)
            end = _parse_slot_dt(record.record_date, s1, start)
            if start and end:
                if end < start:
                    end += timedelta(days=1)
                parsed_slots.append((start, end))
                prev_end = end
            continue

        # malformed pair fallback: ["08:00","11:00,12:00-17:00,..."]
        if re.match(r"^\d{1,2}:\d{2}$", s0):
            first_end = s1.split(",")[0].strip()
            if re.match(r"^\d{1,2}:\d{2}$", first_end):
                start = _parse_slot_dt(record.record_date, s0, prev_end)
                end = _parse_slot_dt(record.record_date, first_end, start)
                if start and end:
                    if end < start:
                        end += timedelta(days=1)
                    parsed_slots.append((start, end))
                    prev_end = end

        for a, b in re.findall(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})", s1):
            start = _parse_slot_dt(record.record_date, a, prev_end)
            end = _parse_slot_dt(record.record_date, b, start)
            if not start or not end:
                continue
            if end < start:
                end += timedelta(days=1)
            parsed_slots.append((start, end))
            prev_end = end

    parsed_slots.sort(key=lambda x: x[0])
    breaks: list[tuple[datetime, datetime]] = []
    for idx in range(len(parsed_slots) - 1):
        curr_end = parsed_slots[idx][1]
        next_start = parsed_slots[idx + 1][0]
        if next_start > curr_end:
            breaks.append((curr_end, next_start))
    return breaks


def _calc_two_punch_hours_with_shift_break(record) -> float | None:
    raw = _extract_raw_punch_data(record)
    times = re.findall(r"(\d{1,2}:\d{2})", raw)
    if len(times) != 2:
        return None
    start = _parse_slot_dt(record.record_date, times[0])
    end = _parse_slot_dt(record.record_date, times[1], start)
    if not start or not end:
        return None
    if end < start:
        end += timedelta(days=1)

    total_seconds = (end - start).total_seconds()
    if total_seconds <= 0:
        return None

    for b_start, b_end in _build_shift_break_windows(record):
        overlap_start = max(start, b_start)
        overlap_end = min(end, b_end)
        if overlap_end > overlap_start:
            total_seconds -= (overlap_end - overlap_start).total_seconds()

    hours = max(total_seconds / 3600.0, 0.0)
    return round(hours, 2)


def _calc_record_work_hours(record) -> tuple[float, int]:
    special_hours = _calc_two_punch_hours_with_shift_break(record)
    if special_hours is not None:
        return special_hours, 0

    in_times: list[datetime] = []
    out_times: list[datetime] = []
    in_seen: set[tuple[int, int, int, int, int]] = set()
    out_seen: set[tuple[int, int, int, int, int]] = set()

    for raw in (record.check_in_times or []):
        dt = _parse_punch_dt(raw, record.record_date)
        if not dt:
            continue
        key = (dt.year, dt.month, dt.day, dt.hour, dt.minute)
        if key in in_seen:
            continue
        in_seen.add(key)
        in_times.append(dt)

    for raw in (record.check_out_times or []):
        dt = _parse_punch_dt(raw, record.record_date)
        if not dt:
            continue
        key = (dt.year, dt.month, dt.day, dt.hour, dt.minute)
        if key in out_seen:
            continue
        out_seen.add(key)
        out_times.append(dt)

    in_times.sort()
    out_times.sort()
    used_out: set[int] = set()
    total_hours = 0.0
    unmatched = 0

    for in_dt in in_times:
        match_idx = None
        match_out = None
        for idx, out_dt in enumerate(out_times):
            if idx in used_out:
                continue
            candidate = out_dt
            if candidate < in_dt:
                candidate = candidate + timedelta(days=1)
            if candidate >= in_dt:
                match_idx = idx
                match_out = candidate
                break
        if match_idx is None or match_out is None:
            unmatched += 1
            continue
        used_out.add(match_idx)
        hours = (match_out - in_dt).total_seconds() / 3600.0
        if 0 < hours <= 20:
            total_hours += hours
        else:
            unmatched += 1

    unmatched += max(len(out_times) - len(used_out), 0)
    return round(total_hours, 2), unmatched


def _accessible_emp_ids() -> list[int]:
    if g.current_user.role == "admin":
        return [e.id for e in Employee.query.with_entities(Employee.id).all()]
    emp_rows = UserEmployeeAssignment.query.filter_by(user_id=g.current_user.id).all()
    dept_rows = UserDepartmentAssignment.query.filter_by(user_id=g.current_user.id).all()
    ids = {r.emp_id for r in emp_rows}
    dept_ids = [r.dept_id for r in dept_rows]
    if dept_ids:
        dept_emp_ids = Employee.query.with_entities(Employee.id).filter(Employee.dept_id.in_(dept_ids)).all()
        ids.update(row.id for row in dept_emp_ids)
    return list(ids)


def _non_manager_emp_ids(emp_ids: list[int]) -> list[int]:
    if not emp_ids:
        return []
    rows = (
        Employee.query.with_entities(Employee.id)
        .filter(Employee.id.in_(emp_ids), Employee.is_manager.is_(False))
        .all()
    )
    allowed = {row.id for row in rows}
    return [emp_id for emp_id in emp_ids if emp_id in allowed]


def _manager_emp_ids(emp_ids: list[int]) -> list[int]:
    if not emp_ids:
        return []
    rows = (
        Employee.query.with_entities(Employee.id)
        .filter(Employee.id.in_(emp_ids), Employee.is_manager.is_(True))
        .all()
    )
    allowed = {row.id for row in rows}
    return [emp_id for emp_id in emp_ids if emp_id in allowed]


def _manager_options() -> ManagerAttendanceOptions:
    month = _resolve_query_month()
    account_set = AccountSet.query.filter_by(month=month).first()
    return ManagerAttendanceOptions(
        month=month,
        factory_rest_days=(account_set.factory_rest_days if account_set else 0.0) or 0.0,
        monthly_benefit_days=(account_set.monthly_benefit_days if account_set else 0.0) or 0.0,
    )


def _pick_emp_id() -> int | None:
    requested = request.args.get("emp_id", type=int)
    allowed = _non_manager_emp_ids(_accessible_emp_ids())
    if requested and requested in allowed:
        return requested
    return allowed[0] if allowed else None


def _requested_emp_ids() -> list[int]:
    ids: list[int] = []
    for raw in request.args.getlist("emp_ids"):
        for part in str(raw).split(","):
            text = part.strip()
            if text.isdigit():
                ids.append(int(text))
    if ids:
        return ids

    single = request.args.get("emp_id", type=int)
    return [single] if single else []


def _keyword_filtered_emp_ids(base_ids: list[int]) -> list[int]:
    keyword = (request.args.get("emp_keyword") or "").strip()
    if not keyword:
        return base_ids
    if not base_ids:
        return []

    like_kw = f"%{keyword}%"
    rows = (
        Employee.query.with_entities(Employee.id)
        .filter(Employee.id.in_(base_ids))
        .filter((Employee.emp_no.like(like_kw)) | (Employee.name.like(like_kw)))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    matched = [row.id for row in rows]
    if not matched:
        return []
    matched_set = set(matched)
    # keep base_ids order
    return [emp_id for emp_id in base_ids if emp_id in matched_set]


def _pick_emp_ids() -> list[int]:
    requested = _requested_emp_ids()
    allowed = _non_manager_emp_ids(_accessible_emp_ids())
    if requested:
        allowed_set = set(allowed)
        filtered = [emp_id for emp_id in requested if emp_id in allowed_set]
        # keep order while deduplicating
        return _keyword_filtered_emp_ids(list(dict.fromkeys(filtered)))
    return _keyword_filtered_emp_ids(allowed)


def _resolve_query_month() -> str:
    active_set = AccountSet.query.filter_by(is_active=True).first()
    return request.args.get("month") or (active_set.month if active_set else datetime.now().strftime("%Y-%m"))


def _build_final_rows(month: str, emp_ids: list[int]) -> list[list[object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    rows: list[list[object]] = []
    date_range = _month_date_range(month)
    datetime_range = _month_datetime_range(month)

    daily_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT) if date_range else {}

    leave_by_emp: dict[int, list[LeaveRecord]] = defaultdict(list)
    if datetime_range:
        start_dt, end_dt = datetime_range
        leave_records = (
            LeaveRecord.query.filter(LeaveRecord.emp_id.in_(emp_ids))
            .filter(LeaveRecord.start_time < end_dt, LeaveRecord.end_time > start_dt)
            .all()
        )
        for record in leave_records:
            leave_by_emp[record.emp_id].append(record)

    for employee in employees:
        daily_rows = daily_by_emp.get(employee.id, [])
        leave_rows = leave_by_emp.get(employee.id, [])

        leave_count = {"病假": 0, "工伤": 0, "丧假": 0, "事假": 0, "补休（调休）": 0, "婚假": 0}
        leave_days = {"病假": 0.0, "工伤": 0.0, "丧假": 0.0, "事假": 0.0, "补休（调休）": 0.0, "婚假": 0.0}
        for row in leave_rows:
            leave_type = _leave_bucket(row.leave_type)
            if not leave_type:
                continue
            leave_count[leave_type] += 1
            leave_days[leave_type] += _normalized_leave_days(_leave_days_in_month(row, month))

        day_work_stats = [_calc_record_work_hours(r) for r in daily_rows]
        attendance_days = round(sum(_attendance_day_value(r) for r in daily_rows), 2)
        half_days = sum(
            1
            for idx, r in enumerate(daily_rows)
            if _punch_count(r) == 2 and 2 <= (day_work_stats[idx][0]) < 5.1
        )
        work_hours = round(sum(x[0] for x in day_work_stats), 2)

        row = [
            employee.department.dept_name if employee.department else "",
            employee.emp_no,
            employee.name,
            attendance_days,
            leave_count["病假"],
            leave_count["工伤"],
            leave_count["丧假"],
            leave_count["事假"],
            leave_count["补休（调休）"],
            leave_count["婚假"],
            round(leave_days["病假"], 2),
            round(leave_days["工伤"], 2),
            round(leave_days["丧假"], 2),
            round(leave_days["事假"], 2),
            round(leave_days["补休（调休）"], 2),
            round(leave_days["婚假"], 2),
            work_hours,
            half_days,
            "",
        ]
        rows.append(row)

    return rows


def _build_abnormal_rows(month: str, emp_ids: list[int]) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    data: list[dict[str, object]] = []
    date_range = _month_date_range(month)
    daily_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT) if date_range else {}

    for employee in employees:
        daily_rows = daily_by_emp.get(employee.id, [])
        abnormal_dates = {
            r.record_date.isoformat()
            for r in daily_rows
            if r.record_date and _raw_punch_count(r) in {1, 3}
        }
        data.append(
            {
                "dept_name": employee.department.dept_name if employee.department else "",
                "emp_no": employee.emp_no,
                "name": employee.name,
                "abnormal_count": len(abnormal_dates),
            }
        )
    return data


def _build_department_hours_rows(month: str, emp_ids: list[int]) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    totals: dict[str, float] = {}

    for employee in employees:
        dept_name = employee.department.dept_name if employee.department else "未分配部门"
        totals.setdefault(dept_name, 0.0)

    if not employees:
        return []
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
    for employee in employees:
        dept_name = employee.department.dept_name if employee.department else "未分配部门"
        for row in rows_by_emp.get(employee.id, []):
            totals.setdefault(dept_name, 0.0)
            totals[dept_name] += _calc_record_work_hours(row)[0]

    return [{"dept_name": k, "total_hours": round(v, 2)} for k, v in sorted(totals.items(), key=lambda x: x[0])]


def _top_level_department_name(department: Department | None) -> str:
    if not department:
        return "未分配部门"
    cursor = department
    visited: set[int] = set()
    while cursor.parent and cursor.id not in visited:
        visited.add(cursor.id)
        cursor = cursor.parent
    return cursor.dept_name or "未分配部门"


def _build_manager_department_hours_rows(month: str, emp_ids: list[int]) -> list[dict[str, object]]:
    employees = (
        Employee.query.options(joinedload(Employee.department))
        .filter(Employee.id.in_(emp_ids))
        .order_by(Employee.emp_no.asc())
        .all()
    )
    totals: dict[str, float] = {}

    for employee in employees:
        dept_name = _top_level_department_name(employee.department)
        totals.setdefault(dept_name, 0.0)

    if not employees:
        return []
    rows_by_emp = attendance_views_by_employee(month, employees, MANAGER_STATS_CONTEXT)
    for employee in employees:
        dept_name = _top_level_department_name(employee.department)
        for row in rows_by_emp.get(employee.id, []):
            totals.setdefault(dept_name, 0.0)
            if row.source == "manager":
                totals[dept_name] += float(row.actual_hours or 0) / 60.0
            else:
                totals[dept_name] += float(row.actual_hours or 0)

    return [{"dept_name": k, "total_hours": round(v, 2)} for k, v in sorted(totals.items(), key=lambda x: x[0])]


@employee_bp.route("/dashboard")
@page_permission_required("employee_dashboard")
def dashboard():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no).all() if emp_ids else []
    return render_template("dashboard.html", employees=employees)


@employee_bp.route("/manager-query")
@page_permission_required("manager_query")
def manager_query_page():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no).all() if emp_ids else []
    return render_template("manager_query.html", employees=employees)


@employee_bp.route("/manager-overtime-query")
@page_permission_required("manager_overtime_query")
def manager_overtime_query_page():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all() if emp_ids else []
    return render_template("manager_overtime_query.html", employees=employees)


@employee_bp.route("/manager-annual-leave-query")
@page_permission_required("manager_annual_leave_query")
def manager_annual_leave_query_page():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all() if emp_ids else []
    return render_template("manager_annual_leave_query.html", employees=employees)


@employee_bp.route("/manager-department-hours-query")
@page_permission_required("manager_department_hours_query")
def manager_department_hours_query_page():
    return render_template("manager_department_hours_query.html")


@employee_bp.route("/abnormal-query")
@page_permission_required("abnormal_query")
def abnormal_query_page():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no).all() if emp_ids else []
    return render_template("abnormal_query.html", employees=employees)


@employee_bp.route("/department-hours-query")
@page_permission_required("department_hours_query")
def department_hours_query_page():
    return render_template("department_hours_query.html")


@employee_bp.route("/punch-records")
@page_permission_required("punch_records")
def punch_records_page():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no).all() if emp_ids else []
    return render_template("punch_records.html", employees=employees)


@employee_bp.route("/summary-download")
@page_permission_required("summary_download")
def summary_download_page():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    employees = Employee.query.filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no).all() if emp_ids else []
    return render_template("summary_download.html", employees=employees)


@employee_bp.route("/api/account-sets", methods=["GET"])
@login_required
def account_sets_api():
    if g.current_user.role != "admin" and not g.current_user.has_any_page_access(
        (*MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)
    ):
        return jsonify({"error": "Forbidden"}), 403
    rows = AccountSet.query.order_by(AccountSet.month.desc()).all()
    return jsonify(
        [
            {
                "id": r.id,
                "month": r.month,
                "name": r.name,
                "is_active": r.is_active,
                "factory_rest_days": r.factory_rest_days or 0,
                "monthly_benefit_days": r.monthly_benefit_days or 0,
            }
            for r in rows
        ]
    )


@employee_bp.route("/api/departments", methods=["GET"])
@login_required
def departments_api():
    if g.current_user.role != "admin" and not g.current_user.has_any_page_access(
        (*MANAGER_PAGE_PERMISSION_KEYS, *EMPLOYEE_PAGE_PERMISSION_KEYS)
    ):
        return jsonify({"error": "Forbidden"}), 403

    emp_ids = _accessible_emp_ids()
    if not emp_ids:
        return jsonify([])

    dept_ids = {
        row.dept_id
        for row in Employee.query.with_entities(Employee.dept_id)
        .filter(Employee.id.in_(emp_ids), Employee.dept_id.isnot(None))
        .all()
    }
    if not dept_ids:
        return jsonify([])

    all_ids = set(dept_ids)
    cursor_ids = set(dept_ids)
    while cursor_ids:
        parents = (
            Department.query.with_entities(Department.parent_id)
            .filter(Department.id.in_(cursor_ids), Department.parent_id.isnot(None))
            .all()
        )
        next_ids = {row.parent_id for row in parents if row.parent_id and row.parent_id not in all_ids}
        all_ids.update(next_ids)
        cursor_ids = next_ids

    depts = Department.query.filter(Department.id.in_(all_ids)).order_by(Department.dept_name.asc()).all()
    return jsonify(
        [{"id": d.id, "dept_no": d.dept_no, "dept_name": d.dept_name, "parent_id": d.parent_id} for d in depts]
    )


@employee_bp.route("/api/punch-records", methods=["GET"])
@page_permission_required("punch_records")
def punch_records_api():
    emp_ids = _pick_emp_ids()
    dept_id = request.args.get("dept_id", type=int)
    if dept_id:
        dept_emp_ids = {
            e.id for e in Employee.query.with_entities(Employee.id).filter(Employee.dept_id == dept_id).all()
        }
        emp_ids = [x for x in emp_ids if x in dept_emp_ids]
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    employees = Employee.query.options(joinedload(Employee.department)).filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all()
    rows = []
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
    for employee in employees:
        rows.extend(rows_by_emp.get(employee.id, []))
    rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

    return jsonify(
        [
            {
                "date": r.record_date.isoformat() if r.record_date else "",
                "emp_no": r.employee.emp_no if r.employee else "",
                "name": r.employee.name if r.employee else "",
                "dept_name": r.employee.department.dept_name if r.employee and r.employee.department else "",
                "raw_punch_data": _extract_raw_punch_data(r),
                "check_in_times": _format_punch_times(r.check_in_times),
                "check_out_times": _format_punch_times(r.check_out_times),
                "punch_count": _raw_punch_count(r),
                "actual_hours": _calc_record_work_hours(r)[0],
                "late_minutes": r.late_minutes or 0,
                "early_leave_minutes": r.early_leave_minutes or 0,
                "exception_reason": r.exception_reason or "",
            }
            for r in rows
        ]
    )


@employee_bp.route("/api/punch-records/export", methods=["GET"])
@page_permission_required("punch_records")
def punch_records_export_api():
    emp_ids = _pick_emp_ids()
    dept_id = request.args.get("dept_id", type=int)
    if dept_id:
        dept_emp_ids = {
            e.id for e in Employee.query.with_entities(Employee.id).filter(Employee.dept_id == dept_id).all()
        }
        emp_ids = [x for x in emp_ids if x in dept_emp_ids]
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    employees = Employee.query.options(joinedload(Employee.department)).filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all()
    rows = []
    rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
    for employee in employees:
        rows.extend(rows_by_emp.get(employee.id, []))
    rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

    punch_headers = [
        "日期",
        "员工编号",
        "员工姓名",
        "部门",
        "原始打卡数据",
        "上班打卡",
        "下班打卡",
        "打卡次数",
        "实出勤小时",
        "迟到分钟",
        "早退分钟",
        "异常原因",
    ]

    row_data = []
    for r in rows:
        row_data.append(
            [
                r.record_date.isoformat() if r.record_date else "",
                r.employee.emp_no if r.employee else "",
                r.employee.name if r.employee else "",
                r.employee.department.dept_name if r.employee and r.employee.department else "",
                _extract_raw_punch_data(r),
                _format_punch_times(r.check_in_times),
                _format_punch_times(r.check_out_times),
                _raw_punch_count(r),
                _calc_record_work_hours(r)[0],
                r.late_minutes or 0,
                r.early_leave_minutes or 0,
                r.exception_reason or "",
            ]
        )

    headers, row_data = _filter_punch_columns(punch_headers, row_data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "打卡数据查询"
    ws.append(headers)
    for row in row_data:
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"打卡数据查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@employee_bp.route("/api/summary", methods=["GET"])
@page_permission_required("employee_dashboard")
def summary_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify({"error": "No employee assigned"}), 404

    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    year = request.args.get("year", type=int) or datetime.now().year

    monthly = AttendanceService.monthly_summary(emp_id, month)
    yearly = AttendanceService.yearly_summary(emp_id, year)
    deduction = AttendanceService.deduction_calc(emp_id, month)
    annual = AttendanceService.annual_leave_balance(emp_id, year)

    return jsonify(
        {
            "emp_id": emp_id,
            "month": month,
            "year": year,
            "monthly": monthly,
            "yearly": yearly,
            "deduction": deduction,
            "annual_leave": annual,
        }
    )


@employee_bp.route("/api/daily-records", methods=["GET"])
@page_permission_required("employee_dashboard")
def daily_records_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify([])

    month = request.args.get("month") or datetime.now().strftime("%Y-%m")
    employee = Employee.query.get(emp_id)
    rows = attendance_views_by_employee(month, [employee], EMPLOYEE_STATS_CONTEXT).get(emp_id, []) if employee else []
    rows.sort(key=lambda row: row.record_date or date.min, reverse=True)
    return jsonify(
        [
            {
                "date": r.record_date.isoformat(),
                "expected_hours": r.expected_hours,
                "actual_hours": r.actual_hours,
                "absent_hours": r.absent_hours,
                "leave_hours": r.leave_hours,
                "leave_type": r.leave_type,
                "overtime_hours": r.overtime_hours,
                "overtime_type": r.overtime_type,
                "late_minutes": r.late_minutes,
                "early_leave_minutes": r.early_leave_minutes,
                "exception_reason": r.exception_reason,
            }
            for r in rows
        ]
    )


@employee_bp.route("/api/overtime", methods=["GET"])
@page_permission_required("employee_dashboard")
def overtime_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify([])

    rows = OvertimeRecord.query.filter_by(emp_id=emp_id).order_by(OvertimeRecord.start_time.desc()).all()
    return jsonify(
        [
            {
                "overtime_no": r.overtime_no,
                "start_time": r.start_time.isoformat() if r.start_time else None,
                "end_time": r.end_time.isoformat() if r.end_time else None,
                "effective_hours": r.effective_hours,
                "is_weekend": r.is_weekend,
                "is_holiday": r.is_holiday,
                "salary_option": r.salary_option,
                "reason": r.reason,
                "approval_status": r.approval_status,
            }
            for r in rows
        ]
    )


@employee_bp.route("/api/leave", methods=["GET"])
@page_permission_required("employee_dashboard")
def leave_api():
    emp_id = _pick_emp_id()
    if not emp_id:
        return jsonify([])

    rows = LeaveRecord.query.filter_by(emp_id=emp_id).order_by(LeaveRecord.start_time.desc()).all()
    return jsonify(
        [
            {
                "leave_no": r.leave_no,
                "apply_date": r.apply_date.isoformat() if r.apply_date else None,
                "leave_type": r.leave_type,
                "start_time": r.start_time.isoformat() if r.start_time else None,
                "end_time": r.end_time.isoformat() if r.end_time else None,
                "duration": r.duration,
                "reason": r.reason,
                "approval_status": r.approval_status,
            }
            for r in rows
        ]
    )


@employee_bp.route("/api/annual-leave", methods=["GET"])
@page_permission_required("employee_dashboard")
def annual_leave_api():
    emp_id = _pick_emp_id()
    year = request.args.get("year", type=int) or datetime.now().year
    if not emp_id:
        return jsonify({"year": year, "total_days": 0, "used_days": 0, "remaining_days": 0})

    row = AnnualLeave.query.filter_by(emp_id=emp_id, year=year).first()
    if not row:
        return jsonify({"year": year, "total_days": 0, "used_days": 0, "remaining_days": 0})
    return jsonify(
        {
            "year": row.year,
            "total_days": row.total_days,
            "used_days": row.used_days,
            "remaining_days": row.remaining_days,
        }
    )


@employee_bp.route("/api/final-data", methods=["GET"])
@page_permission_required("employee_dashboard")
def final_data_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"headers": [], "rows": [], "error": "No employee assigned"})

    month = _resolve_query_month()
    requested_ids = _requested_emp_ids()
    rows = _build_final_rows(month, emp_ids)
    headers, rows = _filter_final_columns(FINAL_HEADERS, rows)

    return jsonify(
        {
            "headers": headers,
            "rows": rows,
            "month": month,
            "emp_ids": emp_ids,
            "mode": "all" if not requested_ids else ("single" if len(emp_ids) == 1 else "multi"),
        }
    )


@employee_bp.route("/api/final-data/export", methods=["GET"])
@page_permission_required("employee_dashboard")
def final_data_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    rows = _build_final_rows(month, emp_ids)
    headers, rows = _filter_final_columns(FINAL_HEADERS, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤数据查询"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"考勤数据查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@employee_bp.route("/api/summary-download/export", methods=["GET"])
@page_permission_required("summary_download")
def summary_download_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee selected"}), 400

    month = _resolve_query_month()
    sheets_raw = (request.args.get("sheets") or "final,punch").strip()
    include_final = "final" in sheets_raw
    include_punch = "punch" in sheets_raw

    wb = openpyxl.Workbook()

    if include_final:
        final_rows = _build_final_rows(month, emp_ids)
        headers, rows = _filter_final_columns(FINAL_HEADERS, final_rows)
        ws = wb.active
        ws.title = "考勤数据查询"
        ws.append(headers)
        for row in rows:
            ws.append(row)
    else:
        wb.remove(wb.active)

    if include_punch:
        employees = Employee.query.options(joinedload(Employee.department)).filter(Employee.id.in_(emp_ids)).order_by(Employee.emp_no.asc()).all()
        punch_rows = []
        rows_by_emp = attendance_views_by_employee(month, employees, EMPLOYEE_STATS_CONTEXT)
        for employee in employees:
            punch_rows.extend(rows_by_emp.get(employee.id, []))
        punch_rows.sort(key=lambda row: ((row.employee.emp_no if row.employee else ""), row.record_date or date.min), reverse=True)

        punch_headers = [
            "日期", "员工编号", "员工姓名", "部门", "原始打卡数据",
            "上班打卡", "下班打卡", "打卡次数", "实出勤小时",
            "迟到分钟", "早退分钟", "异常原因",
        ]

        punch_row_data = []
        for r in punch_rows:
            punch_row_data.append([
                r.record_date.isoformat() if r.record_date else "",
                r.employee.emp_no if r.employee else "",
                r.employee.name if r.employee else "",
                r.employee.department.dept_name if r.employee and r.employee.department else "",
                _extract_raw_punch_data(r),
                _format_punch_times(r.check_in_times),
                _format_punch_times(r.check_out_times),
                _raw_punch_count(r),
                _calc_record_work_hours(r)[0],
                r.late_minutes or 0,
                r.early_leave_minutes or 0,
                r.exception_reason or "",
            ])

        punch_headers, punch_row_data = _filter_punch_columns(punch_headers, punch_row_data)

        ws2 = wb.create_sheet("打卡数据查询")
        ws2.append(punch_headers)
        for row in punch_row_data:
            ws2.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"汇总下载_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@employee_bp.route("/api/manager-attendance", methods=["GET"])
@page_permission_required("manager_query")
def manager_attendance_api():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    options = _manager_options()
    rows = build_manager_rows(options, emp_ids)
    return jsonify(
        {
            "headers": MANAGER_HEADERS,
            "rows": rows_as_table(rows),
            "month": options.month,
            "factory_rest_days": options.factory_rest_days,
            "monthly_benefit_days": options.monthly_benefit_days,
        }
    )


@employee_bp.route("/api/manager-overtime-query", methods=["GET"])
@page_permission_required("manager_overtime_query")
def manager_overtime_query_api():
    from routes.admin import _manager_month_rows, _manager_overtime_values

    year = request.args.get("year", type=int) or datetime.now().year
    values = _manager_overtime_values(year)
    allowed_ids = set(_manager_emp_ids(_accessible_emp_ids()))
    values = {name: row for name, row in values.items() if row.get("emp_id") in allowed_ids}
    requested_ids = set(_requested_emp_ids())
    if requested_ids:
        values = {name: row for name, row in values.items() if row.get("emp_id") in requested_ids}
    return jsonify(
        {
            "year": year,
            "headers": ["部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"],
            "rows": _manager_month_rows(values, "剩余调休天数"),
        }
    )


@employee_bp.route("/api/manager-annual-leave-query", methods=["GET"])
@page_permission_required("manager_annual_leave_query")
def manager_annual_leave_query_api():
    from routes.admin import _annual_leave_value_keys, _manager_annual_leave_values, _manager_month_rows

    year = request.args.get("year", type=int) or datetime.now().year
    values = _manager_annual_leave_values(year)
    allowed_ids = set(_manager_emp_ids(_accessible_emp_ids()))
    values = {name: row for name, row in values.items() if row.get("emp_id") in allowed_ids}
    requested_ids = set(_requested_emp_ids())
    if requested_ids:
        values = {name: row for name, row in values.items() if row.get("emp_id") in requested_ids}
    return jsonify(
        {
            "year": year,
            "headers": ["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"],
            "rows": _manager_month_rows(values, "剩余年休天数", _annual_leave_value_keys()),
        }
    )


@employee_bp.route("/api/manager-department-hours", methods=["GET"])
@page_permission_required("manager_department_hours_query")
def manager_department_hours_api():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    return jsonify(_build_manager_department_hours_rows(month, emp_ids))


@employee_bp.route("/api/manager-department-hours/export", methods=["GET"])
@page_permission_required("manager_department_hours_query")
def manager_department_hours_export_api():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    if not emp_ids:
        return jsonify({"error": "No manager assigned"}), 400

    month = _resolve_query_month()
    rows = _build_manager_department_hours_rows(month, emp_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "管理人员部门工时查询"
    ws.append(["部门名称", "总工时（小时）"])
    for row in rows:
        ws.append([row.get("dept_name", ""), row.get("total_hours", 0)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员部门工时查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _fill_manager_template(ws, rows: list[dict[str, object]], include_actual_attendance_days: bool = True) -> None:
    headers = manager_headers(include_actual_attendance_days)

    if include_actual_attendance_days and ws.max_column == len(MANAGER_HEADERS) - 2 and ws.cell(1, 4).value == "事/病假":
        ws.insert_cols(4)
        for row_idx in range(1, ws.max_row + 1):
            source = ws.cell(row_idx, 3)
            target = ws.cell(row_idx, 4)
            target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
    funeral_col_idx = 9 if include_actual_attendance_days else 8
    late_col_idx = funeral_col_idx + 1
    if ws.max_column == len(headers) - 1 and ws.cell(1, funeral_col_idx).value == "迟到\\早退":
        ws.insert_cols(funeral_col_idx)
        for row_idx in range(1, ws.max_row + 1):
            source = ws.cell(row_idx, late_col_idx)
            target = ws.cell(row_idx, funeral_col_idx)
            target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
    ws.column_dimensions["K" if include_actual_attendance_days else "J"].width = 12.6666666666667
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx).value = header

    by_name = {str(row.get("name", "")).strip(): row for row in rows}
    filled_names: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row_idx, 2).value or "").strip()
        if not name or name not in by_name:
            continue
        item = by_name[name]
        values = [
            item.get("attendance_days", 0),
            *([item.get("actual_attendance_days", 0)] if include_actual_attendance_days else []),
            item.get("personal_sick_days", 0),
            item.get("injury_days", 0),
            item.get("business_trip_days", 0),
            item.get("marriage_days", 0),
            item.get("funeral_days", 0),
            item.get("late_early_minutes", 0),
            item.get("summary", ""),
            item.get("benefit_days", 0),
            item.get("overtime_change", 0),
            item.get("remark", ""),
        ]
        for offset, value in enumerate(values, start=3):
            ws.cell(row_idx, offset).value = value
        filled_names.add(name)

    for item in rows:
        name = str(item.get("name", "")).strip()
        if not name or name in filled_names:
            continue
        ws.append(
            [
                item.get("dept_name", ""),
                name,
                item.get("attendance_days", 0),
                *([item.get("actual_attendance_days", 0)] if include_actual_attendance_days else []),
                item.get("personal_sick_days", 0),
                item.get("injury_days", 0),
                item.get("business_trip_days", 0),
                item.get("marriage_days", 0),
                item.get("funeral_days", 0),
                item.get("late_early_minutes", 0),
                item.get("summary", ""),
                item.get("benefit_days", 0),
                item.get("overtime_change", 0),
                item.get("remark", ""),
            ]
        )


@employee_bp.route("/api/manager-attendance/export", methods=["GET"])
@page_permission_required("manager_query")
def manager_attendance_export_api():
    emp_ids = _manager_emp_ids(_accessible_emp_ids())
    requested_ids = _requested_emp_ids()
    if requested_ids:
        allowed = set(emp_ids)
        emp_ids = [emp_id for emp_id in requested_ids if emp_id in allowed]
    options = _manager_options()
    rows = build_manager_rows(options, emp_ids)
    include_actual_attendance_days = request.args.get("show_actual_attendance_days") == "1"

    template_path = "/home/lewis/文档/考勤/管理人员最终输出输出模板.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        _fill_manager_template(ws, rows, include_actual_attendance_days)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "管理人员查询"
        ws.append(manager_headers(include_actual_attendance_days))
        for row in rows_as_table(rows, include_actual_attendance_days):
            ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员考勤查询_{options.month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@employee_bp.route("/api/abnormal-attendance", methods=["GET"])
@page_permission_required("abnormal_query")
def abnormal_attendance_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    return jsonify(_build_abnormal_rows(month, emp_ids))


@employee_bp.route("/api/abnormal-attendance/export", methods=["GET"])
@page_permission_required("abnormal_query")
def abnormal_attendance_export_api():
    emp_ids = _pick_emp_ids()
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    rows = _build_abnormal_rows(month, emp_ids)

    headers = ["部门名称", "人员编号", "人员姓名", "异常考勤次数"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工异常查询"
    ws.append(headers)
    for r in rows:
        ws.append([r.get("dept_name", ""), r.get("emp_no", ""), r.get("name", ""), r.get("abnormal_count", 0)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"员工异常查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@employee_bp.route("/api/department-hours", methods=["GET"])
@page_permission_required("department_hours_query")
def department_hours_api():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    if not emp_ids:
        return jsonify([])

    month = _resolve_query_month()
    return jsonify(_build_department_hours_rows(month, emp_ids))


@employee_bp.route("/api/department-hours/export", methods=["GET"])
@page_permission_required("department_hours_query")
def department_hours_export_api():
    emp_ids = _non_manager_emp_ids(_accessible_emp_ids())
    if not emp_ids:
        return jsonify({"error": "No employee assigned"}), 400

    month = _resolve_query_month()
    rows = _build_department_hours_rows(month, emp_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工部门工时查询"
    ws.append(["部门名称", "总工时（小时）"])
    for row in rows:
        ws.append([row.get("dept_name", ""), row.get("total_hours", 0)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"员工部门工时查询_{month}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
