from __future__ import annotations

import logging
import os
import subprocess
from collections import defaultdict
from io import BytesIO
from datetime import date, datetime
from typing import Any

from flask import abort, jsonify, request, send_file, g
import openpyxl
from sqlalchemy.orm import joinedload, selectinload

logger = logging.getLogger(__name__)

from models import db
from models.department import Department
from models.employee import Employee
from models.employee import (
    ATTENDANCE_SOURCE_AUTO_FALLBACK,
    ATTENDANCE_SOURCE_EMPLOYEE,
    ATTENDANCE_SOURCE_MANAGER,
    ATTENDANCE_SOURCE_VALUES,
)
from models.employee_shift import EmployeeShiftAssignment
from models.shift import Shift
from models.daily_record import DailyRecord
from models.account_set import AccountSet, AccountSetFactoryRestDay, AccountSetImport
from models.overtime import OvertimeRecord
from models.annual_leave import AnnualLeave
from models.manager_month_stat import ManagerMonthStat
from models.manager_attendance_override import ManagerAttendanceOverride
from models.employee_attendance_override import EmployeeAttendanceOverride
from models.attendance_override_history import AttendanceOverrideHistory
from models.user import (
    ALL_PAGE_PERMISSION_KEYS,
    EMPLOYEE_PAGE_PERMISSION_KEYS,
    HOME_PAGE_PERMISSION_KEYS,
    MANAGER_PAGE_PERMISSION_KEYS,
    PAGE_PERMISSION_LABELS,
    User,
    UserEmployeeAssignment,
    UserDepartmentAssignment,
)
from services.import_service import ImportService
from services.calculation_progress_service import get_calc_progress, update_calc_progress
from services.manager_attendance_service import ManagerAttendanceOptions, build_manager_rows
from utils.helpers import parse_bool_zh




_EMPLOYEE_OVERRIDE_FIELDS = (
    "attendance_days",
    "actual_attendance_days",
    "work_hours",
    "half_days",
    "late_early_minutes",
)

_EMPLOYEE_OVERRIDE_LABELS = {
    "attendance_days": "考勤天数",
    "actual_attendance_days": "实际出勤天数",
    "work_hours": "工时",
    "half_days": "半勤天数",
    "late_early_minutes": "迟到早退",
    "remark": "备注",
}

_MANAGER_ATTENDANCE_OVERRIDE_FIELDS = (
    "attendance_days",
    "injury_days",
    "business_trip_days",
    "marriage_days",
    "funeral_days",
    "late_early_minutes",
)

_MANAGER_OVERRIDE_LABELS = {
    "attendance_days": "出勤天数",
    "injury_days": "工伤",
    "business_trip_days": "出差",
    "marriage_days": "婚假",
    "funeral_days": "丧假",
    "late_early_minutes": "迟到早退",
    "remark": "备注",
}

_DAILY_OVERRIDE_LABELS = {
    "record_date": "日期",
    "status": "考勤状态",
    "is_evening_overtime": "晚上加班",
    "work_hours": "工时",
    "late_minutes": "迟到分钟",
    "early_leave_minutes": "早退分钟",
    "remark": "备注",
}

_DEPARTMENT_ORIGINAL_ID_HEADER = "原始部门ID"
_DEPARTMENT_ORIGINAL_DEPT_NO_HEADER = "原始部门编号"
_DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER = "原始部门名称"
_DEPARTMENT_METADATA_SHEET = "部门导入元数据"
_DEPARTMENT_IMPORT_TEMP_PREFIX = "__IMPORT_TMP__"


def _require_model(model, ident):
    row = db.session.get(model, ident)
    if row is None:
        abort(404)
    return row


def _requested_emp_ids() -> list[int]:
    raw_values = request.args.getlist("emp_ids")
    if not raw_values:
        raw = (request.args.get("emp_ids") or "").strip()
        raw_values = raw.split(",") if raw else []
    seen: set[int] = set()
    emp_ids: list[int] = []
    for raw in raw_values:
        for part in str(raw or "").split(","):
            text = part.strip()
            if not text or not text.isdigit():
                continue
            emp_id = int(text)
            if emp_id in seen:
                continue
            seen.add(emp_id)
            emp_ids.append(emp_id)
    return emp_ids


def _manager_scope_employees(emp_ids: list[int] | None = None):
    query = Employee.query.filter(Employee.is_manager.is_(True))
    if emp_ids is not None:
        if not emp_ids:
            return []
        query = query.filter(Employee.id.in_(emp_ids))
    return _unique_employees(
        query.order_by(Employee.dept_id.asc(), Employee.emp_no.asc(), Employee.name.asc()).all()
    )


def _unique_employees(rows):
    unique_rows = []
    seen_ids = set()
    for row in rows or []:
        if not row or row.id in seen_ids:
            continue
        seen_ids.add(row.id)
        unique_rows.append(row)
    return unique_rows
def _accessible_dept_ids_set() -> set[int]:
    if getattr(g, "current_user", None) is None:
        return set()
    
    if g.current_user.role == "admin":
        return {d.id for d in Department.query.with_entities(Department.id).all()}
        
    dept_rows = UserDepartmentAssignment.query.filter_by(user_id=g.current_user.id).all()
    assigned_dept_ids = {r.dept_id for r in dept_rows}
    
    if not assigned_dept_ids:
        return set()
        
    all_departments = Department.query.all()
    dept_by_id = {d.id: d for d in all_departments}
    
    result_ids = set()
    for d_id in assigned_dept_ids:
        curr = dept_by_id.get(d_id)
        while curr:
            result_ids.add(curr.id)
            if curr.parent_id:
                curr = dept_by_id.get(curr.parent_id)
            else:
                curr = None
                
    return result_ids


def _accessible_emp_ids_set() -> set[int]:
    # 可见范围计算的唯一实现在 query_core._accessible_emp_ids（含部门树展开），
    # 双份实现易漂移导致越权/漏数据；此处仅按 set 口径薄包装。
    # 函数内 import 以维持两路由模块间零模块级依赖（避免循环导入）。
    from routes.query_core import _accessible_emp_ids

    return set(_accessible_emp_ids())



def _serialize_user(user: User, profile_departments_by_id: dict[int, Department] | None = None) -> dict:
    profile_department = None
    if user.profile_dept_id:
        if profile_departments_by_id is not None:
            profile_department = profile_departments_by_id.get(user.profile_dept_id)
        else:
            profile_department = db.session.get(Department, user.profile_dept_id)
    return {
        "id": user.id,
        "username": user.username,
        "profile_emp_no": user.profile_emp_no or "",
        "profile_name": user.profile_name or "",
        "profile_dept_id": user.profile_dept_id,
        "profile_department": {
            "id": profile_department.id,
            "dept_no": profile_department.dept_no,
            "dept_name": profile_department.dept_name,
        } if profile_department else None,
        "role": user.role,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "login_failed_attempts": int(user.login_failed_attempts or 0),
        "login_locked_until": user.login_locked_until.isoformat() if user.login_locked_until else None,
        "login_disabled_until_admin_unlock": bool(user.login_disabled_until_admin_unlock),
        "login_disabled_reason": user.login_disabled_reason,
        "page_permissions": user.effective_page_permissions(),
        "emp_ids": [a.emp_id for a in user.employee_assignments],
        "dept_ids": [a.dept_id for a in user.department_assignments],
        "employees": [
            {
                "id": a.employee.id,
                "emp_no": a.employee.emp_no,
                "name": a.employee.name,
                "dept_id": a.employee.dept_id,
                "dept_no": a.employee.department.dept_no if a.employee.department else "",
                "dept_name": a.employee.department.dept_name if a.employee.department else "",
            }
            for a in user.employee_assignments
            if a.employee
        ],
        "departments": [
            {
                "id": a.department.id,
                "dept_no": a.department.dept_no,
                "dept_name": a.department.dept_name,
                "parent_id": a.department.parent_id,
            }
            for a in user.department_assignments
            if a.department
        ],
    }


def _user_list_query():
    return (
        User.query.options(
            selectinload(User.employee_assignments)
            .joinedload(UserEmployeeAssignment.employee)
            .joinedload(Employee.department),
            selectinload(User.department_assignments).joinedload(UserDepartmentAssignment.department),
        )
        .order_by(User.id.desc())
    )


def _bind_user_profile_identity(user: User, emp_ids: list[int]) -> None:
    if user.profile_emp_no and user.profile_name and user.profile_dept_id:
        return

    first_emp_id = next((emp_id for emp_id in emp_ids if isinstance(emp_id, int)), None)
    if first_emp_id is None:
        return

    employee = db.session.get(Employee, first_emp_id)
    if not employee:
        return

    user.profile_emp_no = employee.emp_no or user.profile_emp_no
    user.profile_name = employee.name or user.profile_name
    user.profile_dept_id = employee.dept_id or user.profile_dept_id


def _default_page_permissions_for_role(role: str) -> dict[str, bool]:
    if role == "admin":
        return {key: True for key in ALL_PAGE_PERMISSION_KEYS}
    return {key: key in HOME_PAGE_PERMISSION_KEYS for key in ALL_PAGE_PERMISSION_KEYS}


def _parse_page_permissions(data: dict | None, role: str, existing_user: User | None = None) -> dict[str, bool]:
    if role == "admin":
        return {key: True for key in ALL_PAGE_PERMISSION_KEYS}

    if not data or "page_permissions" not in data:
        if existing_user and isinstance(existing_user.page_permissions, dict):
            return existing_user.effective_page_permissions()
        return _default_page_permissions_for_role(role)

    raw = data.get("page_permissions") or {}
    if not isinstance(raw, dict):
        return _default_page_permissions_for_role(role)
    return {key: bool(raw.get(key, False)) for key in ALL_PAGE_PERMISSION_KEYS}


def _sync_user_assignments(user: User, emp_ids: list[int]) -> None:
    valid_ids = {
        row.id for row in Employee.query.with_entities(Employee.id).filter(Employee.id.in_(emp_ids)).all()
    }
    current_ids = {a.emp_id for a in user.employee_assignments}
    to_remove = [a for a in user.employee_assignments if a.emp_id not in valid_ids]
    for assignment in to_remove:
        db.session.delete(assignment)
    for emp_id in valid_ids - current_ids:
        db.session.add(UserEmployeeAssignment(user_id=user.id, emp_id=emp_id))


def _sync_user_department_assignments(user: User, dept_ids: list[int]) -> None:
    valid_ids = {
        row.id for row in Department.query.with_entities(Department.id).filter(Department.id.in_(dept_ids)).all()
    }
    current_ids = {a.dept_id for a in user.department_assignments}
    to_remove = [a for a in user.department_assignments if a.dept_id not in valid_ids]
    for assignment in to_remove:
        db.session.delete(assignment)
    for dept_id in valid_ids - current_ids:
        db.session.add(UserDepartmentAssignment(user_id=user.id, dept_id=dept_id))


def _serialize_employee(employee: Employee) -> dict:
    shift = employee.shift_assignment.shift if employee.shift_assignment else None
    return {
        "id": employee.id,
        "emp_no": employee.emp_no,
        "name": employee.name,
        "card_no": employee.card_no or None,
        "is_manager": bool(employee.is_manager),
        "is_nursing": bool(employee.is_nursing),
        "employee_stats_attendance_source": employee.employee_stats_attendance_source or ATTENDANCE_SOURCE_EMPLOYEE,
        "manager_stats_attendance_source": employee.manager_stats_attendance_source or ATTENDANCE_SOURCE_MANAGER,
        "dept_id": employee.dept_id,
        "dept_no": employee.department.dept_no if employee.department else "",
        "dept_name": employee.department.dept_name if employee.department else "",
        "department": employee.department.dept_name if employee.department else "",
        "shift_id": shift.id if shift else None,
        "shift_no": shift.shift_no if shift else "",
        "shift_name": shift.shift_name if shift else "",
        "resigned_at": employee.resigned_at.isoformat() if employee.resigned_at else None,
    }


def _parse_attendance_source(value: Any, default: str) -> str:
    source = str(value or "").strip() or default
    label_map = {
        "员工考勤源文件取值": ATTENDANCE_SOURCE_EMPLOYEE,
        "管理人员考勤源文件取值": ATTENDANCE_SOURCE_MANAGER,
        "自动回退": ATTENDANCE_SOURCE_AUTO_FALLBACK,
    }
    source = label_map.get(source, source)
    if source not in ATTENDANCE_SOURCE_VALUES:
        return default
    return source


def _parse_department_original_id(value: Any) -> int | None:
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    try:
        return int(float(raw_value))
    except (TypeError, ValueError, OverflowError):
        return None


def _load_department_identity_metadata(wb: openpyxl.Workbook) -> dict[int, dict[str, str]]:
    if _DEPARTMENT_METADATA_SHEET not in wb.sheetnames:
        return {}

    metadata_ws = wb[_DEPARTMENT_METADATA_SHEET]
    metadata_rows = [list(r) for r in metadata_ws.iter_rows(values_only=True)]
    if not metadata_rows:
        return {}

    header_idx, header_map = _parse_header_row(
        metadata_rows,
        [
            _DEPARTMENT_ORIGINAL_ID_HEADER,
            _DEPARTMENT_ORIGINAL_DEPT_NO_HEADER,
            _DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER,
        ],
    )
    original_id_idx = header_map.get(_DEPARTMENT_ORIGINAL_ID_HEADER, -1)
    original_dept_no_idx = header_map.get(_DEPARTMENT_ORIGINAL_DEPT_NO_HEADER, -1)
    original_dept_name_idx = header_map.get(_DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER, -1)
    if original_id_idx < 0 or original_dept_no_idx < 0 or original_dept_name_idx < 0:
        return {}

    identities_by_id: dict[int, dict[str, str]] = {}
    for row in metadata_rows[header_idx + 1 :]:
        original_id = (
            _parse_department_original_id(row[original_id_idx])
            if original_id_idx < len(row)
            else None
        )
        if original_id is None:
            continue
        original_dept_no = (
            str(row[original_dept_no_idx]).strip()
            if original_dept_no_idx < len(row) and row[original_dept_no_idx] is not None
            else ""
        )
        original_dept_name = (
            str(row[original_dept_name_idx]).strip()
            if original_dept_name_idx < len(row) and row[original_dept_name_idx] is not None
            else ""
        )
        if not original_dept_no or not original_dept_name:
            continue
        identities_by_id[original_id] = {
            "dept_no": original_dept_no,
            "dept_name": original_dept_name,
        }
    return identities_by_id


def _department_matches_original_identity(
    department: Department | None,
    original_identity: dict[str, str] | None,
) -> bool:
    if not department or not original_identity:
        return False
    return (
        (department.dept_no or "").strip() == original_identity["dept_no"]
        and (department.dept_name or "").strip() == original_identity["dept_name"]
    )


def _factory_rest_unit(period: str) -> float:
    if period == "full":
        return 1.0
    if period in {"am", "pm"}:
        return 0.5
    raise ValueError(f"Unsupported factory rest period: {period}")


def _effective_factory_rest_days(account_set: AccountSet | None) -> float:
    if account_set is None:
        return 0

    if not account_set.factory_rest_entries:
        return (account_set.factory_rest_days or 0)

    return sum(_factory_rest_unit(item.rest_period) for item in account_set.factory_rest_entries)


def _manager_factory_rest_days(account_set: AccountSet | None) -> float:
    if account_set is None:
        return 0
    if not account_set.factory_rest_entries:
        return 0
    return sum(_factory_rest_unit(item.rest_period) for item in account_set.factory_rest_entries)


def _manager_factory_rest_requires_detail(account_set: AccountSet | None) -> bool:
    if account_set is None:
        return False
    return not account_set.factory_rest_entries and float(account_set.factory_rest_days or 0) > 0


def _parse_factory_rest_entries(entries: object, month: str) -> list[dict[str, object]]:
    if entries is None:
        return []
    if not isinstance(entries, list):
        raise ValueError("厂休明细格式不正确")

    normalized: list[dict[str, object]] = []
    seen_dates: set[str] = set()
    month_prefix = f"{month}-"
    for item in entries:
        if not isinstance(item, dict):
            raise ValueError("厂休明细格式不正确")
        date_text = str(item.get("date") or "").strip()
        period = str(item.get("period") or "").strip()
        if not date_text.startswith(month_prefix):
            raise ValueError("厂休日期必须属于当前账套月份")
        try:
            rest_date = datetime.strptime(date_text, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError("厂休日期格式不正确") from exc
        try:
            _factory_rest_unit(period)
        except ValueError as exc:
            raise ValueError("厂休时段仅支持 full/am/pm") from exc
        if rest_date.isoformat() in seen_dates:
            raise ValueError("同一天只能设置一个厂休时段")
        seen_dates.add(rest_date.isoformat())
        normalized.append({"date": rest_date, "period": period})
    return normalized


def _replace_factory_rest_entries(account_set: AccountSet, entries: list[dict[str, object]]) -> float:
    account_set.factory_rest_entries.clear()
    total = 0.0
    for item in entries:
        period = str(item["period"])
        account_set.factory_rest_entries.append(
            AccountSetFactoryRestDay(
                rest_date=item["date"],
                rest_period=period,
            )
        )
        total += _factory_rest_unit(period)
    account_set.factory_rest_days = total
    return total


def _validate_initial_factory_rest_entry_backfill(
    account_set: AccountSet,
    entries: list[dict[str, object]],
) -> None:
    if account_set.factory_rest_entries:
        return

    legacy_days = float(account_set.factory_rest_days or 0)
    if legacy_days <= 0:
        return

    entry_total = sum(_factory_rest_unit(str(item["period"])) for item in entries)
    if round(entry_total, 2) == round(legacy_days, 2):
        return

    raise ValueError(
        f"首次补录厂休明细时，明细汇总必须等于原厂休天数 {legacy_days:g} 天，请一次性补齐全部明细后再保存"
    )


def _serialize_factory_rest_entry(row: AccountSetFactoryRestDay) -> dict:
    return {
        "date": row.rest_date.isoformat() if row.rest_date else None,
        "period": row.rest_period,
        "unit": _factory_rest_unit(row.rest_period),
    }


def _serialize_account_set(row: AccountSet) -> dict:
    success_count = 0
    error_count = 0
    pending_count = 0
    latest_import_at = None
    for item in row.imports:
        if item.status == "ok":
            success_count += 1
        elif item.status == "error":
            error_count += 1
        else:
            pending_count += 1
        if item.created_at and (latest_import_at is None or item.created_at > latest_import_at):
            latest_import_at = item.created_at

    factory_rest_entries = sorted(
        row.factory_rest_entries,
        key=lambda item: (
            item.rest_date.isoformat() if item.rest_date else "",
            item.rest_period or "",
        ),
    )
    serialized_factory_rest_entries = [_serialize_factory_rest_entry(item) for item in factory_rest_entries]

    return {
        "id": row.id,
        "month": row.month,
        "name": row.name,
        "is_active": row.is_active,
        "is_locked": bool(row.is_locked),
        "locked_at": row.locked_at.isoformat() if row.locked_at else None,
        "locked_by": row.locked_by,
        "factory_rest_days": _effective_factory_rest_days(row),
        "factory_rest_entries": serialized_factory_rest_entries,
        "monthly_benefit_days": row.monthly_benefit_days or 0,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "imports_count": len(row.imports),
        "pending_count": pending_count,
        "success_count": success_count,
        "error_count": error_count,
        "latest_import_at": latest_import_at.isoformat() if latest_import_at else None,
    }


def _account_set_file_type(filename: str) -> str:
    if "加班" in filename:
        return "overtime"
    if "请假" in filename:
        return "leave"
    if "管理人员" in filename and "月报" in filename:
        return "manager_monthly"
    if "管理人员" in filename:
        return "manager_daily"
    if "月报" in filename:
        return "monthly"
    return "daily"


def _account_set_for_month(month: str | None) -> AccountSet | None:
    key = (month or "").strip()
    if not key:
        return None
    # 请求级缓存：同请求内同月账套重复查询（修正列表曾放大到 2N 次）直接复用；
    # 写账套的接口均为写完即返回，无写后再读路径
    cache = getattr(g, "_account_set_by_month", None)
    if cache is None:
        cache = {}
        g._account_set_by_month = cache
    if key in cache:
        return cache[key]
    result = AccountSet.query.filter_by(month=key).first()
    cache[key] = result
    return result


def _locked_account_set_error(account_set: AccountSet, action_label: str):
    return jsonify({"error": f"{account_set.month} 账套已锁定，不能{action_label}"}), 400


def _ensure_account_set_unlocked(account_set: AccountSet | None, action_label: str):
    if account_set and account_set.is_locked:
        return _locked_account_set_error(account_set, action_label)
    return None


def _user_display_name(user: User | None) -> str:
    return user.username if user else ""


def _override_field_labels(override_type: str) -> dict[str, str]:
    if override_type == "employee":
        return _EMPLOYEE_OVERRIDE_LABELS
    if override_type == "daily":
        return _DAILY_OVERRIDE_LABELS
    return _MANAGER_OVERRIDE_LABELS



def _override_state_from_row(row: object | None, fields: tuple[str, ...]) -> dict[str, object]:
    state = {field: getattr(row, field) if row else None for field in fields}
    state["remark"] = row.remark if row else ""
    return state


def _has_override_state_changes(before: dict[str, object], after: dict[str, object]) -> bool:
    keys = set(before.keys()) | set(after.keys())
    return any(before.get(key) != after.get(key) for key in keys)


def _changed_override_fields(before: dict[str, object], after: dict[str, object]) -> list[str]:
    keys = [key for key in after.keys() if key in before or key == "remark"]
    return [key for key in keys if before.get(key) != after.get(key)]


def _record_override_history(
    override_type: str,
    emp_id: int,
    month: str,
    action_type: str,
    before_values: dict[str, object],
    after_values: dict[str, object],
    source_file_name: str | None = None,
) -> None:
    history = AttendanceOverrideHistory(
        override_type=override_type,
        emp_id=emp_id,
        month=month,
        action_type=action_type,
        changed_fields_json=_changed_override_fields(before_values, after_values),
        before_values_json=before_values,
        after_values_json=after_values,
        remark=str(after_values.get("remark") or ""),
        source_file_name=(source_file_name or "").strip() or None,
        operator_user_id=g.current_user.id if getattr(g, "current_user", None) else None,
    )
    db.session.add(history)


def _serialize_override_history(row: AttendanceOverrideHistory) -> dict[str, object]:
    labels = _override_field_labels(row.override_type)
    before = row.before_values_json if isinstance(row.before_values_json, dict) else {}
    after = row.after_values_json if isinstance(row.after_values_json, dict) else {}
    changed_fields = list(row.changed_fields_json or [])
    changes = [
        {
            "field": field,
            "label": labels.get(field, field),
            "before": before.get(field),
            "after": after.get(field),
        }
        for field in changed_fields
    ]
    return {
        "id": row.id,
        "override_type": row.override_type,
        "action_type": row.action_type,
        "month": row.month,
        "emp_id": row.emp_id,
        "emp_no": row.employee.emp_no if row.employee else "",
        "employee_name": row.employee.name if row.employee else "",
        "remark": row.remark or "",
        "source_file_name": row.source_file_name or "",
        "operator_name": _user_display_name(row.operator_user),
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "changes": changes,
    }



def _history_rows_for_month(override_type: str, month: str) -> list[dict[str, object]]:
    rows = (
        AttendanceOverrideHistory.query.filter_by(override_type=override_type, month=month)
        .order_by(AttendanceOverrideHistory.created_at.desc(), AttendanceOverrideHistory.id.desc())
        .all()
    )
    return [_serialize_override_history(row) for row in rows]


def _locked_months_for_year(year: int, include_prev_dec: bool = False) -> list[str]:
    months = [f"{year}-{month:02d}" for month in range(1, 13)]
    if include_prev_dec:
        months.insert(0, f"{year - 1}-12")
    account_sets = AccountSet.query.filter(AccountSet.month.in_(months)).all()
    locked_months = {row.month for row in account_sets if row.is_locked}
    return [month for month in months if month in locked_months]


def _ensure_year_months_unlocked(year: int, action_label: str, include_prev_dec: bool = False):
    locked_months = _locked_months_for_year(year, include_prev_dec=include_prev_dec)
    if locked_months:
        return jsonify({"error": f"{', '.join(locked_months)} 账套已锁定，不能{action_label}"}), 400
    return None


def _stat_key_month(year: int, stat_key: str) -> str | None:
    if stat_key == "prev_dec":
        return f"{year - 1}-12"
    if stat_key.startswith("m") and stat_key[1:].isdigit():
        month_no = int(stat_key[1:])
        if 1 <= month_no <= 12:
            return f"{year}-{month_no:02d}"
    return None


def _stat_key_lock_state(year: int, stat_key: str) -> str:
    month = _stat_key_month(year, stat_key)
    if not month:
        return "editable"
    account_set = _account_set_for_month(month)
    if not account_set:
        return "missing_account_set"
    return "locked" if account_set.is_locked else "editable"


def _next_auto_dept_no() -> str:
    index = Department.query.count() + 1
    while True:
        dept_no = f"AUTO-{index:04d}"
        if not Department.query.filter_by(dept_no=dept_no).first():
            return dept_no
        index += 1


def _resolve_department(dept_name: str | None, dept_no: str | None = None) -> Department | None:
    clean_name = (dept_name or "").strip()
    clean_no = (dept_no or "").strip()
    if not clean_name:
        return None

    if clean_no:
        existing_by_no = Department.query.filter_by(dept_no=clean_no).first()
        if existing_by_no:
            return existing_by_no
        department = Department(dept_no=clean_no, dept_name=clean_name)
        db.session.add(department)
        db.session.flush()
        return department

    existing_by_name = Department.query.filter_by(dept_name=clean_name).first()
    if existing_by_name:
        return existing_by_name

    department = Department(dept_no=_next_auto_dept_no(), dept_name=clean_name)
    db.session.add(department)
    db.session.flush()
    return department


def _parse_parent_id(raw: Any) -> int | None:
    if raw in (None, "", "null", "None"):
        return None
    if isinstance(raw, int):
        return raw
    text_value = str(raw).strip()
    if not text_value:
        return None
    if not text_value.isdigit():
        return -1
    return int(text_value)


def _validate_parent_department(parent_id: int | None, current_dept_id: int | None = None) -> tuple[Department | None, str | None]:
    if parent_id is None:
        return None, None
    parent = db.session.get(Department, parent_id)
    if not parent:
        return None, "上级部门不存在"
    if current_dept_id and parent.id == current_dept_id:
        return None, "上级部门不能选择自身"
    if current_dept_id:
        cursor = parent
        while cursor:
            if cursor.id == current_dept_id:
                return None, "上级部门设置非法，会形成循环层级"
            cursor = cursor.parent
    return parent, None


def _parse_header_row(rows: list[list[Any]], expected: list[str]) -> tuple[int, dict[str, int]]:
    best_idx = 0
    best_map: dict[str, int] = {}
    best_score = -1
    limit = min(len(rows), 8)
    for idx in range(limit):
        raw = rows[idx]
        header_map = {}
        for i, value in enumerate(raw):
            text = str(value).strip() if value is not None else ""
            if text:
                header_map[text] = i
        score = sum(1 for key in expected if key in header_map)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_map = header_map
    return best_idx, best_map


def _resolve_shift(shift_no: str | None) -> Shift | None:
    key = (shift_no or "").strip()
    if not key:
        return None
    return Shift.query.filter_by(shift_no=key).first()


def _assign_employee_shift(employee: Employee, shift: Shift | None) -> None:
    assignment = employee.shift_assignment
    if shift is None:
        if assignment:
            db.session.delete(assignment)
        return

    if not assignment:
        assignment = EmployeeShiftAssignment(emp_id=employee.id, shift_id=shift.id)
        db.session.add(assignment)
        employee.shift_assignment = assignment
    else:
        assignment.shift_id = shift.id


def list_account_sets():
    rows = AccountSet.query.order_by(AccountSet.month.desc()).all()
    return jsonify([_serialize_account_set(row) for row in rows])


def create_account_set():
    data = request.json or {}
    month = (data.get("month") or "").strip()
    factory_rest_days = data.get("factory_rest_days", 0)
    monthly_benefit_days = data.get("monthly_benefit_days", 0)
    if not month or len(month) != 7:
        return jsonify({"error": "month is required in YYYY-MM format"}), 400
    if AccountSet.query.filter_by(month=month).first():
        return jsonify({"error": "该月份账套已存在"}), 400
    try:
        entries = _parse_factory_rest_entries(data.get("factory_rest_entries"), month)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    row = AccountSet(
        month=month,
        name=f"{month} 账套",
        factory_rest_days=float(factory_rest_days or 0),
        monthly_benefit_days=float(monthly_benefit_days or 0),
    )
    if AccountSet.query.count() == 0:
        row.is_active = True
    db.session.add(row)
    if "factory_rest_entries" in data:
        _replace_factory_rest_entries(row, entries)
    db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


def update_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "修改账套参数")
    if locked_error:
        return locked_error
    data = request.json or {}
    if "factory_rest_entries" in data:
        try:
            entries = _parse_factory_rest_entries(data.get("factory_rest_entries"), row.month)
            _validate_initial_factory_rest_entry_backfill(row, entries)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        _replace_factory_rest_entries(row, entries)
    elif "factory_rest_days" in data:
        row.factory_rest_days = float(data.get("factory_rest_days") or 0)
    row.monthly_benefit_days = float(data.get("monthly_benefit_days") or 0)
    db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


def activate_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    AccountSet.query.update({AccountSet.is_active: False})
    row.is_active = True
    db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


def lock_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    if not row.is_locked:
        row.is_locked = True
        row.locked_at = datetime.utcnow()
        row.locked_by = g.current_user.id
        db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


def unlock_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    if row.is_locked:
        row.is_locked = False
        row.locked_at = None
        row.locked_by = None
        db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


def reset_account_set_imported(account_set_id: int):
    """清空账套已导入数据：按账套月份删月报/日报/请假/加班及归档，供错传后重置。

    管理人员月统计与年假余额不在清理范围：重传正确文件并重新计算后由计算管线
    自行同步；归档文件与导入记录一并删除，避免残留错文件被再次计算。
    """
    from models.leave import LeaveRecord
    from models.monthly_report import MonthlyReport

    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "清空已导入数据")
    if locked_error:
        return locked_error

    year, month = int(row.month[:4]), int(row.month[5:7])
    start_date = date(year, month, 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    next_date = date(next_year, next_month, 1)
    start_dt = datetime.combine(start_date, datetime.min.time())
    next_dt = datetime.combine(next_date, datetime.min.time())

    deleted = {
        "monthly_reports": MonthlyReport.query.filter(
            MonthlyReport.report_month == row.month
        ).delete(synchronize_session=False),
        "daily_records": DailyRecord.query.filter(
            DailyRecord.record_date >= start_date, DailyRecord.record_date < next_date
        ).delete(synchronize_session=False),
        "leave_records": LeaveRecord.query.filter(
            LeaveRecord.start_time >= start_dt, LeaveRecord.start_time < next_dt
        ).delete(synchronize_session=False),
        "overtime_records": OvertimeRecord.query.filter(
            OvertimeRecord.start_time >= start_dt, OvertimeRecord.start_time < next_dt
        ).delete(synchronize_session=False),
    }

    for record in row.imports:
        path = (record.stored_path or "").strip()
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
    deleted["import_records"] = AccountSetImport.query.filter_by(
        account_set_id=row.id
    ).delete(synchronize_session=False)

    db.session.commit()
    return jsonify({"status": "ok", "month": row.month, "deleted": deleted})


def delete_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "删除账套")
    if locked_error:
        return locked_error

    # best-effort cleanup archived files
    for record in row.imports:
        path = (record.stored_path or "").strip()
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass

    was_active = row.is_active
    db.session.delete(row)
    db.session.commit()

    if was_active:
        fallback = AccountSet.query.order_by(AccountSet.month.desc()).first()
        if fallback:
            AccountSet.query.update({AccountSet.is_active: False})
            fallback.is_active = True
            db.session.commit()

    return jsonify({"status": "ok"})


def get_account_set_calc_progress(account_set_id: int):
    mode = (request.args.get("mode") or "").strip()
    # mode 会拼入进度文件路径，白名单外的值一律按未开始处理
    if mode not in ("employee", "manager", "all"):
        return jsonify({"status": "idle", "percent": 0, "stage": ""})
    progress = get_calc_progress(account_set_id, mode)
    if progress is None:
        return jsonify({"status": "idle", "percent": 0, "stage": ""})
    return jsonify(progress)


def calculate_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "重新计算")
    if locked_error:
        return locked_error
    mode = (request.args.get("mode") or "all").strip()
    records_query = AccountSetImport.query.filter_by(account_set_id=row.id)
    if mode == "employee":
        records_query = records_query.filter(AccountSetImport.file_type.in_(["leave", "overtime", "monthly", "daily"]))
    elif mode == "manager":
        records_query = records_query.filter(AccountSetImport.file_type.in_(["leave", "overtime", "manager_monthly", "manager_daily"]))
    records = records_query.order_by(AccountSetImport.id.asc()).all()

    if not records:
        return jsonify({"status": "error", "message": "该账套暂无可计算文件", "mode": mode}), 400

    success = 0
    failed = 0
    results = []

    # ---- 计算进度（写入进度文件，前端轮询）----
    stage_count = len(records) + (1 if mode == "manager" else 0)
    last_percent = -100

    def _report(stage_index: int, done: int, total: int, stage: str) -> None:
        nonlocal last_percent
        fraction = done / total if total > 0 else 1.0
        percent = min(int((stage_index + fraction) * 100 / stage_count), 100)
        # 行级回调很密集，百分比变化不足 2 时跳过写盘
        if percent < 100 and percent - last_percent < 2:
            return
        last_percent = percent
        update_calc_progress(row.id, mode, percent, stage)

    update_calc_progress(row.id, mode, 0, "开始计算...")

    for stage_idx, rec in enumerate(records):
        stage_text = f"正在导入 {rec.source_filename or ''}（{stage_idx + 1}/{len(records)}）"
        path = (rec.stored_path or "").strip()
        filename = rec.source_filename or os.path.basename(path)
        if not path or not os.path.exists(path):
            failed += 1
            rec.status = "error"
            rec.error_message = "archive file not found"
            rec.imported_count = 0
            results.append({"file": filename, "status": "error", "error": rec.error_message})
            _report(stage_idx, 1, 1, stage_text)
            db.session.commit()
            continue

        try:
            imported = ImportService.import_file(
                path,
                progress_cb=lambda done, total, _idx=stage_idx, _stage=stage_text: _report(_idx, done, total, _stage),
            )
            if imported.get("status") == "ok":
                success += 1
                rec.status = "ok"
                rec.file_type = imported.get("file_type")
                rec.imported_count = imported.get("imported", 0)
                rec.error_message = None
                results.append({"file": filename, "status": "ok", "result": imported})
            else:
                failed += 1
                rec.status = "error"
                rec.error_message = imported.get("message", "import failed")
                rec.imported_count = 0
                results.append({"file": filename, "status": "error", "error": rec.error_message, "result": imported})
        except Exception:
            failed += 1
            logger.exception("账套文件导入失败 file=%s", filename)
            rec.status = "error"
            rec.error_message = "导入失败，请查看服务端日志"
            rec.imported_count = 0
            results.append({"file": filename, "status": "error", "error": "导入失败，请查看服务端日志"})
        db.session.commit()

    manager_stats_sync = None
    if mode == "manager" and failed == 0:
        try:
            manager_options = ManagerAttendanceOptions(
                month=row.month,
                factory_rest_days=_effective_factory_rest_days(row),
                monthly_benefit_days=row.monthly_benefit_days or 0,
            )
            manager_stats_sync = _sync_manager_month_stats(
                row.month,
                manager_options,
                progress_cb=lambda done, total: _report(len(records), done, total, "正在汇总管理人员考勤..."),
            )
            if manager_stats_sync["error_count"]:
                failed += manager_stats_sync["error_count"]
        except Exception:
            db.session.rollback()
            failed += 1
            logger.exception("管理人员考勤统计同步失败 month=%s", row.month)
            manager_stats_sync = {
                "month": row.month,
                "overtime_synced": 0,
                "annual_leave_synced": 0,
                "error_count": 1,
                "errors": ["考勤统计同步失败，请查看服务端日志"],
            }

    update_calc_progress(row.id, mode, 100, "计算完成", status="finished")
    return jsonify(
        {
            "status": "ok" if failed == 0 else "partial",
            "account_set": _serialize_account_set(row),
            "mode": mode,
            "total": len(records),
            "success": success,
            "failed": failed,
            "manager_stats_sync": manager_stats_sync,
            "results": results,
        }
    )


def create_shift():
    data = request.json or {}
    shift_no = (data.get("shift_no") or "").strip()
    shift_name = (data.get("shift_name") or "").strip()
    time_slots = data.get("time_slots") or []
    is_cross_day = bool(data.get("is_cross_day", False))

    if not shift_no or not shift_name:
        return jsonify({"error": "shift_no and shift_name are required"}), 400

    shift = Shift.query.filter_by(shift_no=shift_no).first()
    if shift:
        return jsonify({"error": "shift_no already exists"}), 400

    shift = Shift(
        shift_no=shift_no,
        shift_name=shift_name,
        time_slots=time_slots,
        is_cross_day=is_cross_day,
    )
    db.session.add(shift)

    db.session.commit()
    return jsonify({"status": "ok", "id": shift.id})


def list_shifts():
    rows = Shift.query.order_by(Shift.shift_no.asc()).all()
    return jsonify(
        [
            {
                "id": s.id,
                "shift_no": s.shift_no,
                "shift_name": s.shift_name,
                "time_slots": s.time_slots or [],
                "is_cross_day": s.is_cross_day,
            }
            for s in rows
        ]
    )


def update_shift(shift_id: int):
    data = request.json or {}
    shift_no = (data.get("shift_no") or "").strip()
    shift_name = (data.get("shift_name") or "").strip()
    time_slots = data.get("time_slots") or []
    is_cross_day = bool(data.get("is_cross_day", False))

    shift = _require_model(Shift, shift_id)
    if not shift_no or not shift_name:
        return jsonify({"error": "shift_no and shift_name are required"}), 400

    duplicate = Shift.query.filter(Shift.shift_no == shift_no, Shift.id != shift_id).first()
    if duplicate:
        return jsonify({"error": "shift_no already exists"}), 400

    shift.shift_no = shift_no
    shift.shift_name = shift_name
    shift.time_slots = time_slots
    shift.is_cross_day = is_cross_day
    db.session.commit()
    return jsonify({"status": "ok"})


def delete_shift(shift_id: int):
    shift = _require_model(Shift, shift_id)
    if shift.employee_assignments:
        return jsonify({"error": "该班次已绑定员工，无法删除"}), 400
    if shift.daily_records:
        return jsonify({"error": "该班次已有关联考勤记录，无法删除"}), 400

    db.session.delete(shift)
    db.session.commit()
    return jsonify({"status": "ok"})


RESIGN_DISABLE_REASON = "employee_resigned"


def _disable_accounts_for_resignation(employee: Employee) -> None:
    for assignment in UserEmployeeAssignment.query.filter_by(emp_id=employee.id).all():
        user = assignment.user
        if user is None or user.role == "admin":
            continue
        profile_emp_no = (user.profile_emp_no or "").strip()
        if profile_emp_no and profile_emp_no != employee.emp_no:
            continue
        if not profile_emp_no and (user.username or "").strip().lower() != employee.emp_no.lower():
            continue
        if user.is_login_disabled():
            continue
        user.login_disabled_until_admin_unlock = True
        user.login_disabled_reason = RESIGN_DISABLE_REASON


def _release_resignation_lock_for_users(employee: Employee) -> None:
    for assignment in UserEmployeeAssignment.query.filter_by(emp_id=employee.id).all():
        user = assignment.user
        if user is None or user.login_disabled_reason != RESIGN_DISABLE_REASON:
            continue
        still_bound_to_resigned = (
            UserEmployeeAssignment.query.filter(
                UserEmployeeAssignment.user_id == user.id,
                UserEmployeeAssignment.emp_id != employee.id,
            )
            .join(Employee, Employee.id == UserEmployeeAssignment.emp_id)
            .filter(Employee.resigned_at.isnot(None))
            .first()
        )
        if still_bound_to_resigned:
            continue
        user.login_disabled_until_admin_unlock = False
        user.login_disabled_reason = None


def resign_employee_by_emp_no():
    data = request.json or {}
    emp_no = (data.get("emp_no") or "").strip()
    if not emp_no:
        return jsonify({"error": "请输入工号"}), 400

    raw_date = (data.get("resigned_at") or "").strip()
    if raw_date:
        try:
            resigned_at = datetime.strptime(raw_date, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "离职日期格式不正确，应为 YYYY-MM-DD"}), 400
    else:
        resigned_at = date.today()

    employee = Employee.query.filter_by(emp_no=emp_no).first()
    if not employee:
        return jsonify({"error": "工号不存在"}), 400
    if employee.resigned_at:
        return jsonify({"error": "该员工已离职"}), 400

    employee.resigned_at = resigned_at
    _disable_accounts_for_resignation(employee)
    db.session.commit()
    return jsonify({"status": "ok", "employee": _serialize_employee(employee)})


def reinstate_employee(employee_id: int):
    employee = _require_model(Employee, employee_id)
    if not employee.resigned_at:
        return jsonify({"error": "该员工未离职"}), 400

    employee.resigned_at = None
    _release_resignation_lock_for_users(employee)
    db.session.commit()
    return jsonify({"status": "ok", "employee": _serialize_employee(employee)})


def employees_list():
    allowed_ids = _accessible_emp_ids_set()
    query = Employee.query
    if getattr(g, "current_user", None) and g.current_user.role != "admin":
        if not allowed_ids:
            return jsonify([])
        query = query.filter(Employee.id.in_(allowed_ids))
    status = (request.args.get("status") or "active").strip()
    if status == "active":
        query = query.filter(Employee.resigned_at.is_(None))
    elif status == "resigned":
        query = query.filter(Employee.resigned_at.isnot(None))
    rows = _unique_employees(query.order_by(Employee.emp_no.asc()).all())
    return jsonify([_serialize_employee(e) for e in rows])


def departments_list():
    allowed_ids = _accessible_dept_ids_set()
    query = Department.query
    if getattr(g, "current_user", None) and g.current_user.role != "admin":
        if not allowed_ids:
            return jsonify([])
        query = query.filter(Department.id.in_(allowed_ids))
    rows = query.order_by(Department.dept_name.asc()).all()
    return jsonify(
        [
            {
                "id": d.id,
                "dept_no": d.dept_no,
                "dept_name": d.dept_name,
                "parent_id": d.parent_id,
                "parent_name": d.parent.dept_name if d.parent else "",
                "is_locked": bool(d.is_locked),
            }
            for d in rows
        ]
    )


def manager_overtime_records():
    year = request.args.get("year", type=int) or datetime.now().year
    return jsonify(_manager_month_rows(_manager_overtime_values(year, _requested_emp_ids() or None), "剩余调休天数"))


def update_manager_overtime_summary():
    payload, status = _save_manager_month_stat("overtime")
    return jsonify(payload), status



def _manager_export_months(year: int) -> list[tuple[str, str]]:
    months = [(f"{year - 1}-12", "12月")]
    months.extend((f"{year}-{month:02d}", f"{month}月") for month in range(1, 13))
    return months


def _manager_base_month_values(emp_ids: list[int] | None = None) -> dict[str, dict[str, object]]:
    employees = _manager_scope_employees(emp_ids)
    return {
        employee.name: {
            "emp_id": employee.id,
            "dept_name": employee.department.dept_name if employee.department else "",
            "remark": "",
        }
        for employee in employees
    }


def _month_value_keys() -> list[str]:
    return ["prev_dec", *[f"m{month}" for month in range(1, 13)]]


def _annual_leave_value_keys() -> list[str]:
    return [f"m{month}" for month in range(1, 13)]


def _number_or_blank(value: object) -> float | str:
    if value in (None, ""):
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return ""


def _month_for_stat_key(year: int, key: str) -> str:
    if key == "prev_dec":
        return f"{year - 1}-12"
    return f"{year}-{int(key[1:]):02d}"


def _validate_manager_month_stat(stat_type: str, year: int, values: dict[str, float]) -> str | None:
    if stat_type == "annual_leave":
        total = sum(values.values())
        if any(value < 0 for value in values.values()):
            return "年休使用天数不能为负数"
        if total > 12:
            return "年休一年最多 12 天"
        for key, value in values.items():
            if value > 3:
                return "年休每月使用不能超过 3 天"
            month = _month_for_stat_key(year, key)
            account_set = AccountSet.query.filter_by(month=month).first()
            factory_rest_days = _manager_factory_rest_days(account_set)
            if factory_rest_days + value > 7:
                return f"{month} 厂休+年休不能超过 7 天"

    if stat_type == "overtime":
        balance = 0.0
        for key in _month_value_keys():
            value = values.get(key, 0.0)
            if value < 0 and balance <= 0:
                return "剩余加班天数为 0 时不能使用调休"
            if balance + value < 0:
                return "使用调休天数不能超过当前剩余加班天数"
            balance += value
    return None


def _stat_value_keys(stat_type: str) -> list[str]:
    return _annual_leave_value_keys() if stat_type == "annual_leave" else _month_value_keys()


def _stat_headers(stat_type: str) -> list[str]:
    if stat_type == "annual_leave":
        return ["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"]
    return ["部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"]


def _stat_col_keys(stat_type: str) -> list[tuple[str, str]]:
    if stat_type == "annual_leave":
        return [(f"m{month}", f"{month}月") for month in range(1, 13)]
    return [("prev_dec", "前年累积天数"), *[(f"m{month}", f"{month}月") for month in range(1, 13)]]


def _apply_saved_manager_stats(values_by_name: dict[str, dict[str, object]], year: int, stat_type: str) -> None:
    rows = (
        ManagerMonthStat.query.join(Employee, ManagerMonthStat.emp_id == Employee.id)
        .filter(
            ManagerMonthStat.year == year,
            ManagerMonthStat.stat_type == stat_type,
            Employee.is_manager.is_(True),
        )
        .all()
    )
    for row in rows:
        if not row.employee or row.employee.name not in values_by_name:
            continue
        values = values_by_name[row.employee.name]
        for key in _stat_value_keys(stat_type):
            values[key] = getattr(row, key)
        if stat_type == "annual_leave":
            values["prev_dec"] = ""
        values["remaining"] = row.remaining
        values["remark"] = row.remark or ""


def _upsert_manager_month_stat(stat_type: str, emp_id: int, year: int, values: dict[str, float], remark: str) -> tuple[dict[str, str], int]:
    employee = _require_model(Employee, emp_id)
    if not employee.is_manager:
        return {"error": "employee is not manager"}, 400
    row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
    error = _validate_manager_month_stat(stat_type, year, values)
    if error:
        return {"error": error}, 400
    if not row:
        row = ManagerMonthStat(emp_id=employee.id, year=year, stat_type=stat_type)
        db.session.add(row)
    if stat_type == "annual_leave":
        row.prev_dec = 0
    for key, value in values.items():
        setattr(row, key, value)
    row.remaining = round(12 - sum(values.values()), 2) if stat_type == "annual_leave" else round(sum(values.values()), 2)
    row.remark = (remark or "").strip()
    db.session.commit()
    return {"status": "ok"}, 200


def _save_manager_month_stat(stat_type: str) -> tuple[dict[str, str], int]:
    data = request.json or {}
    emp_id = int(data.get("emp_id") or 0)
    year = int(data.get("year") or datetime.now().year)
    keys = _stat_value_keys(stat_type)
    submitted_values = {key: float(_number_or_blank(data.get(key)) or 0) for key in keys}
    employee = _require_model(Employee, emp_id)
    if not employee.is_manager:
        return {"error": "employee is not manager"}, 400

    row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
    values = {
        key: float(getattr(row, key) or 0) if row else 0.0
        for key in keys
    }
    skipped_locked_months: list[str] = []
    for key in keys:
        month = _stat_key_month(year, key)
        if _stat_key_lock_state(year, key) == "locked":
            if month:
                skipped_locked_months.append(month)
            continue
        values[key] = submitted_values[key]

    payload, status = _upsert_manager_month_stat(stat_type, emp_id, year, values, data.get("remark") or "")
    if status != 200:
        return payload, status
    if skipped_locked_months:
        payload["skipped_locked_months"] = sorted(set(skipped_locked_months))
        payload["warning"] = f"已跳过锁定月份：{'、'.join(payload['skipped_locked_months'])}"
    return payload, status

def _validate_month(value: str | None) -> str | None:
    text = (value or "").strip()
    try:
        datetime.strptime(text, "%Y-%m")
    except ValueError:
        return None
    return text


def _manager_attendance_options(month: str) -> ManagerAttendanceOptions:
    # 请求级缓存：options 为纯只读派生值，同请求内重复构建无意义
    cache = getattr(g, "_manager_attendance_options_cache", None)
    if cache is None:
        cache = {}
        g._manager_attendance_options_cache = cache
    if month in cache:
        return cache[month]
    account = AccountSet.query.filter_by(month=month).first()
    options = ManagerAttendanceOptions(
        month=month,
        factory_rest_days=_manager_factory_rest_days(account),
        monthly_benefit_days=(account.monthly_benefit_days if account else 0) or 0,
    )
    cache[month] = options
    return options


def _manager_attendance_override_payload(row: ManagerAttendanceOverride | None) -> dict[str, object]:
    payload = {field: getattr(row, field) if row else None for field in _MANAGER_ATTENDANCE_OVERRIDE_FIELDS}
    payload["remark"] = row.remark if row else ""
    payload["updated_at"] = row.updated_at.isoformat() if row and row.updated_at else None
    payload["updated_by_name"] = _user_display_name(row.updated_by_user) if row else ""
    return payload


def _manager_attendance_row(emp_id: int, month: str, include_overrides: bool) -> dict[str, object] | None:
    options = _manager_attendance_options(month)
    rows = build_manager_rows(options, [emp_id], include_overrides=include_overrides)
    return rows[0] if rows else None


def _manager_attendance_response(emp_id: int, month: str) -> tuple[dict[str, object], int]:
    employee = db.session.get(Employee, emp_id)
    if not employee or not employee.is_manager:
        return {"error": "employee is not manager"}, 400
    automatic = _manager_attendance_row(emp_id, month, include_overrides=False)
    applied = _manager_attendance_row(emp_id, month, include_overrides=True)
    override = ManagerAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    return {
        "employee": _serialize_employee(employee),
        "month": month,
        "automatic": automatic,
        "override": _manager_attendance_override_payload(override),
        "applied": applied,
        "history": _history_rows_for_month("manager", month),
    }, 200


def _manager_attendance_list_response(emp_ids: list[int], month: str) -> tuple[dict[str, object], int]:
    if not emp_ids:
        emp_ids = list(_accessible_emp_ids_set())
        if not emp_ids:
            return {"rows": [], "month": month}, 200
    employees = {
        employee.id: employee
        for employee in Employee.query.filter(
            Employee.id.in_(emp_ids),
            Employee.is_manager.is_(True),
        ).all()
    }
    overrides = {
        row.emp_id: row
        for row in ManagerAttendanceOverride.query.filter(
            ManagerAttendanceOverride.month == month,
            ManagerAttendanceOverride.emp_id.in_(emp_ids),
        ).all()
    }
    rows: list[dict[str, object]] = []
    options = _manager_attendance_options(month)
    valid_ids = list(employees.keys())
    automatic_by_emp = {row["emp_id"]: row for row in build_manager_rows(options, valid_ids, include_overrides=False)}
    applied_by_emp = {row["emp_id"]: row for row in build_manager_rows(options, valid_ids, include_overrides=True)}
    for emp_id in emp_ids:
        employee = employees.get(emp_id)
        if not employee:
            continue
        rows.append(
            {
                "employee": _serialize_employee(employee),
                "automatic": automatic_by_emp.get(emp_id),
                "override": _manager_attendance_override_payload(overrides.get(emp_id)),
                "applied": applied_by_emp.get(emp_id),
            }
        )
    return {"rows": rows, "month": month}, 200


def _nullable_float(data: dict[str, object], key: str) -> tuple[float | None, str | None]:
    value = data.get(key)
    if value in (None, ""):
        return None, None
    try:
        parsed = round(float(value), 2)
    except (TypeError, ValueError):
        return None, f"{key} 必须是数字"
    if parsed < 0:
        return None, f"{key} 不能为负数"
    return parsed, None


def _nullable_int(data: dict[str, object], key: str) -> tuple[int | None, str | None]:
    value = data.get(key)
    if value in (None, ""):
        return None, None
    try:
        parsed = int(round(float(value)))
    except (TypeError, ValueError):
        return None, f"{key} 必须是整数"
    if parsed < 0:
        return None, f"{key} 不能为负数"
    return parsed, None


def _override_workbook_response(wb: openpyxl.Workbook, filename: str):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _employee_override_export_headers() -> list[str]:
    return [
        "月份",
        "工号",
        "姓名",
        "系统考勤天数",
        "系统实际出勤天数",
        "系统工时",
        "系统半勤天数",
        "系统迟到早退",
        "考勤天数",
        "实际出勤天数",
        "工时",
        "半勤天数",
        "迟到早退",
        "备注",
    ]


def _manager_override_export_headers() -> list[str]:
    return [
        "月份",
        "工号",
        "姓名",
        "系统出勤天数",
        "系统工伤",
        "系统出差",
        "系统婚假",
        "系统丧假",
        "系统迟到早退",
        "出勤天数",
        "工伤",
        "出差",
        "婚假",
        "丧假",
        "迟到早退",
        "备注",
    ]


def _build_employee_override_export_workbook(month: str, include_real_rows: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工考勤修正"
    ws.append(_employee_override_export_headers())
    if include_real_rows:
        employees = (
            Employee.query.filter_by(is_manager=False)
            .order_by(Employee.dept_id.asc(), Employee.emp_no.asc(), Employee.name.asc())
            .all()
        )
        automatic_by_emp = _employee_automatic_rows_by_emp([e.id for e in employees], month)
        overrides_by_emp = {
            row.emp_id: row
            for row in EmployeeAttendanceOverride.query.filter(
                EmployeeAttendanceOverride.month == month
            ).all()
        }
        for employee in employees:
            automatic = automatic_by_emp.get(employee.id) or {}
            override = overrides_by_emp.get(employee.id)
            ws.append(
                [
                    month,
                    employee.emp_no,
                    employee.name,
                    automatic.get("attendance_days"),
                    automatic.get("actual_attendance_days"),
                    automatic.get("work_hours"),
                    automatic.get("half_days"),
                    automatic.get("late_early_minutes"),
                    override.attendance_days if override else "",
                    override.actual_attendance_days if override else "",
                    override.work_hours if override else "",
                    override.half_days if override else "",
                    override.late_early_minutes if override else "",
                    override.remark if override and override.remark else "",
                ]
            )
    else:
        ws.append(["2026-05", "1001001", "张三", 20, 18, 160, 0, 5, "", "", "", "", "", "留空表示不覆盖"])
    return wb


def _build_manager_override_export_workbook(month: str, include_real_rows: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "管理人员考勤修正"
    ws.append(_manager_override_export_headers())
    if include_real_rows:
        employees = _manager_scope_employees()
        options = _manager_attendance_options(month)
        automatic_by_emp = {
            row["emp_id"]: row
            for row in build_manager_rows(options, [e.id for e in employees], include_overrides=False)
        }
        overrides_by_emp = {
            row.emp_id: row
            for row in ManagerAttendanceOverride.query.filter(
                ManagerAttendanceOverride.month == month
            ).all()
        }
        for employee in employees:
            automatic = automatic_by_emp.get(employee.id) or {}
            override = overrides_by_emp.get(employee.id)
            ws.append(
                [
                    month,
                    employee.emp_no,
                    employee.name,
                    automatic.get("attendance_days"),
                    automatic.get("injury_days"),
                    automatic.get("business_trip_days"),
                    automatic.get("marriage_days"),
                    automatic.get("funeral_days"),
                    automatic.get("late_early_minutes"),
                    override.attendance_days if override else "",
                    override.injury_days if override else "",
                    override.business_trip_days if override else "",
                    override.marriage_days if override else "",
                    override.funeral_days if override else "",
                    override.late_early_minutes if override else "",
                    override.remark if override and override.remark else "",
                ]
            )
    else:
        ws.append(["2026-05", "2001001", "李经理", 22, 0, 2, 0, 0, 0, "", "", "", "", "", "", "留空表示不覆盖"])
    return wb


def _import_summary(success_count: int, skipped_count: int, failed_count: int, changed_count: int, errors: list[str]) -> dict[str, object]:
    return {
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        "changed_count": changed_count,
        "errors": errors,
    }


def _apply_employee_override_updates(
    row: EmployeeAttendanceOverride | None,
    emp_id: int,
    month: str,
    updates: dict[str, object],
    action_type: str,
    source_file_name: str | None = None,
) -> bool:
    before_values = _override_state_from_row(row, _EMPLOYEE_OVERRIDE_FIELDS)
    after_values = dict(before_values)
    after_values.update(updates)
    if not _has_override_state_changes(before_values, after_values):
        return False
    if not row:
        row = EmployeeAttendanceOverride(emp_id=emp_id, month=month)
        db.session.add(row)
    for field in _EMPLOYEE_OVERRIDE_FIELDS:
        setattr(row, field, after_values.get(field))
    row.remark = str(after_values.get("remark") or "")
    row.updated_by = g.current_user.id
    _record_override_history("employee", emp_id, month, action_type, before_values, after_values, source_file_name)
    return True


def _apply_manager_override_updates(
    row: ManagerAttendanceOverride | None,
    emp_id: int,
    month: str,
    updates: dict[str, object],
    action_type: str,
    source_file_name: str | None = None,
) -> bool:
    before_values = _override_state_from_row(row, _MANAGER_ATTENDANCE_OVERRIDE_FIELDS)
    after_values = dict(before_values)
    after_values.update(updates)
    if not _has_override_state_changes(before_values, after_values):
        return False
    if not row:
        row = ManagerAttendanceOverride(emp_id=emp_id, month=month)
        db.session.add(row)
    for field in _MANAGER_ATTENDANCE_OVERRIDE_FIELDS:
        setattr(row, field, after_values.get(field))
    row.remark = str(after_values.get("remark") or "")
    row.updated_by = g.current_user.id
    _record_override_history("manager", emp_id, month, action_type, before_values, after_values, source_file_name)
    return True


def _stat_key_for_month(month: str) -> tuple[int, str]:
    year, month_no = [int(part) for part in month.split("-", 1)]
    return year, f"m{month_no}"


def _sync_manager_month_stats(
    month: str,
    options: ManagerAttendanceOptions | None = None,
    progress_cb=None,
) -> dict[str, object]:
    """重算指定月份的管理人员考勤，并把加班/年休结果写回 manager_month_stats。

    管理人员考勤修正（单条保存/删除、批量导入）会改变 attendance_days，
    进而影响加班天数（出勤+厂休>当月天数 即产生加班）。加班查询页直接读取
    manager_month_stats，因此修正后必须重算并同步，否则两边数据不一致。

    本函数复用账套文件导入的同步路径：build_manager_rows(sync_month_stats=True)
    会在计算时直接写回 stat 表，再由 _sync_manager_stats_from_manager_rows 校验。
    options 缺省时按 _manager_attendance_options(month) 构造，与查询页保持一致。
    """
    if options is None:
        options = _manager_attendance_options(month)
    manager_rows = build_manager_rows(options, sync_month_stats=True, progress_cb=progress_cb)
    result = _sync_manager_stats_from_manager_rows(month, manager_rows)
    db.session.commit()
    return result


def _sync_manager_stats_from_manager_rows(month: str, manager_rows: list[dict[str, object]]) -> dict[str, object]:
    """Stats are now written directly inside build_manager_rows.
    This function provides a compatibility layer — it validates the results
    but no longer writes to the stat tables to avoid double-writing.
    """
    year, key = _stat_key_for_month(month)
    employees_by_name = {employee.name: employee for employee in _manager_scope_employees()}
    errors: list[str] = []

    for item in manager_rows:
        name = str(item.get("name") or "").strip()
        employee = employees_by_name.get(name)
        if not employee:
            errors.append(f"{name or '空姓名'}：未找到管理人员")
            continue

        for stat_type, source_key, label in (
            ("overtime", "overtime_change", "加班变化"),
            ("annual_leave", "benefit_days", "福利天数"),
        ):
            stat_row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
            values = {
                stat_key: float(getattr(stat_row, stat_key) or 0) if stat_row else 0.0
                for stat_key in _stat_value_keys(stat_type)
            }
            # Values already written by build_manager_rows; validate only
            error = _validate_manager_month_stat(stat_type, year, values)
            if error:
                errors.append(f"{employee.name} {label}：{error}")

    return {
        "month": month,
        "overtime_synced": 0,
        "annual_leave_synced": 0,
        "error_count": len(errors),
        "errors": errors,
    }


def _download_manager_stat_template(stat_type: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "年休统计表" if stat_type == "annual_leave" else "加班统计表"
    ws.append(_stat_headers(stat_type))
    employees = _manager_scope_employees()
    for employee in employees:
        values = [employee.department.dept_name if employee.department else "", employee.name]
        values.extend("" for _key, _label in _stat_col_keys(stat_type))
        values.extend([12 if stat_type == "annual_leave" else 0, ""])
        ws.append(values)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "管理人员年休导入示例.xlsx" if stat_type == "annual_leave" else "管理人员加班导入示例.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _header_map(ws) -> dict[str, int]:
    headers = {}
    for cell in ws[1]:
        text = str(cell.value or "").strip()
        if text:
            headers[text] = cell.column
    return headers


def _import_manager_stat_file(stat_type: str, year: int):
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "请选择要导入的Excel文件"}), 400

    wb = openpyxl.load_workbook(file, data_only=True)
    try:
        ws = wb.active
        headers = _header_map(ws)
        if "姓名" not in headers:
            return jsonify({"error": "导入文件缺少姓名列"}), 400

        imported = 0
        errors: list[str] = []
        skipped_locked_months: set[str] = set()
        employees_by_name = {employee.name: employee for employee in _manager_scope_employees()}
        for row_idx in range(2, ws.max_row + 1):
            name = str(ws.cell(row_idx, headers["姓名"]).value or "").strip()
            if not name:
                continue
            employee = employees_by_name.get(name)
            if not employee:
                errors.append(f"第{row_idx}行：未找到管理人员 {name}")
                continue

            existing_row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
            values = {
                key: float(getattr(existing_row, key) or 0) if existing_row else 0.0
                for key in _stat_value_keys(stat_type)
            }
            for key, label in _stat_col_keys(stat_type):
                col_idx = headers.get(label)
                month = _stat_key_month(year, key)
                if _stat_key_lock_state(year, key) == "locked":
                    if month:
                        skipped_locked_months.add(month)
                    continue
                values[key] = float(_number_or_blank(ws.cell(row_idx, col_idx).value if col_idx else None) or 0)
            remark_col = headers.get("备注")
            remark = str(ws.cell(row_idx, remark_col).value or "").strip() if remark_col else ""
            payload, status = _upsert_manager_month_stat(stat_type, employee.id, year, values, remark)
            if status != 200:
                errors.append(f"第{row_idx}行：{payload.get('error', '保存失败')}")
                continue
            imported += 1

        response = {
            "status": "ok",
            "imported": imported,
            "errors": errors,
            "error_count": len(errors),
            "skipped_locked_months": sorted(skipped_locked_months),
        }
        if skipped_locked_months:
            response["warning"] = f"已跳过锁定月份：{'、'.join(response['skipped_locked_months'])}"
        return jsonify(response)
    finally:
        wb.close()


def _manager_overtime_values(year: int, emp_ids: list[int] | None = None) -> dict[str, dict[str, object]]:
    values_by_name = _manager_base_month_values(emp_ids)
    for values in values_by_name.values():
        values["remaining"] = 0
    _apply_saved_manager_stats(values_by_name, year, "overtime")
    return values_by_name


def _manager_annual_leave_values(year: int, emp_ids: list[int] | None = None) -> dict[str, dict[str, object]]:
    values_by_name = _manager_base_month_values(emp_ids)
    for values in values_by_name.values():
        values["remaining"] = 12
    _apply_saved_manager_stats(values_by_name, year, "annual_leave")
    return values_by_name


def _manager_month_rows(values_by_name: dict[str, dict[str, object]], remaining_label: str, keys: list[str] | None = None) -> list[dict[str, object]]:
    keys = keys or _month_value_keys()
    rows = []
    for name, values in values_by_name.items():
        row = {
            "emp_id": values.get("emp_id"),
            "dept_name": values.get("dept_name", ""),
            "name": name,
            "remaining_label": remaining_label,
            "remaining": values.get("remaining", ""),
            "remark": values.get("remark", ""),
        }
        for key in keys:
            row[key] = values.get(key, "")
        rows.append(row)
    return rows


def _fill_named_month_template(ws, values_by_name: dict[str, dict[str, object]], total_key: str, keys: list[str] | None = None) -> None:
    keys = keys or _month_value_keys()
    total_col = 3 + len(keys)
    remark_col = total_col + 1
    filled: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row_idx, 2).value or "").strip()
        for col_idx in range(3, remark_col + 1):
            ws.cell(row_idx, col_idx).value = ""
        if not name or name not in values_by_name:
            continue
        values = values_by_name[name]
        for col_idx, key in enumerate(keys, start=3):
            ws.cell(row_idx, col_idx).value = values.get(key, "")
        ws.cell(row_idx, total_col).value = values.get(total_key, "")
        ws.cell(row_idx, remark_col).value = values.get("remark", "")
        filled.add(name)

    for name, values in values_by_name.items():
        if name in filled:
            continue
        ws.append(
            [
                values.get("dept_name", ""),
                name,
                *[values.get(key, "") for key in keys],
                values.get(total_key, ""),
                values.get("remark", ""),
            ]
        )


def _export_manager_overtime_workbook(year: int):
    values_by_name = _manager_overtime_values(year)
    month_keys = _manager_export_months(year)
    template_path = "/home/lewis/文档/考勤/加班单查询表.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["部门", "姓名", *[label for _month, label in month_keys], "剩余调休天数", "备注"])
    ws.cell(1, 3).value = "前年累积天数"
    _fill_named_month_template(ws, values_by_name, "remaining")

    output = BytesIO()
    try:
        wb.save(output)
    finally:
        wb.close()
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员加班信息_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _export_manager_annual_leave_workbook(year: int):
    values_by_name = _manager_annual_leave_values(year)
    template_path = "/home/lewis/文档/考勤/加班单查询表.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        ws.delete_cols(3)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"])
    ws.title = "年休统计表"
    ws.cell(1, 15).value = "剩余年休天数"
    ws.cell(1, 16).value = "备注"
    _fill_named_month_template(ws, values_by_name, "remaining", _annual_leave_value_keys())

    output = BytesIO()
    try:
        wb.save(output)
    finally:
        wb.close()
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员年休信息_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def manager_annual_leave_records():
    year = request.args.get("year", type=int) or datetime.now().year
    return jsonify(
        _manager_month_rows(
            _manager_annual_leave_values(year, _requested_emp_ids() or None),
            "剩余年休天数",
            _annual_leave_value_keys(),
        )
    )


def update_manager_annual_leave_record():
    data = request.json or {}
    if any(key in data for key in _annual_leave_value_keys()) or "remaining" in data:
        payload, status = _save_manager_month_stat("annual_leave")
        return jsonify(payload), status

    emp_id = int(data.get("emp_id") or 0)
    year = int(data.get("year") or datetime.now().year)
    locked_error = _ensure_year_months_unlocked(year, "修改管理人员年休统计")
    if locked_error:
        return locked_error
    employee = _require_model(Employee, emp_id)
    if not employee.is_manager:
        return jsonify({"error": "employee is not manager"}), 400
    row = AnnualLeave.query.filter_by(emp_id=employee.id, year=year).first()
    if not row:
        row = AnnualLeave(emp_id=employee.id, year=year)
        db.session.add(row)
    row.total_days = float(data.get("total_days") or 0)
    row.used_days = float(data.get("used_days") or 0)
    row.remaining_days = float(data.get("remaining_days") or 0)
    db.session.commit()
    return jsonify({"status": "ok", "id": row.id})


def create_department():
    data = request.json or {}
    dept_no = (data.get("dept_no") or "").strip()
    dept_name = (data.get("dept_name") or "").strip()
    parent_id = _parse_parent_id(data.get("parent_id"))
    is_locked = bool(data.get("is_locked"))
    if not dept_no or not dept_name:
        return jsonify({"error": "dept_no and dept_name are required"}), 400
    if parent_id == -1:
        return jsonify({"error": "parent_id must be integer"}), 400
    if Department.query.filter_by(dept_no=dept_no).first():
        return jsonify({"error": "dept_no already exists"}), 400
    parent, err = _validate_parent_department(parent_id)
    if err:
        return jsonify({"error": err}), 400

    department = Department(
        dept_no=dept_no,
        dept_name=dept_name,
        parent_id=parent.id if parent else None,
        is_locked=is_locked,
    )
    db.session.add(department)
    db.session.commit()
    return jsonify({"status": "ok", "id": department.id})


def update_department(dept_id: int):
    data = request.json or {}
    dept_no = (data.get("dept_no") or "").strip()
    dept_name = (data.get("dept_name") or "").strip()
    parent_id = _parse_parent_id(data.get("parent_id"))
    is_locked = bool(data.get("is_locked"))
    if not dept_no or not dept_name:
        return jsonify({"error": "dept_no and dept_name are required"}), 400
    if parent_id == -1:
        return jsonify({"error": "parent_id must be integer"}), 400

    department = _require_model(Department, dept_id)
    duplicate = Department.query.filter(Department.dept_no == dept_no, Department.id != dept_id).first()
    if duplicate:
        return jsonify({"error": "dept_no already exists"}), 400
    parent, err = _validate_parent_department(parent_id, dept_id)
    if err:
        return jsonify({"error": err}), 400

    department.dept_no = dept_no
    department.dept_name = dept_name
    department.parent_id = parent.id if parent else None
    department.is_locked = is_locked
    db.session.commit()
    return jsonify({"status": "ok"})


def delete_department(dept_id: int):
    department = _require_model(Department, dept_id)
    if department.children:
        return jsonify({"error": "该部门存在下级部门，无法删除"}), 400
    if department.employees:
        return jsonify({"error": "该部门已绑定员工，无法删除"}), 400
    if department.user_assignments:
        return jsonify({"error": "该部门已绑定账号权限，无法删除"}), 400

    db.session.delete(department)
    db.session.commit()
    return jsonify({"status": "ok"})


def batch_operate_departments():
    data = request.json or {}
    ids = data.get("ids") or []
    action = (data.get("action") or "").strip()

    if action not in {"delete", "set_parent", "lock", "unlock"}:
        return jsonify({"error": "unsupported action"}), 400
    if not ids:
        return jsonify({"error": "ids are required"}), 400

    departments = Department.query.filter(Department.id.in_(ids)).all()
    if not departments:
        return jsonify({"error": "departments not found"}), 404

    if action == "delete":
        blocked = []
        for department in departments:
            if department.children:
                blocked.append(f"{department.dept_name}（存在下级部门）")
            elif department.employees:
                blocked.append(f"{department.dept_name}（已绑定员工）")
            elif department.user_assignments:
                blocked.append(f"{department.dept_name}（已绑定账号权限）")
        if blocked:
            return jsonify({"error": f"以下部门不可删除：{', '.join(blocked)}"}), 400

        for department in departments:
            db.session.delete(department)
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(departments)})

    if action in {"lock", "unlock"}:
        locked = action == "lock"
        for department in departments:
            department.is_locked = locked
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(departments)})

    parent_id = _parse_parent_id(data.get("parent_id"))
    if parent_id == -1:
        return jsonify({"error": "parent_id must be integer"}), 400
    selected_ids = {d.id for d in departments}
    if parent_id in selected_ids:
        return jsonify({"error": "上级部门不能选择已选部门"}), 400
    parent, err = _validate_parent_department(parent_id)
    if err:
        return jsonify({"error": err}), 400

    for department in departments:
        if parent:
            cursor = parent
            while cursor:
                if cursor.id == department.id:
                    return jsonify({"error": f"{department.dept_name} 的上级部门设置会形成循环层级"}), 400
                cursor = cursor.parent
        department.parent_id = parent.id if parent else None
    db.session.commit()
    return jsonify({"status": "ok", "action": action, "affected": len(departments)})


def delete_unbound_departments():
    all_departments = Department.query.all()
    deleted = 0
    skipped_locked = 0
    skipped_employee_bound = 0
    skipped_account_bound = 0

    for department in all_departments:
        if department.children:
            continue
        if department.is_locked:
            skipped_locked += 1
            continue
        if department.employees:
            skipped_employee_bound += 1
            continue
        if department.user_assignments:
            skipped_account_bound += 1
            continue
        db.session.delete(department)
        deleted += 1

    db.session.commit()
    return jsonify(
        {
            "status": "ok",
            "deleted": deleted,
            "skipped_locked": skipped_locked,
            "skipped_employee_bound": skipped_employee_bound,
            "skipped_account_bound": skipped_account_bound,
        }
    )


def _build_departments_workbook(
    rows: list[tuple[str, str, str, str]],
    include_identity_metadata: bool = False,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "部门导入模板"
    ws.append(["部门编号", "部门名称", "上级部门编号", _DEPARTMENT_ORIGINAL_ID_HEADER])
    ws.column_dimensions["D"].hidden = True

    for row in rows:
        ws.append(list(row[:4]))

    if include_identity_metadata:
        metadata_ws = wb.create_sheet(_DEPARTMENT_METADATA_SHEET)
        metadata_ws.sheet_state = "hidden"
        metadata_ws.append(
            [
                _DEPARTMENT_ORIGINAL_ID_HEADER,
                _DEPARTMENT_ORIGINAL_DEPT_NO_HEADER,
                _DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER,
            ]
        )
        for row in rows:
            metadata_ws.append([row[3], row[0], row[1]])
    return wb


def _card_no_conflict(card_no: str, exclude_employee_id: int | None = None) -> bool:
    if not card_no:
        return False
    query = Employee.query.filter_by(card_no=card_no)
    if exclude_employee_id is not None:
        query = query.filter(Employee.id != exclude_employee_id)
    return query.first() is not None


def create_employee():
    data = request.json or {}
    emp_no = (data.get("emp_no") or "").strip()
    name = (data.get("name") or "").strip()
    card_no = (data.get("card_no") or "").strip() or None
    dept_name = (data.get("dept_name") or "").strip()
    shift_no = (data.get("shift_no") or "").strip()
    is_manager = bool(data.get("is_manager"))
    is_nursing = bool(data.get("is_nursing"))
    employee_stats_attendance_source = _parse_attendance_source(
        data.get("employee_stats_attendance_source"), ATTENDANCE_SOURCE_EMPLOYEE
    )
    manager_stats_attendance_source = _parse_attendance_source(
        data.get("manager_stats_attendance_source"), ATTENDANCE_SOURCE_MANAGER
    )

    if not emp_no or not name:
        return jsonify({"error": "emp_no and name are required"}), 400
    if Employee.query.filter_by(emp_no=emp_no).first():
        return jsonify({"error": "emp_no already exists"}), 400
    if _card_no_conflict(card_no):
        return jsonify({"error": "card_no already exists"}), 400

    department = _resolve_department(dept_name) if dept_name else None
    employee = Employee(
        emp_no=emp_no,
        name=name,
        card_no=card_no,
        dept_id=department.id if department else None,
        is_manager=is_manager,
        is_nursing=is_nursing,
        employee_stats_attendance_source=employee_stats_attendance_source,
        manager_stats_attendance_source=manager_stats_attendance_source,
    )
    db.session.add(employee)
    db.session.flush()
    _assign_employee_shift(employee, _resolve_shift(shift_no))
    db.session.commit()
    return jsonify({"status": "ok", "employee": _serialize_employee(employee)})


def update_employee(employee_id: int):
    data = request.json or {}
    emp_no = (data.get("emp_no") or "").strip()
    name = (data.get("name") or "").strip()
    card_no = (data.get("card_no") or "").strip() or None
    dept_name = (data.get("dept_name") or "").strip()
    shift_no = (data.get("shift_no") or "").strip()
    is_manager = bool(data.get("is_manager"))
    is_nursing = data.get("is_nursing")
    employee_stats_attendance_source = data.get("employee_stats_attendance_source")
    manager_stats_attendance_source = data.get("manager_stats_attendance_source")
    if is_nursing is not None:
        is_nursing = bool(is_nursing)

    employee = _require_model(Employee, employee_id)

    if not emp_no or not name:
        return jsonify({"error": "emp_no and name are required"}), 400

    duplicate = Employee.query.filter(Employee.emp_no == emp_no, Employee.id != employee_id).first()
    if duplicate:
        return jsonify({"error": "emp_no already exists"}), 400
    if _card_no_conflict(card_no, exclude_employee_id=employee_id):
        return jsonify({"error": "card_no already exists"}), 400

    employee.emp_no = emp_no
    employee.name = name
    employee.card_no = card_no
    employee.is_manager = is_manager
    if is_nursing is not None:
        employee.is_nursing = is_nursing
    employee.employee_stats_attendance_source = _parse_attendance_source(
        employee_stats_attendance_source, employee.employee_stats_attendance_source or ATTENDANCE_SOURCE_EMPLOYEE
    )
    employee.manager_stats_attendance_source = _parse_attendance_source(
        manager_stats_attendance_source, employee.manager_stats_attendance_source or ATTENDANCE_SOURCE_MANAGER
    )
    if dept_name:
        department = _resolve_department(dept_name)
        employee.dept_id = department.id if department else None
    else:
        employee.dept_id = None
    _assign_employee_shift(employee, _resolve_shift(shift_no))

    db.session.commit()
    return jsonify({"status": "ok", "employee": _serialize_employee(employee)})


def delete_employee(employee_id: int):
    employee = _require_model(Employee, employee_id)
    db.session.delete(employee)
    db.session.commit()
    return jsonify({"status": "ok"})


def batch_operate_employees():
    data = request.json or {}
    ids = data.get("ids") or []
    action = (data.get("action") or "").strip()

    if not ids:
        return jsonify({"error": "ids are required"}), 400

    employees = Employee.query.filter(Employee.id.in_(ids)).all()
    if not employees:
        return jsonify({"error": "employees not found"}), 404

    if action == "delete":
        for employee in employees:
            db.session.delete(employee)
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_department":
        dept_name = (data.get("dept_name") or "").strip()
        department = _resolve_department(dept_name) if dept_name else None
        for employee in employees:
            employee.dept_id = department.id if department else None
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_shift":
        shift_no = (data.get("shift_no") or "").strip()
        shift = _resolve_shift(shift_no) if shift_no else None
        for employee in employees:
            _assign_employee_shift(employee, shift)
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_manager":
        is_manager = bool(data.get("is_manager"))
        for employee in employees:
            employee.is_manager = is_manager
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_nursing":
        is_nursing = bool(data.get("is_nursing"))
        for employee in employees:
            employee.is_nursing = is_nursing
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_employee_stats_attendance_source":
        source = _parse_attendance_source(data.get("employee_stats_attendance_source"), ATTENDANCE_SOURCE_EMPLOYEE)
        for employee in employees:
            employee.employee_stats_attendance_source = source
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_manager_stats_attendance_source":
        source = _parse_attendance_source(data.get("manager_stats_attendance_source"), ATTENDANCE_SOURCE_MANAGER)
        for employee in employees:
            employee.manager_stats_attendance_source = source
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_name":
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name is required"}), 400
        if len(employees) != 1:
            return jsonify({"error": "set_name only supports single selected employee"}), 400
        employees[0].name = name
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": 1})

    if action == "set_emp_no":
        emp_no = (data.get("emp_no") or "").strip()
        if not emp_no:
            return jsonify({"error": "emp_no is required"}), 400
        if len(employees) != 1:
            return jsonify({"error": "set_emp_no only supports single selected employee"}), 400
        duplicate = Employee.query.filter(Employee.emp_no == emp_no, Employee.id != employees[0].id).first()
        if duplicate:
            return jsonify({"error": "emp_no already exists"}), 400
        employees[0].emp_no = emp_no
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": 1})

    return jsonify({"error": "unsupported action"}), 400



def _employee_override_values(override: EmployeeAttendanceOverride | None) -> dict[str, float | int | None]:
    return {field: getattr(override, field) if override else None for field in _EMPLOYEE_OVERRIDE_FIELDS}


def _employee_automatic_row(emp_id: int, month: str) -> dict[str, object] | None:
    return _employee_automatic_rows_by_emp([emp_id], month).get(emp_id)


def _employee_automatic_rows_by_emp(emp_ids: list[int], month: str) -> dict[int, dict[str, object]]:
    from routes.query_core import _build_final_rows

    if not emp_ids:
        return {}
    # include_overrides=False：automatic 为「系统原始 + 逐日修正」，不含月度修正（月度优先级最高，applied 单独体现）
    # _build_final_rows 接收 emp_id 列表，但返回行内只有 emp_no（唯一），故用 emp_no 反查 emp_id
    emp_no_by_id = {e.id: e.emp_no for e in Employee.query.filter(Employee.id.in_(emp_ids)).all()}
    rows_by_emp_no = {row[1]: row for row in _build_final_rows(month, emp_ids, include_overrides=False)}
    late_by_emp = _employee_late_early_minutes_by_emp(emp_ids, month)
    result: dict[int, dict[str, object]] = {}
    for emp_id, emp_no in emp_no_by_id.items():
        row = rows_by_emp_no.get(emp_no)
        if not row:
            continue
        result[emp_id] = {
            "attendance_days": row[3],
            "actual_attendance_days": row[4],
            "work_hours": row[17],
            "half_days": row[18],
            "late_early_minutes": late_by_emp.get(emp_id, 0),
        }
    return result


def _employee_late_early_minutes(emp_id: int, month: str) -> int:
    return _employee_late_early_minutes_by_emp([emp_id], month).get(emp_id, 0)


def _employee_late_early_minutes_by_emp(emp_ids: list[int], month: str) -> dict[int, int]:
    from routes.query_core import _month_date_range
    from services.daily_override_service import (
        daily_override_maps,
        effective_early_leave_minutes,
        effective_late_minutes,
    )

    if not emp_ids:
        return {}
    date_range = _month_date_range(month)
    if not date_range:
        return {emp_id: 0 for emp_id in emp_ids}
    start_date, end_date = date_range
    records_by_emp: dict[int, list[DailyRecord]] = defaultdict(list)
    for record in (
        DailyRecord.query.filter(DailyRecord.emp_id.in_(emp_ids))
        .filter(DailyRecord.record_date >= start_date, DailyRecord.record_date < end_date)
        .all()
    ):
        records_by_emp[record.emp_id].append(record)
    overrides_by_emp = daily_override_maps(month, emp_ids)
    totals: dict[int, int] = {}
    for emp_id in emp_ids:
        overrides = overrides_by_emp.get(emp_id, {})
        total = 0
        seen_dates = set()
        for record in records_by_emp.get(emp_id, []):
            override = overrides.get(record.record_date)
            total += effective_late_minutes(record.late_minutes, override)
            total += effective_early_leave_minutes(record.early_leave_minutes, override)
            seen_dates.add(record.record_date)
        # 修正表有记录但无 DailyRecord 的日期（缺勤日补修正）也计入
        for day, override in overrides.items():
            if day not in seen_dates:
                total += int(override.late_minutes or 0) + int(override.early_leave_minutes or 0)
        totals[emp_id] = total
    return totals


def _employee_override_payload(row: EmployeeAttendanceOverride | None) -> dict[str, object]:
    payload = {field: getattr(row, field) if row else None for field in _EMPLOYEE_OVERRIDE_FIELDS}
    payload["remark"] = row.remark if row else ""
    payload["updated_at"] = row.updated_at.isoformat() if row and row.updated_at else None
    payload["updated_by_name"] = _user_display_name(row.updated_by_user) if row else ""
    return payload


def _employee_override_response(emp_id: int, month: str) -> tuple[dict[str, object], int]:
    employee = db.session.get(Employee, emp_id)
    if not employee or employee.is_manager:
        return {"error": "employee is a manager, not a regular employee"}, 400
    automatic = _employee_automatic_row(emp_id, month)
    override = EmployeeAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    override_data = _employee_override_values(override)
    applied: dict[str, object] = {}
    if automatic:
        for field in _EMPLOYEE_OVERRIDE_FIELDS:
            applied[field] = override_data[field] if override_data[field] is not None else automatic.get(field)
    return {
        "employee": _serialize_employee(employee),
        "month": month,
        "automatic": automatic,
        "override": _employee_override_payload(override),
        "applied": applied,
        "history": _history_rows_for_month("employee", month),
    }, 200


def _employee_override_list_response(emp_ids: list[int], month: str) -> tuple[dict[str, object], int]:
    if not emp_ids:
        emp_ids = list(_accessible_emp_ids_set())
        if not emp_ids:
            return {"rows": [], "month": month}, 200
    employees = {
        employee.id: employee
        for employee in Employee.query.filter(
            Employee.id.in_(emp_ids),
            Employee.is_manager.is_(False),
        ).all()
    }
    overrides = {
        row.emp_id: row
        for row in EmployeeAttendanceOverride.query.filter(
            EmployeeAttendanceOverride.month == month,
            EmployeeAttendanceOverride.emp_id.in_(emp_ids),
        ).all()
    }
    rows: list[dict[str, object]] = []
    automatic_by_emp = _employee_automatic_rows_by_emp(list(employees.keys()), month)
    for emp_id in emp_ids:
        employee = employees.get(emp_id)
        if not employee:
            continue
        override = overrides.get(emp_id)
        override_data = _employee_override_values(override)
        automatic = automatic_by_emp.get(emp_id)
        applied: dict[str, object] = {}
        if automatic:
            for field in _EMPLOYEE_OVERRIDE_FIELDS:
                applied[field] = override_data[field] if override_data[field] is not None else automatic.get(field)
        rows.append(
            {
                "employee": _serialize_employee(employee),
                "automatic": automatic,
                "override": _employee_override_payload(override),
                "applied": applied,
            }
        )
    return {"rows": rows, "month": month}, 200

