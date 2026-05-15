from __future__ import annotations

import os
import subprocess
from io import BytesIO
from datetime import datetime
from typing import Any

from flask import Blueprint, abort, jsonify, redirect, render_template, request, send_file, g
import openpyxl

from models import db
from models.department import Department
from models.employee import Employee
from models.employee import (
    ATTENDANCE_SOURCE_AUTO_FALLBACK,
    ATTENDANCE_SOURCE_EMPLOYEE,
    ATTENDANCE_SOURCE_MANAGER,
    ATTENDANCE_SOURCE_VALUES,
)
from models.employee_shift import EmployeeShiftAssignment
from models.shift import Shift
from models.daily_record import DailyRecord
from models.account_set import AccountSet, AccountSetImport
from models.overtime import OvertimeRecord
from models.annual_leave import AnnualLeave
from models.manager_month_stat import ManagerMonthStat
from models.manager_attendance_override import ManagerAttendanceOverride
from models.employee_attendance_override import EmployeeAttendanceOverride
from models.attendance_override_history import AttendanceOverrideHistory
from models.user import (
    ALL_PAGE_PERMISSION_KEYS,
    EMPLOYEE_PAGE_PERMISSION_KEYS,
    HOME_PAGE_PERMISSION_KEYS,
    MANAGER_PAGE_PERMISSION_KEYS,
    PAGE_PERMISSION_LABELS,
    User,
    UserEmployeeAssignment,
    UserDepartmentAssignment,
)
from services.import_service import ImportService
from services.manager_attendance_service import ManagerAttendanceOptions, build_manager_rows
from routes.auth import admin_required
from utils.helpers import parse_bool_zh


admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


_EMPLOYEE_OVERRIDE_FIELDS = (
    "attendance_days",
    "work_hours",
    "half_days",
    "late_early_minutes",
)

_EMPLOYEE_OVERRIDE_LABELS = {
    "attendance_days": "考勤天数",
    "work_hours": "工时",
    "half_days": "半勤天数",
    "late_early_minutes": "迟到早退",
    "remark": "备注",
}

_MANAGER_ATTENDANCE_OVERRIDE_FIELDS = (
    "attendance_days",
    "injury_days",
    "business_trip_days",
    "marriage_days",
    "funeral_days",
    "late_early_minutes",
)

_MANAGER_OVERRIDE_LABELS = {
    "attendance_days": "出勤天数",
    "injury_days": "工伤",
    "business_trip_days": "出差",
    "marriage_days": "婚假",
    "funeral_days": "丧假",
    "late_early_minutes": "迟到早退",
    "remark": "备注",
}

_DEPARTMENT_ORIGINAL_ID_HEADER = "原始部门ID"
_DEPARTMENT_ORIGINAL_DEPT_NO_HEADER = "原始部门编号"
_DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER = "原始部门名称"
_DEPARTMENT_METADATA_SHEET = "部门导入元数据"
_DEPARTMENT_IMPORT_TEMP_PREFIX = "__IMPORT_TMP__"


def _require_model(model, ident):
    row = db.session.get(model, ident)
    if row is None:
        abort(404)
    return row


def _manager_scope_employees():
    return (
        Employee.query.filter(Employee.is_manager.is_(True))
        .order_by(Employee.dept_id.asc(), Employee.emp_no.asc(), Employee.name.asc())
        .all()
    )


def _convert_uploaded_xls_to_xlsx(xls_path: str) -> str | None:
    xlsx_path = f"{os.path.splitext(xls_path)[0]}.xlsx"
    try:
        subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                os.path.dirname(xls_path),
                xls_path,
            ],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except Exception:
        return None
    return xlsx_path if os.path.exists(xlsx_path) else None


def _serialize_user(user: User) -> dict:
    profile_department = db.session.get(Department, user.profile_dept_id) if user.profile_dept_id else None
    return {
        "id": user.id,
        "username": user.username,
        "profile_emp_no": user.profile_emp_no or "",
        "profile_name": user.profile_name or "",
        "profile_dept_id": user.profile_dept_id,
        "profile_department": {
            "id": profile_department.id,
            "dept_no": profile_department.dept_no,
            "dept_name": profile_department.dept_name,
        } if profile_department else None,
        "role": user.role,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "page_permissions": user.effective_page_permissions(),
        "emp_ids": [a.emp_id for a in user.employee_assignments],
        "dept_ids": [a.dept_id for a in user.department_assignments],
        "employees": [
            {
                "id": a.employee.id,
                "emp_no": a.employee.emp_no,
                "name": a.employee.name,
                "dept_id": a.employee.dept_id,
                "dept_no": a.employee.department.dept_no if a.employee.department else "",
                "dept_name": a.employee.department.dept_name if a.employee.department else "",
            }
            for a in user.employee_assignments
            if a.employee
        ],
        "departments": [
            {
                "id": a.department.id,
                "dept_no": a.department.dept_no,
                "dept_name": a.department.dept_name,
                "parent_id": a.department.parent_id,
            }
            for a in user.department_assignments
            if a.department
        ],
    }


def _bind_user_profile_identity(user: User, emp_ids: list[int]) -> None:
    if user.profile_emp_no and user.profile_name and user.profile_dept_id:
        return

    first_emp_id = next((emp_id for emp_id in emp_ids if isinstance(emp_id, int)), None)
    if first_emp_id is None:
        return

    employee = db.session.get(Employee, first_emp_id)
    if not employee:
        return

    user.profile_emp_no = employee.emp_no or user.profile_emp_no
    user.profile_name = employee.name or user.profile_name
    user.profile_dept_id = employee.dept_id or user.profile_dept_id


def _default_page_permissions_for_role(role: str) -> dict[str, bool]:
    if role == "admin":
        return {key: True for key in ALL_PAGE_PERMISSION_KEYS}
    return {key: True for key in ALL_PAGE_PERMISSION_KEYS}


def _parse_page_permissions(data: dict | None, role: str, existing_user: User | None = None) -> dict[str, bool]:
    if role == "admin":
        return {key: True for key in ALL_PAGE_PERMISSION_KEYS}

    if not data or "page_permissions" not in data:
        if existing_user and isinstance(existing_user.page_permissions, dict):
            return existing_user.effective_page_permissions()
        return _default_page_permissions_for_role(role)

    raw = data.get("page_permissions") or {}
    if not isinstance(raw, dict):
        return _default_page_permissions_for_role(role)
    return {key: bool(raw.get(key, False)) for key in ALL_PAGE_PERMISSION_KEYS}


def _sync_user_assignments(user: User, emp_ids: list[int]) -> None:
    valid_ids = {
        row.id for row in Employee.query.with_entities(Employee.id).filter(Employee.id.in_(emp_ids)).all()
    }
    current_ids = {a.emp_id for a in user.employee_assignments}
    to_remove = [a for a in user.employee_assignments if a.emp_id not in valid_ids]
    for assignment in to_remove:
        db.session.delete(assignment)
    for emp_id in valid_ids - current_ids:
        db.session.add(UserEmployeeAssignment(user_id=user.id, emp_id=emp_id))


def _sync_user_department_assignments(user: User, dept_ids: list[int]) -> None:
    valid_ids = {
        row.id for row in Department.query.with_entities(Department.id).filter(Department.id.in_(dept_ids)).all()
    }
    current_ids = {a.dept_id for a in user.department_assignments}
    to_remove = [a for a in user.department_assignments if a.dept_id not in valid_ids]
    for assignment in to_remove:
        db.session.delete(assignment)
    for dept_id in valid_ids - current_ids:
        db.session.add(UserDepartmentAssignment(user_id=user.id, dept_id=dept_id))


def _serialize_employee(employee: Employee) -> dict:
    shift = employee.shift_assignment.shift if employee.shift_assignment else None
    return {
        "id": employee.id,
        "emp_no": employee.emp_no,
        "name": employee.name,
        "is_manager": bool(employee.is_manager),
        "is_nursing": bool(employee.is_nursing),
        "employee_stats_attendance_source": employee.employee_stats_attendance_source or ATTENDANCE_SOURCE_EMPLOYEE,
        "manager_stats_attendance_source": employee.manager_stats_attendance_source or ATTENDANCE_SOURCE_MANAGER,
        "dept_id": employee.dept_id,
        "dept_no": employee.department.dept_no if employee.department else "",
        "dept_name": employee.department.dept_name if employee.department else "",
        "department": employee.department.dept_name if employee.department else "",
        "shift_id": shift.id if shift else None,
        "shift_no": shift.shift_no if shift else "",
        "shift_name": shift.shift_name if shift else "",
    }


def _parse_attendance_source(value: Any, default: str) -> str:
    source = str(value or "").strip() or default
    if source not in ATTENDANCE_SOURCE_VALUES:
        return default
    return source


def _parse_department_original_id(value: Any) -> int | None:
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    try:
        return int(float(raw_value))
    except (TypeError, ValueError, OverflowError):
        return None


def _load_department_identity_metadata(wb: openpyxl.Workbook) -> dict[int, dict[str, str]]:
    if _DEPARTMENT_METADATA_SHEET not in wb.sheetnames:
        return {}

    metadata_ws = wb[_DEPARTMENT_METADATA_SHEET]
    metadata_rows = [list(r) for r in metadata_ws.iter_rows(values_only=True)]
    if not metadata_rows:
        return {}

    header_idx, header_map = _parse_header_row(
        metadata_rows,
        [
            _DEPARTMENT_ORIGINAL_ID_HEADER,
            _DEPARTMENT_ORIGINAL_DEPT_NO_HEADER,
            _DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER,
        ],
    )
    original_id_idx = header_map.get(_DEPARTMENT_ORIGINAL_ID_HEADER, -1)
    original_dept_no_idx = header_map.get(_DEPARTMENT_ORIGINAL_DEPT_NO_HEADER, -1)
    original_dept_name_idx = header_map.get(_DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER, -1)
    if original_id_idx < 0 or original_dept_no_idx < 0 or original_dept_name_idx < 0:
        return {}

    identities_by_id: dict[int, dict[str, str]] = {}
    for row in metadata_rows[header_idx + 1 :]:
        original_id = (
            _parse_department_original_id(row[original_id_idx])
            if original_id_idx < len(row)
            else None
        )
        if original_id is None:
            continue
        original_dept_no = (
            str(row[original_dept_no_idx]).strip()
            if original_dept_no_idx < len(row) and row[original_dept_no_idx] is not None
            else ""
        )
        original_dept_name = (
            str(row[original_dept_name_idx]).strip()
            if original_dept_name_idx < len(row) and row[original_dept_name_idx] is not None
            else ""
        )
        if not original_dept_no or not original_dept_name:
            continue
        identities_by_id[original_id] = {
            "dept_no": original_dept_no,
            "dept_name": original_dept_name,
        }
    return identities_by_id


def _department_matches_original_identity(
    department: Department | None,
    original_identity: dict[str, str] | None,
) -> bool:
    if not department or not original_identity:
        return False
    return (
        (department.dept_no or "").strip() == original_identity["dept_no"]
        and (department.dept_name or "").strip() == original_identity["dept_name"]
    )


def _serialize_account_set(row: AccountSet) -> dict:
    success_count = 0
    error_count = 0
    pending_count = 0
    latest_import_at = None
    for item in row.imports:
        if item.status == "ok":
            success_count += 1
        elif item.status == "error":
            error_count += 1
        else:
            pending_count += 1
        if item.created_at and (latest_import_at is None or item.created_at > latest_import_at):
            latest_import_at = item.created_at

    return {
        "id": row.id,
        "month": row.month,
        "name": row.name,
        "is_active": row.is_active,
        "is_locked": bool(row.is_locked),
        "locked_at": row.locked_at.isoformat() if row.locked_at else None,
        "locked_by": row.locked_by,
        "factory_rest_days": row.factory_rest_days or 0,
        "monthly_benefit_days": row.monthly_benefit_days or 0,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "imports_count": len(row.imports),
        "pending_count": pending_count,
        "success_count": success_count,
        "error_count": error_count,
        "latest_import_at": latest_import_at.isoformat() if latest_import_at else None,
    }


def _account_set_file_type(filename: str) -> str:
    if "加班" in filename:
        return "overtime"
    if "请假" in filename:
        return "leave"
    if "管理人员" in filename and "月报" in filename:
        return "manager_monthly"
    if "管理人员" in filename:
        return "manager_daily"
    if "月报" in filename:
        return "monthly"
    return "daily"


def _account_set_for_month(month: str | None) -> AccountSet | None:
    key = (month or "").strip()
    if not key:
        return None
    return AccountSet.query.filter_by(month=key).first()


def _locked_account_set_error(account_set: AccountSet, action_label: str):
    return jsonify({"error": f"{account_set.month} 账套已锁定，不能{action_label}"}), 400


def _ensure_account_set_unlocked(account_set: AccountSet | None, action_label: str):
    if account_set and account_set.is_locked:
        return _locked_account_set_error(account_set, action_label)
    return None


def _user_display_name(user: User | None) -> str:
    return user.username if user else ""


def _override_field_labels(override_type: str) -> dict[str, str]:
    return _EMPLOYEE_OVERRIDE_LABELS if override_type == "employee" else _MANAGER_OVERRIDE_LABELS


def _override_field_names(override_type: str) -> tuple[str, ...]:
    return _EMPLOYEE_OVERRIDE_FIELDS if override_type == "employee" else _MANAGER_ATTENDANCE_OVERRIDE_FIELDS


def _override_state_from_row(row: object | None, fields: tuple[str, ...]) -> dict[str, object]:
    state = {field: getattr(row, field) if row else None for field in fields}
    state["remark"] = row.remark if row else ""
    return state


def _has_override_state_changes(before: dict[str, object], after: dict[str, object]) -> bool:
    keys = set(before.keys()) | set(after.keys())
    return any(before.get(key) != after.get(key) for key in keys)


def _changed_override_fields(before: dict[str, object], after: dict[str, object]) -> list[str]:
    keys = [key for key in after.keys() if key in before or key == "remark"]
    return [key for key in keys if before.get(key) != after.get(key)]


def _record_override_history(
    override_type: str,
    emp_id: int,
    month: str,
    action_type: str,
    before_values: dict[str, object],
    after_values: dict[str, object],
    source_file_name: str | None = None,
) -> None:
    history = AttendanceOverrideHistory(
        override_type=override_type,
        emp_id=emp_id,
        month=month,
        action_type=action_type,
        changed_fields_json=_changed_override_fields(before_values, after_values),
        before_values_json=before_values,
        after_values_json=after_values,
        remark=str(after_values.get("remark") or ""),
        source_file_name=(source_file_name or "").strip() or None,
        operator_user_id=g.current_user.id if getattr(g, "current_user", None) else None,
    )
    db.session.add(history)


def _serialize_override_history(row: AttendanceOverrideHistory) -> dict[str, object]:
    labels = _override_field_labels(row.override_type)
    before = row.before_values_json if isinstance(row.before_values_json, dict) else {}
    after = row.after_values_json if isinstance(row.after_values_json, dict) else {}
    changed_fields = list(row.changed_fields_json or [])
    changes = [
        {
            "field": field,
            "label": labels.get(field, field),
            "before": before.get(field),
            "after": after.get(field),
        }
        for field in changed_fields
    ]
    return {
        "id": row.id,
        "override_type": row.override_type,
        "action_type": row.action_type,
        "month": row.month,
        "emp_id": row.emp_id,
        "emp_no": row.employee.emp_no if row.employee else "",
        "employee_name": row.employee.name if row.employee else "",
        "remark": row.remark or "",
        "source_file_name": row.source_file_name or "",
        "operator_name": _user_display_name(row.operator_user),
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "changes": changes,
    }


def _history_rows(override_type: str, emp_id: int, month: str) -> list[dict[str, object]]:
    rows = (
        AttendanceOverrideHistory.query.filter_by(override_type=override_type, emp_id=emp_id, month=month)
        .order_by(AttendanceOverrideHistory.created_at.desc(), AttendanceOverrideHistory.id.desc())
        .all()
    )
    return [_serialize_override_history(row) for row in rows]


def _history_rows_for_month(override_type: str, month: str) -> list[dict[str, object]]:
    rows = (
        AttendanceOverrideHistory.query.filter_by(override_type=override_type, month=month)
        .order_by(AttendanceOverrideHistory.created_at.desc(), AttendanceOverrideHistory.id.desc())
        .all()
    )
    return [_serialize_override_history(row) for row in rows]


def _locked_months_for_year(year: int, include_prev_dec: bool = False) -> list[str]:
    months = [f"{year}-{month:02d}" for month in range(1, 13)]
    if include_prev_dec:
        months.insert(0, f"{year - 1}-12")
    locked = []
    for month in months:
        account_set = _account_set_for_month(month)
        if account_set and account_set.is_locked:
            locked.append(month)
    return locked


def _ensure_year_months_unlocked(year: int, action_label: str, include_prev_dec: bool = False):
    locked_months = _locked_months_for_year(year, include_prev_dec=include_prev_dec)
    if locked_months:
        return jsonify({"error": f"{', '.join(locked_months)} 账套已锁定，不能{action_label}"}), 400
    return None


def _stat_key_month(year: int, stat_key: str) -> str | None:
    if stat_key == "prev_dec":
        return f"{year - 1}-12"
    if stat_key.startswith("m") and stat_key[1:].isdigit():
        month_no = int(stat_key[1:])
        if 1 <= month_no <= 12:
            return f"{year}-{month_no:02d}"
    return None


def _stat_key_lock_state(year: int, stat_key: str) -> str:
    month = _stat_key_month(year, stat_key)
    if not month:
        return "editable"
    account_set = _account_set_for_month(month)
    if not account_set:
        return "missing_account_set"
    return "locked" if account_set.is_locked else "editable"


def _next_auto_dept_no() -> str:
    index = Department.query.count() + 1
    while True:
        dept_no = f"AUTO-{index:04d}"
        if not Department.query.filter_by(dept_no=dept_no).first():
            return dept_no
        index += 1


def _resolve_department(dept_name: str | None, dept_no: str | None = None) -> Department | None:
    clean_name = (dept_name or "").strip()
    clean_no = (dept_no or "").strip()
    if not clean_name:
        return None

    if clean_no:
        existing_by_no = Department.query.filter_by(dept_no=clean_no).first()
        if existing_by_no:
            return existing_by_no
        department = Department(dept_no=clean_no, dept_name=clean_name)
        db.session.add(department)
        db.session.flush()
        return department

    existing_by_name = Department.query.filter_by(dept_name=clean_name).first()
    if existing_by_name:
        return existing_by_name

    department = Department(dept_no=_next_auto_dept_no(), dept_name=clean_name)
    db.session.add(department)
    db.session.flush()
    return department


def _parse_parent_id(raw: Any) -> int | None:
    if raw in (None, "", "null", "None"):
        return None
    if isinstance(raw, int):
        return raw
    text_value = str(raw).strip()
    if not text_value:
        return None
    if not text_value.isdigit():
        return -1
    return int(text_value)


def _validate_parent_department(parent_id: int | None, current_dept_id: int | None = None) -> tuple[Department | None, str | None]:
    if parent_id is None:
        return None, None
    parent = db.session.get(Department, parent_id)
    if not parent:
        return None, "上级部门不存在"
    if current_dept_id and parent.id == current_dept_id:
        return None, "上级部门不能选择自身"
    if current_dept_id:
        cursor = parent
        while cursor:
            if cursor.id == current_dept_id:
                return None, "上级部门设置非法，会形成循环层级"
            cursor = cursor.parent
    return parent, None


def _parse_header_row(rows: list[list[Any]], expected: list[str]) -> tuple[int, dict[str, int]]:
    best_idx = 0
    best_map: dict[str, int] = {}
    best_score = -1
    limit = min(len(rows), 8)
    for idx in range(limit):
        raw = rows[idx]
        header_map = {}
        for i, value in enumerate(raw):
            text = str(value).strip() if value is not None else ""
            if text:
                header_map[text] = i
        score = sum(1 for key in expected if key in header_map)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_map = header_map
    return best_idx, best_map


def _resolve_shift(shift_no: str | None) -> Shift | None:
    key = (shift_no or "").strip()
    if not key:
        return None
    return Shift.query.filter_by(shift_no=key).first()


def _assign_employee_shift(employee: Employee, shift: Shift | None) -> None:
    assignment = employee.shift_assignment
    if shift is None:
        if assignment:
            db.session.delete(assignment)
        return

    if not assignment:
        assignment = EmployeeShiftAssignment(emp_id=employee.id, shift_id=shift.id)
        db.session.add(assignment)
        employee.shift_assignment = assignment
    else:
        assignment.shift_id = shift.id


@admin_bp.route("/dashboard")
@admin_required
def dashboard():
    return render_template("admin/dashboard.html")


@admin_bp.route("/account-sets", methods=["GET"])
@admin_required
def list_account_sets():
    rows = AccountSet.query.order_by(AccountSet.month.desc()).all()
    return jsonify([_serialize_account_set(row) for row in rows])


@admin_bp.route("/account-sets", methods=["POST"])
@admin_required
def create_account_set():
    data = request.json or {}
    month = (data.get("month") or "").strip()
    factory_rest_days = request.json.get("factory_rest_days", 0) if request.json else 0
    monthly_benefit_days = request.json.get("monthly_benefit_days", 0) if request.json else 0
    if not month or len(month) != 7:
        return jsonify({"error": "month is required in YYYY-MM format"}), 400
    if AccountSet.query.filter_by(month=month).first():
        return jsonify({"error": "该月份账套已存在"}), 400

    row = AccountSet(
        month=month,
        name=f"{month} 账套",
        factory_rest_days=float(factory_rest_days or 0),
        monthly_benefit_days=float(monthly_benefit_days or 0),
    )
    if AccountSet.query.count() == 0:
        row.is_active = True
    db.session.add(row)
    db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


@admin_bp.route("/account-sets/<int:account_set_id>", methods=["PUT"])
@admin_required
def update_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "修改账套参数")
    if locked_error:
        return locked_error
    data = request.json or {}
    row.factory_rest_days = float(data.get("factory_rest_days") or 0)
    row.monthly_benefit_days = float(data.get("monthly_benefit_days") or 0)
    db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


@admin_bp.route("/account-sets/<int:account_set_id>/activate", methods=["POST"])
@admin_required
def activate_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    AccountSet.query.update({AccountSet.is_active: False})
    row.is_active = True
    db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


@admin_bp.route("/account-sets/<int:account_set_id>/lock", methods=["POST"])
@admin_required
def lock_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    if not row.is_locked:
        row.is_locked = True
        row.locked_at = datetime.utcnow()
        row.locked_by = g.current_user.id
        db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


@admin_bp.route("/account-sets/<int:account_set_id>/unlock", methods=["POST"])
@admin_required
def unlock_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    if row.is_locked:
        row.is_locked = False
        row.locked_at = None
        row.locked_by = None
        db.session.commit()
    return jsonify({"status": "ok", "account_set": _serialize_account_set(row)})


@admin_bp.route("/account-sets/<int:account_set_id>", methods=["DELETE"])
@admin_required
def delete_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "删除账套")
    if locked_error:
        return locked_error

    # best-effort cleanup archived files
    for record in row.imports:
        path = (record.stored_path or "").strip()
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass

    was_active = row.is_active
    db.session.delete(row)
    db.session.commit()

    if was_active:
        fallback = AccountSet.query.order_by(AccountSet.month.desc()).first()
        if fallback:
            AccountSet.query.update({AccountSet.is_active: False})
            fallback.is_active = True
            db.session.commit()

    return jsonify({"status": "ok"})


@admin_bp.route("/account-sets/<int:account_set_id>/calculate", methods=["POST"])
@admin_required
def calculate_account_set(account_set_id: int):
    row = _require_model(AccountSet, account_set_id)
    locked_error = _ensure_account_set_unlocked(row, "重新计算")
    if locked_error:
        return locked_error
    mode = (request.args.get("mode") or "all").strip()
    records_query = AccountSetImport.query.filter_by(account_set_id=row.id)
    if mode == "employee":
        records_query = records_query.filter(AccountSetImport.file_type.in_(["leave", "overtime", "monthly", "daily"]))
    elif mode == "manager":
        records_query = records_query.filter(AccountSetImport.file_type.in_(["leave", "overtime", "manager_monthly", "manager_daily"]))
    records = records_query.order_by(AccountSetImport.id.asc()).all()

    if not records:
        return jsonify({"status": "error", "message": "该账套暂无可计算文件", "mode": mode}), 400

    success = 0
    failed = 0
    results = []

    for rec in records:
        path = (rec.stored_path or "").strip()
        filename = rec.source_filename or os.path.basename(path)
        if not path or not os.path.exists(path):
            failed += 1
            rec.status = "error"
            rec.error_message = "archive file not found"
            rec.imported_count = 0
            results.append({"file": filename, "status": "error", "error": rec.error_message})
            db.session.commit()
            continue

        try:
            imported = ImportService.import_file(path)
            if imported.get("status") == "ok":
                success += 1
                rec.status = "ok"
                rec.file_type = imported.get("file_type")
                rec.imported_count = imported.get("imported", 0)
                rec.error_message = None
                results.append({"file": filename, "status": "ok", "result": imported})
            else:
                failed += 1
                rec.status = "error"
                rec.error_message = imported.get("message", "import failed")
                rec.imported_count = 0
                results.append({"file": filename, "status": "error", "error": rec.error_message, "result": imported})
        except Exception as exc:
            failed += 1
            rec.status = "error"
            rec.error_message = str(exc)
            rec.imported_count = 0
            results.append({"file": filename, "status": "error", "error": str(exc)})
        db.session.commit()

    manager_stats_sync = None
    if mode == "manager" and failed == 0:
        try:
            manager_options = ManagerAttendanceOptions(
                month=row.month,
                factory_rest_days=row.factory_rest_days or 0,
                monthly_benefit_days=row.monthly_benefit_days or 0,
            )
            manager_rows = build_manager_rows(manager_options, sync_month_stats=True)
            manager_stats_sync = _sync_manager_stats_from_manager_rows(row.month, manager_rows)
            if manager_stats_sync["error_count"]:
                failed += manager_stats_sync["error_count"]
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            failed += 1
            manager_stats_sync = {
                "month": row.month,
                "overtime_synced": 0,
                "annual_leave_synced": 0,
                "error_count": 1,
                "errors": [str(exc)],
            }

    return jsonify(
        {
            "status": "ok" if failed == 0 else "partial",
            "account_set": _serialize_account_set(row),
            "mode": mode,
            "total": len(records),
            "success": success,
            "failed": failed,
            "manager_stats_sync": manager_stats_sync,
            "results": results,
        }
    )


@admin_bp.route("/employees/manage")
@admin_required
def employees_page():
    shifts = Shift.query.order_by(Shift.shift_no.asc()).all()
    return render_template("admin/employees.html", shifts=shifts)


@admin_bp.route("/shifts/manage")
@admin_required
def shifts_page():
    return render_template("admin/shifts.html")


@admin_bp.route("/departments/manage")
@admin_required
def departments_page():
    return render_template("admin/departments.html")


@admin_bp.route("/manager-overtime")
@admin_required
def manager_overtime_page():
    return render_template("admin/manager_overtime.html")


@admin_bp.route("/manager-annual-leave")
@admin_required
def manager_annual_leave_page():
    return render_template("admin/manager_annual_leave.html")


@admin_bp.route("/shifts", methods=["POST"])
@admin_required
def create_shift():
    data = request.json or {}
    shift_no = (data.get("shift_no") or "").strip()
    shift_name = (data.get("shift_name") or "").strip()
    time_slots = data.get("time_slots") or []
    is_cross_day = bool(data.get("is_cross_day", False))

    if not shift_no or not shift_name:
        return jsonify({"error": "shift_no and shift_name are required"}), 400

    shift = Shift.query.filter_by(shift_no=shift_no).first()
    if shift:
        return jsonify({"error": "shift_no already exists"}), 400

    shift = Shift(
        shift_no=shift_no,
        shift_name=shift_name,
        time_slots=time_slots,
        is_cross_day=is_cross_day,
    )
    db.session.add(shift)

    db.session.commit()
    return jsonify({"status": "ok", "id": shift.id})


@admin_bp.route("/shifts", methods=["GET"])
@admin_required
def list_shifts():
    rows = Shift.query.order_by(Shift.shift_no.asc()).all()
    return jsonify(
        [
            {
                "id": s.id,
                "shift_no": s.shift_no,
                "shift_name": s.shift_name,
                "time_slots": s.time_slots or [],
                "is_cross_day": s.is_cross_day,
            }
            for s in rows
        ]
    )


@admin_bp.route("/shifts/<int:shift_id>", methods=["PUT"])
@admin_required
def update_shift(shift_id: int):
    data = request.json or {}
    shift_no = (data.get("shift_no") or "").strip()
    shift_name = (data.get("shift_name") or "").strip()
    time_slots = data.get("time_slots") or []
    is_cross_day = bool(data.get("is_cross_day", False))

    shift = _require_model(Shift, shift_id)
    if not shift_no or not shift_name:
        return jsonify({"error": "shift_no and shift_name are required"}), 400

    duplicate = Shift.query.filter(Shift.shift_no == shift_no, Shift.id != shift_id).first()
    if duplicate:
        return jsonify({"error": "shift_no already exists"}), 400

    shift.shift_no = shift_no
    shift.shift_name = shift_name
    shift.time_slots = time_slots
    shift.is_cross_day = is_cross_day
    db.session.commit()
    return jsonify({"status": "ok"})


@admin_bp.route("/shifts/<int:shift_id>", methods=["DELETE"])
@admin_required
def delete_shift(shift_id: int):
    shift = _require_model(Shift, shift_id)
    if shift.employee_assignments:
        return jsonify({"error": "该班次已绑定员工，无法删除"}), 400
    if shift.daily_records:
        return jsonify({"error": "该班次已有关联考勤记录，无法删除"}), 400

    db.session.delete(shift)
    db.session.commit()
    return jsonify({"status": "ok"})


@admin_bp.route("/employees", methods=["GET"])
@admin_required
def employees_list():
    rows = Employee.query.order_by(Employee.emp_no.asc()).all()
    return jsonify([_serialize_employee(e) for e in rows])


@admin_bp.route("/departments", methods=["GET"])
@admin_required
def departments_list():
    rows = Department.query.order_by(Department.dept_name.asc()).all()
    return jsonify(
        [
            {
                "id": d.id,
                "dept_no": d.dept_no,
                "dept_name": d.dept_name,
                "parent_id": d.parent_id,
                "parent_name": d.parent.dept_name if d.parent else "",
                "is_locked": bool(d.is_locked),
            }
            for d in rows
        ]
    )


@admin_bp.route("/manager-overtime/records", methods=["GET"])
@admin_required
def manager_overtime_records():
    year = request.args.get("year", type=int) or datetime.now().year
    return jsonify(_manager_month_rows(_manager_overtime_values(year), "剩余调休天数"))


@admin_bp.route("/manager-overtime/records", methods=["PUT"])
@admin_required
def update_manager_overtime_summary():
    payload, status = _save_manager_month_stat("overtime")
    return jsonify(payload), status


@admin_bp.route("/manager-overtime/records/<int:record_id>", methods=["PUT"])
@admin_required
def update_manager_overtime_record(record_id: int):
    row = _require_model(OvertimeRecord, record_id)
    if not row.employee or not row.employee.is_manager:
        return jsonify({"error": "record is not a manager overtime record"}), 400
    month = row.date.strftime("%Y-%m") if row.date else None
    account_set = _account_set_for_month(month)
    locked_error = _ensure_account_set_unlocked(account_set, "修改管理人员加班记录")
    if locked_error:
        return locked_error
    data = request.json or {}
    row.effective_hours = float(data.get("effective_hours") or 0)
    row.salary_option = (data.get("salary_option") or "").strip()
    row.is_weekend = bool(data.get("is_weekend"))
    row.is_holiday = bool(data.get("is_holiday"))
    row.approval_status = (data.get("approval_status") or "").strip()
    row.reason = (data.get("reason") or "").strip()
    db.session.commit()
    return jsonify({"status": "ok"})


def _manager_export_months(year: int) -> list[tuple[str, str]]:
    months = [(f"{year - 1}-12", "12月")]
    months.extend((f"{year}-{month:02d}", f"{month}月") for month in range(1, 13))
    return months


def _manager_base_month_values() -> dict[str, dict[str, object]]:
    employees = _manager_scope_employees()
    return {
        employee.name: {
            "emp_id": employee.id,
            "dept_name": employee.department.dept_name if employee.department else "",
            "remark": "",
        }
        for employee in employees
    }


def _month_value_keys() -> list[str]:
    return ["prev_dec", *[f"m{month}" for month in range(1, 13)]]


def _annual_leave_value_keys() -> list[str]:
    return [f"m{month}" for month in range(1, 13)]


def _number_or_blank(value: object) -> float | str:
    if value in (None, ""):
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return ""


def _month_for_stat_key(year: int, key: str) -> str:
    if key == "prev_dec":
        return f"{year - 1}-12"
    return f"{year}-{int(key[1:]):02d}"


def _validate_manager_month_stat(stat_type: str, year: int, values: dict[str, float]) -> str | None:
    if stat_type == "annual_leave":
        total = sum(values.values())
        if any(value < 0 for value in values.values()):
            return "年休使用天数不能为负数"
        if total > 12:
            return "年休一年最多 12 天"
        for key, value in values.items():
            if value > 3:
                return "年休每月使用不能超过 3 天"
            month = _month_for_stat_key(year, key)
            account_set = AccountSet.query.filter_by(month=month).first()
            factory_rest_days = (account_set.factory_rest_days if account_set else 0) or 0
            if factory_rest_days + value > 7:
                return f"{month} 厂休+年休不能超过 7 天"

    if stat_type == "overtime":
        balance = 0.0
        for key in _month_value_keys():
            value = values.get(key, 0.0)
            if value < 0 and balance <= 0:
                return "剩余加班天数为 0 时不能使用调休"
            if balance + value < 0:
                return "使用调休天数不能超过当前剩余加班天数"
            balance += value
    return None


def _stat_value_keys(stat_type: str) -> list[str]:
    return _annual_leave_value_keys() if stat_type == "annual_leave" else _month_value_keys()


def _stat_headers(stat_type: str) -> list[str]:
    if stat_type == "annual_leave":
        return ["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"]
    return ["部门", "姓名", "前年累积天数", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余调休天数", "备注"]


def _stat_col_keys(stat_type: str) -> list[tuple[str, str]]:
    if stat_type == "annual_leave":
        return [(f"m{month}", f"{month}月") for month in range(1, 13)]
    return [("prev_dec", "前年累积天数"), *[(f"m{month}", f"{month}月") for month in range(1, 13)]]


def _apply_saved_manager_stats(values_by_name: dict[str, dict[str, object]], year: int, stat_type: str) -> None:
    rows = (
        ManagerMonthStat.query.join(Employee, ManagerMonthStat.emp_id == Employee.id)
        .filter(
            ManagerMonthStat.year == year,
            ManagerMonthStat.stat_type == stat_type,
            Employee.is_manager.is_(True),
        )
        .all()
    )
    for row in rows:
        if not row.employee or row.employee.name not in values_by_name:
            continue
        values = values_by_name[row.employee.name]
        for key in _stat_value_keys(stat_type):
            values[key] = getattr(row, key)
        if stat_type == "annual_leave":
            values["prev_dec"] = ""
        values["remaining"] = row.remaining
        values["remark"] = row.remark or ""


def _upsert_manager_month_stat(stat_type: str, emp_id: int, year: int, values: dict[str, float], remark: str) -> tuple[dict[str, str], int]:
    employee = _require_model(Employee, emp_id)
    if not employee.is_manager:
        return {"error": "employee is not manager"}, 400
    row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
    error = _validate_manager_month_stat(stat_type, year, values)
    if error:
        return {"error": error}, 400
    if not row:
        row = ManagerMonthStat(emp_id=employee.id, year=year, stat_type=stat_type)
        db.session.add(row)
    if stat_type == "annual_leave":
        row.prev_dec = 0
    for key, value in values.items():
        setattr(row, key, value)
    row.remaining = round(12 - sum(values.values()), 2) if stat_type == "annual_leave" else round(sum(values.values()), 2)
    row.remark = (remark or "").strip()
    db.session.commit()
    return {"status": "ok"}, 200


def _save_manager_month_stat(stat_type: str) -> tuple[dict[str, str], int]:
    data = request.json or {}
    emp_id = int(data.get("emp_id") or 0)
    year = int(data.get("year") or datetime.now().year)
    keys = _stat_value_keys(stat_type)
    submitted_values = {key: float(_number_or_blank(data.get(key)) or 0) for key in keys}
    employee = _require_model(Employee, emp_id)
    if not employee.is_manager:
        return {"error": "employee is not manager"}, 400

    row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
    values = {
        key: float(getattr(row, key) or 0) if row else 0.0
        for key in keys
    }
    skipped_locked_months: list[str] = []
    for key in keys:
        month = _stat_key_month(year, key)
        if _stat_key_lock_state(year, key) == "locked":
            if month:
                skipped_locked_months.append(month)
            continue
        values[key] = submitted_values[key]

    payload, status = _upsert_manager_month_stat(stat_type, emp_id, year, values, data.get("remark") or "")
    if status != 200:
        return payload, status
    if skipped_locked_months:
        payload["skipped_locked_months"] = sorted(set(skipped_locked_months))
        payload["warning"] = f"已跳过锁定月份：{'、'.join(payload['skipped_locked_months'])}"
    return payload, status

def _validate_month(value: str | None) -> str | None:
    text = (value or "").strip()
    try:
        datetime.strptime(text, "%Y-%m")
    except ValueError:
        return None
    return text


def _manager_attendance_options(month: str) -> ManagerAttendanceOptions:
    account = AccountSet.query.filter_by(month=month).first()
    return ManagerAttendanceOptions(
        month=month,
        factory_rest_days=(account.factory_rest_days if account else 0) or 0,
        monthly_benefit_days=(account.monthly_benefit_days if account else 0) or 0,
    )


def _manager_attendance_override_payload(row: ManagerAttendanceOverride | None) -> dict[str, object]:
    payload = {field: getattr(row, field) if row else None for field in _MANAGER_ATTENDANCE_OVERRIDE_FIELDS}
    payload["remark"] = row.remark if row else ""
    payload["updated_at"] = row.updated_at.isoformat() if row and row.updated_at else None
    payload["updated_by_name"] = _user_display_name(row.updated_by_user) if row else ""
    return payload


def _manager_attendance_row(emp_id: int, month: str, include_overrides: bool) -> dict[str, object] | None:
    options = _manager_attendance_options(month)
    rows = build_manager_rows(options, [emp_id], include_overrides=include_overrides)
    return rows[0] if rows else None


def _manager_attendance_response(emp_id: int, month: str) -> tuple[dict[str, object], int]:
    employee = db.session.get(Employee, emp_id)
    if not employee or not employee.is_manager:
        return {"error": "employee is not manager"}, 400
    automatic = _manager_attendance_row(emp_id, month, include_overrides=False)
    applied = _manager_attendance_row(emp_id, month, include_overrides=True)
    override = ManagerAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    return {
        "employee": _serialize_employee(employee),
        "month": month,
        "automatic": automatic,
        "override": _manager_attendance_override_payload(override),
        "applied": applied,
        "history": _history_rows_for_month("manager", month),
    }, 200


def _nullable_float(data: dict[str, object], key: str) -> tuple[float | None, str | None]:
    value = data.get(key)
    if value in (None, ""):
        return None, None
    try:
        parsed = round(float(value), 2)
    except (TypeError, ValueError):
        return None, f"{key} 必须是数字"
    if parsed < 0:
        return None, f"{key} 不能为负数"
    return parsed, None


def _nullable_int(data: dict[str, object], key: str) -> tuple[int | None, str | None]:
    value = data.get(key)
    if value in (None, ""):
        return None, None
    try:
        parsed = int(round(float(value)))
    except (TypeError, ValueError):
        return None, f"{key} 必须是整数"
    if parsed < 0:
        return None, f"{key} 不能为负数"
    return parsed, None


def _override_workbook_response(wb: openpyxl.Workbook, filename: str):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _employee_override_export_headers() -> list[str]:
    return [
        "月份",
        "工号",
        "姓名",
        "系统考勤天数",
        "系统工时",
        "系统半勤天数",
        "系统迟到早退",
        "考勤天数",
        "工时",
        "半勤天数",
        "迟到早退",
        "备注",
    ]


def _manager_override_export_headers() -> list[str]:
    return [
        "月份",
        "工号",
        "姓名",
        "系统出勤天数",
        "系统工伤",
        "系统出差",
        "系统婚假",
        "系统丧假",
        "系统迟到早退",
        "出勤天数",
        "工伤",
        "出差",
        "婚假",
        "丧假",
        "迟到早退",
        "备注",
    ]


def _build_employee_override_export_workbook(month: str, include_real_rows: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工考勤修正"
    ws.append(_employee_override_export_headers())
    if include_real_rows:
        employees = (
            Employee.query.filter_by(is_manager=False)
            .order_by(Employee.dept_id.asc(), Employee.emp_no.asc(), Employee.name.asc())
            .all()
        )
        for employee in employees:
            automatic = _employee_automatic_row(employee.id, month) or {}
            override = EmployeeAttendanceOverride.query.filter_by(emp_id=employee.id, month=month).first()
            ws.append(
                [
                    month,
                    employee.emp_no,
                    employee.name,
                    automatic.get("attendance_days"),
                    automatic.get("work_hours"),
                    automatic.get("half_days"),
                    automatic.get("late_early_minutes"),
                    override.attendance_days if override else "",
                    override.work_hours if override else "",
                    override.half_days if override else "",
                    override.late_early_minutes if override else "",
                    override.remark if override and override.remark else "",
                ]
            )
    else:
        ws.append(["2026-05", "1001001", "张三", 20, 160, 0, 5, "", "", "", "", "留空表示不覆盖"])
    return wb


def _build_manager_override_export_workbook(month: str, include_real_rows: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "管理人员考勤修正"
    ws.append(_manager_override_export_headers())
    if include_real_rows:
        for employee in _manager_scope_employees():
            automatic = _manager_attendance_row(employee.id, month, include_overrides=False) or {}
            override = ManagerAttendanceOverride.query.filter_by(emp_id=employee.id, month=month).first()
            ws.append(
                [
                    month,
                    employee.emp_no,
                    employee.name,
                    automatic.get("attendance_days"),
                    automatic.get("injury_days"),
                    automatic.get("business_trip_days"),
                    automatic.get("marriage_days"),
                    automatic.get("funeral_days"),
                    automatic.get("late_early_minutes"),
                    override.attendance_days if override else "",
                    override.injury_days if override else "",
                    override.business_trip_days if override else "",
                    override.marriage_days if override else "",
                    override.funeral_days if override else "",
                    override.late_early_minutes if override else "",
                    override.remark if override and override.remark else "",
                ]
            )
    else:
        ws.append(["2026-05", "2001001", "李经理", 22, 0, 2, 0, 0, 0, "", "", "", "", "", "", "留空表示不覆盖"])
    return wb


def _import_summary(success_count: int, skipped_count: int, failed_count: int, changed_count: int, errors: list[str]) -> dict[str, object]:
    return {
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        "changed_count": changed_count,
        "errors": errors,
    }


def _apply_employee_override_updates(
    row: EmployeeAttendanceOverride | None,
    emp_id: int,
    month: str,
    updates: dict[str, object],
    action_type: str,
    source_file_name: str | None = None,
) -> bool:
    before_values = _override_state_from_row(row, _EMPLOYEE_OVERRIDE_FIELDS)
    after_values = dict(before_values)
    after_values.update(updates)
    if not _has_override_state_changes(before_values, after_values):
        return False
    if not row:
        row = EmployeeAttendanceOverride(emp_id=emp_id, month=month)
        db.session.add(row)
    for field in _EMPLOYEE_OVERRIDE_FIELDS:
        setattr(row, field, after_values.get(field))
    row.remark = str(after_values.get("remark") or "")
    row.updated_by = g.current_user.id
    _record_override_history("employee", emp_id, month, action_type, before_values, after_values, source_file_name)
    return True


def _apply_manager_override_updates(
    row: ManagerAttendanceOverride | None,
    emp_id: int,
    month: str,
    updates: dict[str, object],
    action_type: str,
    source_file_name: str | None = None,
) -> bool:
    before_values = _override_state_from_row(row, _MANAGER_ATTENDANCE_OVERRIDE_FIELDS)
    after_values = dict(before_values)
    after_values.update(updates)
    if not _has_override_state_changes(before_values, after_values):
        return False
    if not row:
        row = ManagerAttendanceOverride(emp_id=emp_id, month=month)
        db.session.add(row)
    for field in _MANAGER_ATTENDANCE_OVERRIDE_FIELDS:
        setattr(row, field, after_values.get(field))
    row.remark = str(after_values.get("remark") or "")
    row.updated_by = g.current_user.id
    _record_override_history("manager", emp_id, month, action_type, before_values, after_values, source_file_name)
    return True


def _stat_key_for_month(month: str) -> tuple[int, str]:
    year, month_no = [int(part) for part in month.split("-", 1)]
    return year, f"m{month_no}"


def _sync_manager_stats_from_manager_rows(month: str, manager_rows: list[dict[str, object]]) -> dict[str, object]:
    """Stats are now written directly inside build_manager_rows.
    This function provides a compatibility layer — it validates the results
    but no longer writes to the stat tables to avoid double-writing.
    """
    year, key = _stat_key_for_month(month)
    employees_by_name = {employee.name: employee for employee in _manager_scope_employees()}
    errors: list[str] = []

    for item in manager_rows:
        name = str(item.get("name") or "").strip()
        employee = employees_by_name.get(name)
        if not employee:
            errors.append(f"{name or '空姓名'}：未找到管理人员")
            continue

        for stat_type, source_key, label in (
            ("overtime", "overtime_change", "加班变化"),
            ("annual_leave", "benefit_days", "福利天数"),
        ):
            stat_row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
            values = {
                stat_key: float(getattr(stat_row, stat_key) or 0) if stat_row else 0.0
                for stat_key in _stat_value_keys(stat_type)
            }
            # Values already written by build_manager_rows; validate only
            error = _validate_manager_month_stat(stat_type, year, values)
            if error:
                errors.append(f"{employee.name} {label}：{error}")

    return {
        "month": month,
        "overtime_synced": 0,
        "annual_leave_synced": 0,
        "error_count": len(errors),
        "errors": errors,
    }


def _download_manager_stat_template(stat_type: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "年休统计表" if stat_type == "annual_leave" else "加班统计表"
    ws.append(_stat_headers(stat_type))
    employees = _manager_scope_employees()
    for employee in employees:
        values = [employee.department.dept_name if employee.department else "", employee.name]
        values.extend("" for _key, _label in _stat_col_keys(stat_type))
        values.extend([12 if stat_type == "annual_leave" else 0, ""])
        ws.append(values)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "管理人员年休导入示例.xlsx" if stat_type == "annual_leave" else "管理人员加班导入示例.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _header_map(ws) -> dict[str, int]:
    headers = {}
    for cell in ws[1]:
        text = str(cell.value or "").strip()
        if text:
            headers[text] = cell.column
    return headers


def _import_manager_stat_file(stat_type: str, year: int):
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "请选择要导入的Excel文件"}), 400

    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = _header_map(ws)
    if "姓名" not in headers:
        return jsonify({"error": "导入文件缺少姓名列"}), 400

    imported = 0
    errors: list[str] = []
    skipped_locked_months: set[str] = set()
    employees_by_name = {employee.name: employee for employee in _manager_scope_employees()}
    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row_idx, headers["姓名"]).value or "").strip()
        if not name:
            continue
        employee = employees_by_name.get(name)
        if not employee:
            errors.append(f"第{row_idx}行：未找到管理人员 {name}")
            continue

        existing_row = ManagerMonthStat.query.filter_by(emp_id=employee.id, year=year, stat_type=stat_type).first()
        values = {
            key: float(getattr(existing_row, key) or 0) if existing_row else 0.0
            for key in _stat_value_keys(stat_type)
        }
        for key, label in _stat_col_keys(stat_type):
            col_idx = headers.get(label)
            month = _stat_key_month(year, key)
            if _stat_key_lock_state(year, key) == "locked":
                if month:
                    skipped_locked_months.add(month)
                continue
            values[key] = float(_number_or_blank(ws.cell(row_idx, col_idx).value if col_idx else None) or 0)
        remark_col = headers.get("备注")
        remark = str(ws.cell(row_idx, remark_col).value or "").strip() if remark_col else ""
        payload, status = _upsert_manager_month_stat(stat_type, employee.id, year, values, remark)
        if status != 200:
            errors.append(f"第{row_idx}行：{payload.get('error', '保存失败')}")
            continue
        imported += 1

    response = {
        "status": "ok",
        "imported": imported,
        "errors": errors,
        "error_count": len(errors),
        "skipped_locked_months": sorted(skipped_locked_months),
    }
    if skipped_locked_months:
        response["warning"] = f"已跳过锁定月份：{'、'.join(response['skipped_locked_months'])}"
    return jsonify(response)


def _manager_overtime_values(year: int) -> dict[str, dict[str, object]]:
    values_by_name = _manager_base_month_values()
    for values in values_by_name.values():
        values["remaining"] = 0
    _apply_saved_manager_stats(values_by_name, year, "overtime")
    return values_by_name


def _manager_annual_leave_values(year: int) -> dict[str, dict[str, object]]:
    values_by_name = _manager_base_month_values()
    for values in values_by_name.values():
        values["remaining"] = 12
    _apply_saved_manager_stats(values_by_name, year, "annual_leave")
    return values_by_name


def _manager_month_rows(values_by_name: dict[str, dict[str, object]], remaining_label: str, keys: list[str] | None = None) -> list[dict[str, object]]:
    keys = keys or _month_value_keys()
    rows = []
    for name, values in values_by_name.items():
        row = {
            "emp_id": values.get("emp_id"),
            "dept_name": values.get("dept_name", ""),
            "name": name,
            "remaining_label": remaining_label,
            "remaining": values.get("remaining", ""),
            "remark": values.get("remark", ""),
        }
        for key in keys:
            row[key] = values.get(key, "")
        rows.append(row)
    return rows


def _fill_named_month_template(ws, values_by_name: dict[str, dict[str, object]], total_key: str, keys: list[str] | None = None) -> None:
    keys = keys or _month_value_keys()
    total_col = 3 + len(keys)
    remark_col = total_col + 1
    filled: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row_idx, 2).value or "").strip()
        for col_idx in range(3, remark_col + 1):
            ws.cell(row_idx, col_idx).value = ""
        if not name or name not in values_by_name:
            continue
        values = values_by_name[name]
        for col_idx, key in enumerate(keys, start=3):
            ws.cell(row_idx, col_idx).value = values.get(key, "")
        ws.cell(row_idx, total_col).value = values.get(total_key, "")
        ws.cell(row_idx, remark_col).value = values.get("remark", "")
        filled.add(name)

    for name, values in values_by_name.items():
        if name in filled:
            continue
        ws.append(
            [
                values.get("dept_name", ""),
                name,
                *[values.get(key, "") for key in keys],
                values.get(total_key, ""),
                values.get("remark", ""),
            ]
        )


def _export_manager_overtime_workbook(year: int):
    values_by_name = _manager_overtime_values(year)
    month_keys = _manager_export_months(year)
    template_path = "/home/lewis/文档/考勤/加班单查询表.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["部门", "姓名", *[label for _month, label in month_keys], "剩余调休天数", "备注"])
    ws.cell(1, 3).value = "前年累积天数"
    _fill_named_month_template(ws, values_by_name, "remaining")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员加班信息_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _export_manager_annual_leave_workbook(year: int):
    values_by_name = _manager_annual_leave_values(year)
    template_path = "/home/lewis/文档/考勤/加班单查询表.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        ws.delete_cols(3)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["部门", "姓名", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "剩余年休天数", "备注"])
    ws.title = "年休统计表"
    ws.cell(1, 15).value = "剩余年休天数"
    ws.cell(1, 16).value = "备注"
    _fill_named_month_template(ws, values_by_name, "remaining", _annual_leave_value_keys())

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"管理人员年休信息_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.route("/manager-annual-leave/records", methods=["GET"])
@admin_required
def manager_annual_leave_records():
    year = request.args.get("year", type=int) or datetime.now().year
    return jsonify(_manager_month_rows(_manager_annual_leave_values(year), "剩余年休天数", _annual_leave_value_keys()))


@admin_bp.route("/manager-annual-leave/records", methods=["PUT"])
@admin_required
def update_manager_annual_leave_record():
    data = request.json or {}
    if any(key in data for key in _annual_leave_value_keys()) or "remaining" in data:
        payload, status = _save_manager_month_stat("annual_leave")
        return jsonify(payload), status

    emp_id = int(data.get("emp_id") or 0)
    year = int(data.get("year") or datetime.now().year)
    locked_error = _ensure_year_months_unlocked(year, "修改管理人员年休统计")
    if locked_error:
        return locked_error
    employee = _require_model(Employee, emp_id)
    if not employee.is_manager:
        return jsonify({"error": "employee is not manager"}), 400
    row = AnnualLeave.query.filter_by(emp_id=employee.id, year=year).first()
    if not row:
        row = AnnualLeave(emp_id=employee.id, year=year)
        db.session.add(row)
    row.total_days = float(data.get("total_days") or 0)
    row.used_days = float(data.get("used_days") or 0)
    row.remaining_days = float(data.get("remaining_days") or 0)
    db.session.commit()
    return jsonify({"status": "ok", "id": row.id})


@admin_bp.route("/departments", methods=["POST"])
@admin_required
def create_department():
    data = request.json or {}
    dept_no = (data.get("dept_no") or "").strip()
    dept_name = (data.get("dept_name") or "").strip()
    parent_id = _parse_parent_id(data.get("parent_id"))
    is_locked = bool(data.get("is_locked"))
    if not dept_no or not dept_name:
        return jsonify({"error": "dept_no and dept_name are required"}), 400
    if parent_id == -1:
        return jsonify({"error": "parent_id must be integer"}), 400
    if Department.query.filter_by(dept_no=dept_no).first():
        return jsonify({"error": "dept_no already exists"}), 400
    parent, err = _validate_parent_department(parent_id)
    if err:
        return jsonify({"error": err}), 400

    department = Department(
        dept_no=dept_no,
        dept_name=dept_name,
        parent_id=parent.id if parent else None,
        is_locked=is_locked,
    )
    db.session.add(department)
    db.session.commit()
    return jsonify({"status": "ok", "id": department.id})


@admin_bp.route("/departments/<int:dept_id>", methods=["PUT"])
@admin_required
def update_department(dept_id: int):
    data = request.json or {}
    dept_no = (data.get("dept_no") or "").strip()
    dept_name = (data.get("dept_name") or "").strip()
    parent_id = _parse_parent_id(data.get("parent_id"))
    is_locked = bool(data.get("is_locked"))
    if not dept_no or not dept_name:
        return jsonify({"error": "dept_no and dept_name are required"}), 400
    if parent_id == -1:
        return jsonify({"error": "parent_id must be integer"}), 400

    department = _require_model(Department, dept_id)
    duplicate = Department.query.filter(Department.dept_no == dept_no, Department.id != dept_id).first()
    if duplicate:
        return jsonify({"error": "dept_no already exists"}), 400
    parent, err = _validate_parent_department(parent_id, dept_id)
    if err:
        return jsonify({"error": err}), 400

    department.dept_no = dept_no
    department.dept_name = dept_name
    department.parent_id = parent.id if parent else None
    department.is_locked = is_locked
    db.session.commit()
    return jsonify({"status": "ok"})


@admin_bp.route("/departments/<int:dept_id>", methods=["DELETE"])
@admin_required
def delete_department(dept_id: int):
    department = _require_model(Department, dept_id)
    if department.children:
        return jsonify({"error": "该部门存在下级部门，无法删除"}), 400
    if department.employees:
        return jsonify({"error": "该部门已绑定员工，无法删除"}), 400
    if department.user_assignments:
        return jsonify({"error": "该部门已绑定账号权限，无法删除"}), 400

    db.session.delete(department)
    db.session.commit()
    return jsonify({"status": "ok"})


@admin_bp.route("/departments/batch", methods=["POST"])
@admin_required
def batch_operate_departments():
    data = request.json or {}
    ids = data.get("ids") or []
    action = (data.get("action") or "").strip()

    if action not in {"delete", "set_parent", "lock", "unlock"}:
        return jsonify({"error": "unsupported action"}), 400
    if not ids:
        return jsonify({"error": "ids are required"}), 400

    departments = Department.query.filter(Department.id.in_(ids)).all()
    if not departments:
        return jsonify({"error": "departments not found"}), 404

    if action == "delete":
        blocked = []
        for department in departments:
            if department.children:
                blocked.append(f"{department.dept_name}（存在下级部门）")
            elif department.employees:
                blocked.append(f"{department.dept_name}（已绑定员工）")
            elif department.user_assignments:
                blocked.append(f"{department.dept_name}（已绑定账号权限）")
        if blocked:
            return jsonify({"error": f"以下部门不可删除：{', '.join(blocked)}"}), 400

        for department in departments:
            db.session.delete(department)
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(departments)})

    if action in {"lock", "unlock"}:
        locked = action == "lock"
        for department in departments:
            department.is_locked = locked
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(departments)})

    parent_id = _parse_parent_id(data.get("parent_id"))
    if parent_id == -1:
        return jsonify({"error": "parent_id must be integer"}), 400
    selected_ids = {d.id for d in departments}
    if parent_id in selected_ids:
        return jsonify({"error": "上级部门不能选择已选部门"}), 400
    parent, err = _validate_parent_department(parent_id)
    if err:
        return jsonify({"error": err}), 400

    for department in departments:
        if parent:
            cursor = parent
            while cursor:
                if cursor.id == department.id:
                    return jsonify({"error": f"{department.dept_name} 的上级部门设置会形成循环层级"}), 400
                cursor = cursor.parent
        department.parent_id = parent.id if parent else None
    db.session.commit()
    return jsonify({"status": "ok", "action": action, "affected": len(departments)})


@admin_bp.route("/departments/delete-unbound", methods=["POST"])
@admin_required
def delete_unbound_departments():
    all_departments = Department.query.all()
    deleted = 0
    skipped_locked = 0
    skipped_employee_bound = 0
    skipped_account_bound = 0

    for department in all_departments:
        if department.children:
            continue
        if department.is_locked:
            skipped_locked += 1
            continue
        if department.employees:
            skipped_employee_bound += 1
            continue
        if department.user_assignments:
            skipped_account_bound += 1
            continue
        db.session.delete(department)
        deleted += 1

    db.session.commit()
    return jsonify(
        {
            "status": "ok",
            "deleted": deleted,
            "skipped_locked": skipped_locked,
            "skipped_employee_bound": skipped_employee_bound,
            "skipped_account_bound": skipped_account_bound,
        }
    )


def _build_departments_workbook(
    rows: list[tuple[str, str, str, str]],
    include_identity_metadata: bool = False,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "部门导入模板"
    ws.append(["部门编号", "部门名称", "上级部门编号", _DEPARTMENT_ORIGINAL_ID_HEADER])
    ws.column_dimensions["D"].hidden = True

    for row in rows:
        ws.append(list(row[:4]))

    if include_identity_metadata:
        metadata_ws = wb.create_sheet(_DEPARTMENT_METADATA_SHEET)
        metadata_ws.sheet_state = "hidden"
        metadata_ws.append(
            [
                _DEPARTMENT_ORIGINAL_ID_HEADER,
                _DEPARTMENT_ORIGINAL_DEPT_NO_HEADER,
                _DEPARTMENT_ORIGINAL_DEPT_NAME_HEADER,
            ]
        )
        for row in rows:
            metadata_ws.append([row[3], row[0], row[1]])
    return wb


@admin_bp.route("/employees", methods=["POST"])
@admin_required
def create_employee():
    data = request.json or {}
    emp_no = (data.get("emp_no") or "").strip()
    name = (data.get("name") or "").strip()
    dept_name = (data.get("dept_name") or "").strip()
    shift_no = (data.get("shift_no") or "").strip()
    is_manager = bool(data.get("is_manager"))
    is_nursing = bool(data.get("is_nursing"))
    employee_stats_attendance_source = _parse_attendance_source(
        data.get("employee_stats_attendance_source"), ATTENDANCE_SOURCE_EMPLOYEE
    )
    manager_stats_attendance_source = _parse_attendance_source(
        data.get("manager_stats_attendance_source"), ATTENDANCE_SOURCE_MANAGER
    )

    if not emp_no or not name:
        return jsonify({"error": "emp_no and name are required"}), 400
    if Employee.query.filter_by(emp_no=emp_no).first():
        return jsonify({"error": "emp_no already exists"}), 400

    department = _resolve_department(dept_name) if dept_name else None
    employee = Employee(
        emp_no=emp_no,
        name=name,
        dept_id=department.id if department else None,
        is_manager=is_manager,
        is_nursing=is_nursing,
        employee_stats_attendance_source=employee_stats_attendance_source,
        manager_stats_attendance_source=manager_stats_attendance_source,
    )
    db.session.add(employee)
    db.session.flush()
    _assign_employee_shift(employee, _resolve_shift(shift_no))
    db.session.commit()
    return jsonify({"status": "ok", "employee": _serialize_employee(employee)})


@admin_bp.route("/employees/<int:employee_id>", methods=["PUT"])
@admin_required
def update_employee(employee_id: int):
    data = request.json or {}
    emp_no = (data.get("emp_no") or "").strip()
    name = (data.get("name") or "").strip()
    dept_name = (data.get("dept_name") or "").strip()
    shift_no = (data.get("shift_no") or "").strip()
    is_manager = bool(data.get("is_manager"))
    is_nursing = data.get("is_nursing")
    employee_stats_attendance_source = data.get("employee_stats_attendance_source")
    manager_stats_attendance_source = data.get("manager_stats_attendance_source")
    if is_nursing is not None:
        is_nursing = bool(is_nursing)

    employee = _require_model(Employee, employee_id)

    if not emp_no or not name:
        return jsonify({"error": "emp_no and name are required"}), 400

    duplicate = Employee.query.filter(Employee.emp_no == emp_no, Employee.id != employee_id).first()
    if duplicate:
        return jsonify({"error": "emp_no already exists"}), 400

    employee.emp_no = emp_no
    employee.name = name
    employee.is_manager = is_manager
    if is_nursing is not None:
        employee.is_nursing = is_nursing
    employee.employee_stats_attendance_source = _parse_attendance_source(
        employee_stats_attendance_source, employee.employee_stats_attendance_source or ATTENDANCE_SOURCE_EMPLOYEE
    )
    employee.manager_stats_attendance_source = _parse_attendance_source(
        manager_stats_attendance_source, employee.manager_stats_attendance_source or ATTENDANCE_SOURCE_MANAGER
    )
    if dept_name:
        department = _resolve_department(dept_name)
        employee.dept_id = department.id if department else None
    else:
        employee.dept_id = None
    _assign_employee_shift(employee, _resolve_shift(shift_no))

    db.session.commit()
    return jsonify({"status": "ok", "employee": _serialize_employee(employee)})


@admin_bp.route("/employees/<int:employee_id>", methods=["DELETE"])
@admin_required
def delete_employee(employee_id: int):
    employee = _require_model(Employee, employee_id)
    db.session.delete(employee)
    db.session.commit()
    return jsonify({"status": "ok"})


@admin_bp.route("/employees/batch", methods=["POST"])
@admin_required
def batch_operate_employees():
    data = request.json or {}
    ids = data.get("ids") or []
    action = (data.get("action") or "").strip()

    if not ids:
        return jsonify({"error": "ids are required"}), 400

    employees = Employee.query.filter(Employee.id.in_(ids)).all()
    if not employees:
        return jsonify({"error": "employees not found"}), 404

    if action == "delete":
        for employee in employees:
            db.session.delete(employee)
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_department":
        dept_name = (data.get("dept_name") or "").strip()
        department = _resolve_department(dept_name) if dept_name else None
        for employee in employees:
            employee.dept_id = department.id if department else None
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_shift":
        shift_no = (data.get("shift_no") or "").strip()
        shift = _resolve_shift(shift_no) if shift_no else None
        for employee in employees:
            _assign_employee_shift(employee, shift)
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_manager":
        is_manager = bool(data.get("is_manager"))
        for employee in employees:
            employee.is_manager = is_manager
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_nursing":
        is_nursing = bool(data.get("is_nursing"))
        for employee in employees:
            employee.is_nursing = is_nursing
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_employee_stats_attendance_source":
        source = _parse_attendance_source(data.get("employee_stats_attendance_source"), ATTENDANCE_SOURCE_EMPLOYEE)
        for employee in employees:
            employee.employee_stats_attendance_source = source
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_manager_stats_attendance_source":
        source = _parse_attendance_source(data.get("manager_stats_attendance_source"), ATTENDANCE_SOURCE_MANAGER)
        for employee in employees:
            employee.manager_stats_attendance_source = source
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": len(employees)})

    if action == "set_name":
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name is required"}), 400
        if len(employees) != 1:
            return jsonify({"error": "set_name only supports single selected employee"}), 400
        employees[0].name = name
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": 1})

    if action == "set_emp_no":
        emp_no = (data.get("emp_no") or "").strip()
        if not emp_no:
            return jsonify({"error": "emp_no is required"}), 400
        if len(employees) != 1:
            return jsonify({"error": "set_emp_no only supports single selected employee"}), 400
        duplicate = Employee.query.filter(Employee.emp_no == emp_no, Employee.id != employees[0].id).first()
        if duplicate:
            return jsonify({"error": "emp_no already exists"}), 400
        employees[0].emp_no = emp_no
        db.session.commit()
        return jsonify({"status": "ok", "action": action, "affected": 1})

    return jsonify({"error": "unsupported action"}), 400


@admin_bp.route("/daily-records/<int:record_id>/annotate", methods=["POST"])
@admin_required
def annotate_record(record_id: int):
    data = request.json or {}
    reason = (data.get("exception_reason") or "").strip()
    record = _require_model(DailyRecord, record_id)
    record.exception_reason = reason
    db.session.commit()
    return jsonify({"status": "ok"})


def _employee_override_values(override: EmployeeAttendanceOverride | None) -> dict[str, float | int | None]:
    return {field: getattr(override, field) if override else None for field in _EMPLOYEE_OVERRIDE_FIELDS}


def _employee_automatic_row(emp_id: int, month: str) -> dict[str, object] | None:
    from routes.employee import _build_final_rows

    rows = _build_final_rows(month, [emp_id])
    if not rows:
        return None
    row = rows[0]
    return {
        "attendance_days": row[3],
        "work_hours": row[16],
        "half_days": row[17],
        "late_early_minutes": _employee_late_early_minutes(emp_id, month),
    }


def _employee_late_early_minutes(emp_id: int, month: str) -> int:
    from routes.employee import _month_date_range

    date_range = _month_date_range(month)
    if not date_range:
        return 0
    start_date, end_date = date_range
    records = (
        DailyRecord.query.filter_by(emp_id=emp_id)
        .filter(DailyRecord.record_date >= start_date, DailyRecord.record_date < end_date)
        .all()
    )
    return sum((r.late_minutes or 0) + (r.early_leave_minutes or 0) for r in records)


def _employee_override_payload(row: EmployeeAttendanceOverride | None) -> dict[str, object]:
    payload = {field: getattr(row, field) if row else None for field in _EMPLOYEE_OVERRIDE_FIELDS}
    payload["remark"] = row.remark if row else ""
    payload["updated_at"] = row.updated_at.isoformat() if row and row.updated_at else None
    payload["updated_by_name"] = _user_display_name(row.updated_by_user) if row else ""
    return payload


def _employee_override_response(emp_id: int, month: str) -> tuple[dict[str, object], int]:
    employee = db.session.get(Employee, emp_id)
    if not employee or employee.is_manager:
        return {"error": "employee is a manager, not a regular employee"}, 400
    automatic = _employee_automatic_row(emp_id, month)
    override = EmployeeAttendanceOverride.query.filter_by(emp_id=emp_id, month=month).first()
    override_data = _employee_override_values(override)
    applied: dict[str, object] = {}
    if automatic:
        for field in _EMPLOYEE_OVERRIDE_FIELDS:
            applied[field] = override_data[field] if override_data[field] is not None else automatic.get(field)
    return {
        "employee": _serialize_employee(employee),
        "month": month,
        "automatic": automatic,
        "override": _employee_override_payload(override),
        "applied": applied,
        "history": _history_rows_for_month("employee", month),
    }, 200


@admin_bp.route("/")
@admin_required
def admin_root():
    return redirect("/admin/dashboard")


from .admin_accounts import register_admin_account_routes
from .admin_attendance_overrides import register_admin_attendance_override_routes
from .admin_imports import register_admin_import_routes


register_admin_account_routes(admin_bp)
register_admin_attendance_override_routes(admin_bp)
register_admin_import_routes(admin_bp)
