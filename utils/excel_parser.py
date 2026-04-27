from __future__ import annotations

import os
import subprocess
import tempfile
from typing import Any

import openpyxl
import xlrd


class ExcelParser:
    @staticmethod
    def read_rows(file_path: str) -> list[list[Any]]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsx":
            return ExcelParser._read_xlsx(file_path)
        if ext == ".xls":
            return ExcelParser._read_xls(file_path)
        raise ValueError(f"Unsupported file type: {ext}")

    @staticmethod
    def _read_xlsx(file_path: str) -> list[list[Any]]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows

    @staticmethod
    def _read_xls(file_path: str) -> list[list[Any]]:
        # Prefer converting to xlsx first because some legacy xls files have
        # encoding/structure issues in xlrd, while libreoffice conversion is stable.
        converted_rows = ExcelParser._read_xls_via_libreoffice(file_path)
        if converted_rows is not None:
            return converted_rows

        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                row.append(ExcelParser._convert_xls_cell(cell, wb.datemode))
            rows.append(row)
        return rows

    @staticmethod
    def _read_xls_via_libreoffice(file_path: str) -> list[list[Any]] | None:
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                profile_dir = os.path.join(tmpdir, "lo-profile")
                os.makedirs(profile_dir, exist_ok=True)
                env = os.environ.copy()
                env["HOME"] = tmpdir
                env["XDG_CONFIG_HOME"] = tmpdir
                subprocess.run(
                    [
                        "libreoffice",
                        f"-env:UserInstallation=file://{profile_dir}",
                        "--headless",
                        "--convert-to",
                        "xlsx",
                        "--outdir",
                        tmpdir,
                        file_path,
                    ],
                    check=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    env=env,
                )
                basename = os.path.splitext(os.path.basename(file_path))[0]
                converted_path = os.path.join(tmpdir, f"{basename}.xlsx")
                if not os.path.exists(converted_path):
                    return None
                return ExcelParser._read_xlsx(converted_path)
        except Exception:
            return None

    @staticmethod
    def _convert_xls_cell(cell: xlrd.sheet.Cell, datemode: int) -> Any:
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_datetime(cell.value, datemode)
                return dt
            except Exception:
                return cell.value
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            if float(cell.value).is_integer():
                return int(cell.value)
            return float(cell.value)
        if cell.ctype == xlrd.XL_CELL_TEXT and isinstance(cell.value, str):
            return ExcelParser._fix_mojibake(cell.value)
        return cell.value

    @staticmethod
    def _fix_mojibake(text: str) -> str:
        try:
            repaired = text.encode("latin1").decode("gbk")
            if any("\u4e00" <= ch <= "\u9fff" for ch in repaired):
                return repaired
        except Exception:
            pass
        return text
